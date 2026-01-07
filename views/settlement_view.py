#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
结算管理视图
负责处理与出纳的结算上缴信息管理
"""

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from models.settlement import Settlement
from models.charge import Charge
from models.payment import Payment

class SettlementView:
    """结算管理视图类"""
    
    def __init__(self, parent, language_utils=None):
        """
        初始化结算管理视图
        :param parent: 父窗口组件
        :param language_utils: 语言工具类实例
        """
        self.parent = parent
        self.language_utils = language_utils
        self.settlement_list = []
        self.selected_settlement = None
        self.is_initial_load = True  # 添加标志变量，区分初始加载和手动刷新
        self.create_widgets()
        self.load_settlement_list()
    
    def get_text(self, key):
        """
        获取翻译文本
        :param key: 文本键名
        :return: 翻译后的文本
        """
        return self.language_utils.get_text(key) if self.language_utils else key
        
    def update_language(self):
        """
        更新界面语言
        """
        # 更新筛选框文本
        self.filter_labels['settle_month_label']['text'] = self.get_text('settle_month') + ':'
        self.filter_buttons['search_btn']['text'] = self.get_text('search')
        self.filter_buttons['reset_btn']['text'] = self.get_text('reset')
        
        # 更新操作按钮文本
        self.action_buttons['add_settlement_btn']['text'] = self.get_text('add_settlement')
        self.action_buttons['edit_settlement_btn']['text'] = self.get_text('edit_settlement')
        self.action_buttons['delete_settlement_btn']['text'] = self.get_text('delete_settlement')
        self.action_buttons['generate_report_btn']['text'] = self.get_text('generate_settlement_report')
        self.action_buttons['refresh_btn']['text'] = self.get_text('refresh_list')
        
        # 更新列表列标题
        self.settlement_tree.heading("serial", text=self.get_text('serial'))
        self.settlement_tree.heading("settle_month", text=self.get_text('settle_month'))
        self.settlement_tree.heading("settle_date", text=self.get_text('settle_date'))
        self.settlement_tree.heading("total_amount", text=self.get_text('total_amount'))
        self.settlement_tree.heading("cashier", text=self.get_text('cashier'))
        self.settlement_tree.heading("notes", text=self.get_text('notes'))
        
        # 更新表单文本
        self.form_title['text'] = self.get_text('settlement_details')
        self.form_labels['settle_month_label']['text'] = self.get_text('settle_month') + ':'
        self.calculate_btn['text'] = self.get_text('calculate_monthly_received')
        self.form_labels['total_amount_label']['text'] = self.get_text('total_amount') + ':'
        self.form_labels['settle_date_label']['text'] = self.get_text('settle_date') + ':'
        self.form_labels['cashier_label']['text'] = self.get_text('cashier') + ':'
        self.form_labels['notes_label']['text'] = self.get_text('notes') + ':'
        
        # 更新表单按钮文本
        self.form_buttons['save_btn']['text'] = self.get_text('save')
        self.form_buttons['cancel_btn']['text'] = self.get_text('cancel')
        
        # 重新加载列表，确保状态文本正确
        self.load_settlement_list()
    
    def create_widgets(self):
        """
        创建结算管理界面组件
        """
        # 创建主框架
        main_frame = ttk.Frame(self.parent)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 顶部：筛选和操作按钮
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        
        # 筛选条件
        filter_frame = ttk.Frame(top_frame)
        filter_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 存储筛选框标签
        self.filter_labels = {}
        # 存储筛选框按钮
        self.filter_buttons = {}
        # 存储操作按钮
        self.action_buttons = {}
        # 存储表单标签
        self.form_labels = {}
        # 存储表单按钮
        self.form_buttons = {}
        
        self.filter_labels['settle_month_label'] = ttk.Label(filter_frame, text=self.get_text('settle_month') + ':')
        self.filter_labels['settle_month_label'].pack(side=tk.LEFT, padx=5)
        
        self.month_var = tk.StringVar()
        self.month_var.set(datetime.now().strftime("%Y-%m"))
        self.month_entry = ttk.Entry(filter_frame, textvariable=self.month_var, width=10)
        self.month_entry.pack(side=tk.LEFT, padx=5)
        
        self.filter_buttons['search_btn'] = ttk.Button(filter_frame, text=self.get_text('search'), command=self.search_settlements)
        self.filter_buttons['search_btn'].pack(side=tk.LEFT, padx=5)
        
        self.filter_buttons['reset_btn'] = ttk.Button(filter_frame, text=self.get_text('reset'), command=self.reset_filter)
        self.filter_buttons['reset_btn'].pack(side=tk.LEFT, padx=5)
        
        # 操作按钮
        action_frame = ttk.Frame(top_frame)
        action_frame.pack(side=tk.RIGHT)
        
        self.action_buttons['add_settlement_btn'] = ttk.Button(action_frame, text=self.get_text('add_settlement'), command=self.add_settlement)
        self.action_buttons['add_settlement_btn'].pack(side=tk.LEFT, padx=5)
        
        self.action_buttons['edit_settlement_btn'] = ttk.Button(action_frame, text=self.get_text('edit_settlement'), command=self.edit_settlement)
        self.action_buttons['edit_settlement_btn'].pack(side=tk.LEFT, padx=5)
        
        self.action_buttons['delete_settlement_btn'] = ttk.Button(action_frame, text=self.get_text('delete_settlement'), command=self.delete_settlement)
        self.action_buttons['delete_settlement_btn'].pack(side=tk.LEFT, padx=5)
        
        self.action_buttons['generate_report_btn'] = ttk.Button(action_frame, text=self.get_text('generate_settlement_report'), command=self.generate_settlement_report)
        self.action_buttons['generate_report_btn'].pack(side=tk.LEFT, padx=5)
        
        self.action_buttons['refresh_btn'] = ttk.Button(action_frame, text=self.get_text('refresh_list'), command=self.load_settlement_list)
        self.action_buttons['refresh_btn'].pack(side=tk.LEFT, padx=5)
        
        # 中部：结算记录列表
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 列表容器，用于放置Treeview和滚动条
        tree_container = ttk.Frame(list_frame)
        tree_container.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        # 列表控件 - 添加序号列，隐藏ID列
        columns = ("serial", "id", "settle_month", "settle_date", "total_amount", "cashier", "notes")
        self.settlement_tree = ttk.Treeview(tree_container, columns=columns, show="headings")
        
        # 配置Treeview的tag，用于实现隔行背景色
        # 奇数行使用深灰色，偶数行使用白色，与charge_view.py保持一致
        self.settlement_tree.tag_configure('odd', background='#c0c0c0')
        self.settlement_tree.tag_configure('even', background='#ffffff')
        
        # 设置列标题和宽度
        self.settlement_tree.heading("serial", text=self.get_text('serial'))
        self.settlement_tree.column("serial", width=50, anchor="center")
        
        # 隐藏ID列
        self.settlement_tree.heading("id", text="ID")
        self.settlement_tree.column("id", width=0, stretch=False)
        
        self.settlement_tree.heading("settle_month", text=self.get_text('settle_month'))
        self.settlement_tree.column("settle_month", width=100, anchor="center")
        
        self.settlement_tree.heading("settle_date", text=self.get_text('settle_date'))
        self.settlement_tree.column("settle_date", width=120, anchor="center")
        
        self.settlement_tree.heading("total_amount", text=self.get_text('total_amount'))
        self.settlement_tree.column("total_amount", width=100, anchor="e")
        
        self.settlement_tree.heading("cashier", text=self.get_text('cashier'))
        self.settlement_tree.column("cashier", width=100, anchor="center")
        
        self.settlement_tree.heading("notes", text=self.get_text('notes'))
        self.settlement_tree.column("notes", width=200, anchor="w")
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.settlement_tree.yview)
        self.settlement_tree.configure(yscroll=scrollbar.set)
        
        self.settlement_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 统计标签容器
        stats_frame = ttk.Frame(list_frame)
        stats_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=5)
        
        # 总计金额标签
        self.total_amount_label = ttk.Label(stats_frame, text="总计：0.00 元", font=("Arial", 10, "bold"))
        self.total_amount_label.pack(side=tk.RIGHT, anchor="e")
        
        # 右侧：结算明细表单
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 表单标题
        self.form_title = ttk.Label(right_frame, text=self.get_text('settlement_details'), font=("Arial", 14))
        self.form_title.pack(side=tk.TOP, pady=10)
        
        # 表单框架
        form_frame = ttk.Frame(right_frame)
        form_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 表单控件
        row = 0
        
        # 结算月份
        self.form_labels['settle_month_label'] = ttk.Label(form_frame, text=self.get_text('settle_month') + ':')
        self.form_labels['settle_month_label'].grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        
        self.form_settle_month = ttk.Entry(form_frame)
        self.form_settle_month.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        self.form_settle_month.insert(0, datetime.now().strftime("%Y-%m"))
        row += 1
        
        # 计算当月已收金额按钮
        self.calculate_btn = ttk.Button(form_frame, text=self.get_text('calculate_monthly_received'), command=self.calculate_monthly_received)
        self.calculate_btn.grid(row=row, column=0, columnspan=2, pady=5)
        row += 1
        
        # 结算金额
        self.form_labels['total_amount_label'] = ttk.Label(form_frame, text=self.get_text('total_amount') + ':')
        self.form_labels['total_amount_label'].grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        
        self.form_total_amount = ttk.Entry(form_frame)
        self.form_total_amount.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 结算日期
        self.form_labels['settle_date_label'] = ttk.Label(form_frame, text=self.get_text('settle_date') + ':')
        self.form_labels['settle_date_label'].grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        
        self.form_settle_date = ttk.Entry(form_frame)
        self.form_settle_date.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        self.form_settle_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        row += 1
        
        # 出纳姓名
        self.form_labels['cashier_label'] = ttk.Label(form_frame, text=self.get_text('cashier') + ':')
        self.form_labels['cashier_label'].grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        
        self.form_cashier = ttk.Entry(form_frame)
        self.form_cashier.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 备注
        self.form_labels['notes_label'] = ttk.Label(form_frame, text=self.get_text('notes') + ':')
        self.form_labels['notes_label'].grid(row=row, column=0, sticky=tk.NE, padx=5, pady=5)
        
        self.form_notes = tk.Text(form_frame, height=5, width=40)
        self.form_notes.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 表单操作按钮
        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=row, column=0, columnspan=2, pady=20)
        
        self.form_buttons['save_btn'] = ttk.Button(button_frame, text=self.get_text('save'), command=self.save_settlement)
        self.form_buttons['save_btn'].pack(side=tk.LEFT, padx=10)
        
        self.form_buttons['cancel_btn'] = ttk.Button(button_frame, text=self.get_text('cancel'), command=self.clear_form)
        self.form_buttons['cancel_btn'].pack(side=tk.LEFT, padx=10)
        
        # 绑定列表选择事件
        self.settlement_tree.bind("<<TreeviewSelect>>", self.on_settlement_select)
    
    def load_settlement_list(self):
        """
        加载结算记录列表
        添加刷新列表数据的功能，确保用户能实时获取最新数据，并提供视觉反馈
        """
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            # 清空现有列表
            for item in self.settlement_tree.get_children():
                self.settlement_tree.delete(item)
            
            # 获取结算记录
            self.settlement_list = Settlement.get_all()
            
            # 添加到列表控件 - 带有序号列
            for index, settlement in enumerate(self.settlement_list, start=1):
                # 根据索引应用奇偶行标签
                row_tag = 'odd' if index % 2 != 0 else 'even'
                self.settlement_tree.insert("", tk.END, values=(index, 
                                                              settlement.id, 
                                                              settlement.settle_month, 
                                                              settlement.settle_date, 
                                                              settlement.total_amount, 
                                                              settlement.cashier, 
                                                              settlement.notes), tags=(row_tag,))
            
            # 更新统计标签
            self.update_stats_labels()
            
            # 只有在手动刷新时显示成功提示，初始加载时不显示
            if not self.is_initial_load:
                messagebox.showinfo(self.get_text('success'), f"{self.get_text('settlement_management')} {self.get_text('records')} {self.get_text('refreshed')}\n{self.get_text('total_loaded')}: {len(self.settlement_list)} {self.get_text('records')}")
            else:
                # 初始加载完成后，将标志设置为False
                self.is_initial_load = False
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('failed_to_refresh')} {self.get_text('settlement_management')} {self.get_text('records')}: {str(e)}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def search_settlements(self):
        """
        搜索结算记录
        """
        # 这里简化实现，只按月份过滤
        month = self.month_var.get()
        if not month:
            self.load_settlement_list()
            return
        
        # 清空现有列表
        for item in self.settlement_tree.get_children():
            self.settlement_tree.delete(item)
        
        # 获取指定月份的结算记录
        settlement = Settlement.get_by_month(month)
        if settlement:
            # 搜索结果只有一条，序号为1，应用奇偶行标签
            row_tag = 'odd' if 1 % 2 != 0 else 'even'
            self.settlement_tree.insert("", tk.END, values=(1, 
                                                          settlement.id, 
                                                          settlement.settle_month, 
                                                          settlement.settle_date, 
                                                          settlement.total_amount, 
                                                          settlement.cashier, 
                                                          settlement.notes), tags=(row_tag,))
        
        # 更新统计标签
        self.update_stats_labels()
    
    def reset_filter(self):
        """
        重置筛选条件
        """
        self.month_var.set(datetime.now().strftime("%Y-%m"))
        self.load_settlement_list()
    
    def update_stats_labels(self):
        """
        更新统计标签
        计算当前列表中所有结算记录的金额总和并显示
        """
        try:
            # 计算所有显示的结算记录的金额总和
            total_amount = 0.0
            
            # 遍历Treeview中的所有项目
            for item in self.settlement_tree.get_children():
                values = self.settlement_tree.item(item, "values")
                if values and len(values) > 4:  # 确保有总金额数据
                    try:
                        # 第5列（索引4）是total_amount
                        amount = float(values[4])
                        total_amount += amount
                    except (ValueError, TypeError):
                        # 跳过无效的金额值
                        continue
            
            # 更新总计金额标签
            self.total_amount_label.config(text=f"{self.get_text('total')}：{total_amount:.2f} {self.get_text('yuan')}")
        except Exception as e:
            # 出现错误时，确保标签有默认值
            self.total_amount_label.config(text=f"{self.get_text('total')}：0.00 {self.get_text('yuan')}")
            print(f"{self.get_text('failed_to_update_stats')}: {str(e)}")
    
    def on_settlement_select(self, event):
        """
        列表选择事件处理
        """
        selected_items = self.settlement_tree.selection()
        if selected_items:
            item_id = selected_items[0]
            values = self.settlement_tree.item(item_id, "values")
            settlement_id_str = values[1]  # ID值现在在索引1位置，因为添加了序号列
            
            try:
                # 将字符串ID转换为整数，因为settlement.id是整数类型
                settlement_id = int(settlement_id_str)
                
                # 查找对应的结算记录
                self.selected_settlement = next((s for s in self.settlement_list if s.id == settlement_id), None)
                if self.selected_settlement:
                    self.fill_form()
                else:
                    # 如果找不到记录，可能是数据已过期，刷新列表
                    messagebox.showwarning(self.get_text('warning'), self.get_text('data_expired_please_refresh'))
                    self.load_settlement_list()
            except ValueError:
                # 如果ID转换失败，显示错误信息
                messagebox.showerror(self.get_text('error'), self.get_text('settlement_id_invalid'))
                self.load_settlement_list()
    
    def fill_form(self):
        """
        填充表单
        """
        if self.selected_settlement:
            # 结算月份
            self.form_settle_month.delete(0, tk.END)
            self.form_settle_month.insert(0, self.selected_settlement.settle_month)
            
            # 结算金额
            self.form_total_amount.delete(0, tk.END)
            self.form_total_amount.insert(0, str(self.selected_settlement.total_amount))
            
            # 结算日期
            self.form_settle_date.delete(0, tk.END)
            self.form_settle_date.insert(0, self.selected_settlement.settle_date)
            
            # 出纳姓名
            self.form_cashier.delete(0, tk.END)
            self.form_cashier.insert(0, self.selected_settlement.cashier)
            
            # 备注
            self.form_notes.delete(1.0, tk.END)
            self.form_notes.insert(1.0, self.selected_settlement.notes)
    
    def clear_form(self):
        """
        清空表单
        """
        self.selected_settlement = None
        self.form_settle_month.delete(0, tk.END)
        self.form_settle_month.insert(0, datetime.now().strftime("%Y-%m"))
        self.form_total_amount.delete(0, tk.END)
        self.form_settle_date.delete(0, tk.END)
        self.form_settle_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.form_cashier.delete(0, tk.END)
        self.form_notes.delete(1.0, tk.END)
    
    def add_settlement(self):
        """
        添加结算记录
        实现创建新结算记录的功能，包含表单初始化和用户反馈
        """
        # 清空表单，准备添加新记录
        self.clear_form()
        
        # 添加明确的用户反馈，提示用户正在新增结算记录
        messagebox.showinfo(self.get_text('info'), f"{self.get_text('edit_mode')} {self.get_text('add_settlement')} {self.get_text('records')}\n\n{self.get_text('operation_guide')}:\n{self.get_text('guide_step_1')}\n{self.get_text('guide_step_2')}\n{self.get_text('guide_step_3')}\n{self.get_text('guide_step_4')}")
    
    def edit_settlement(self):
        """
        编辑结算记录
        """
        # 首先检查是否有选中的项目
        selected_items = self.settlement_tree.selection()
        if not selected_items:
            messagebox.showwarning(self.get_text('warning'), self.get_text('please_select_settlement_to_edit'))
            return
        
        # 如果没有selected_settlement，尝试从选中的项目中获取
        if not self.selected_settlement:
            item_id = selected_items[0]
            values = self.settlement_tree.item(item_id, "values")
            settlement_id_str = values[1]
            
            try:
                settlement_id = int(settlement_id_str)
                # 重新查找结算记录
                self.selected_settlement = next((s for s in self.settlement_list if s.id == settlement_id), None)
                
                if not self.selected_settlement:
                    # 如果还是找不到，刷新列表并提示
                    messagebox.showwarning(self.get_text('warning'), self.get_text('data_expired_please_refresh'))
                    self.load_settlement_list()
                    return
            except ValueError:
                messagebox.showerror(self.get_text('error'), self.get_text('settlement_id_invalid'))
                self.load_settlement_list()
                return
        
        # 确保表单已填充数据
        self.fill_form()
        
        # 添加明确的用户反馈，提示用户正在编辑结算记录
        messagebox.showinfo(self.get_text('info'), f"{self.get_text('edit_mode')}, {self.get_text('editing')} {self.selected_settlement.settle_month} {self.get_text('month')} {self.get_text('settlement_management')} {self.get_text('records')}\n{self.get_text('please_modify_form_and_save')}")
        
        # 高亮显示正在编辑的行
        for item in self.settlement_tree.get_children():
            values = self.settlement_tree.item(item, "values")
            if int(values[1]) == self.selected_settlement.id:  # 确保类型匹配
                self.settlement_tree.selection_set(item)
                self.settlement_tree.focus(item)
                break
    
    def delete_settlement(self):
        """
        删除结算记录
        提供删除结算记录的功能，包含二次确认机制和详细的记录信息
        """
        # 首先检查是否有选中的项目
        selected_items = self.settlement_tree.selection()
        if not selected_items:
            messagebox.showwarning(self.get_text('warning'), self.get_text('please_select_settlement_to_delete'))
            return
        
        # 如果没有selected_settlement，尝试从选中的项目中获取
        if not self.selected_settlement:
            item_id = selected_items[0]
            values = self.settlement_tree.item(item_id, "values")
            settlement_id_str = values[1]
            
            try:
                settlement_id = int(settlement_id_str)
                # 重新查找结算记录
                self.selected_settlement = next((s for s in self.settlement_list if s.id == settlement_id), None)
                
                if not self.selected_settlement:
                    # 如果还是找不到，刷新列表并提示
                    messagebox.showwarning(self.get_text('warning'), self.get_text('data_expired_please_refresh'))
                    self.load_settlement_list()
                    return
            except ValueError:
                messagebox.showerror(self.get_text('error'), self.get_text('settlement_id_invalid'))
                self.load_settlement_list()
                return
        
        # 显示更详细的确认对话框，包含要删除的结算记录信息
        confirm_message = f"{self.get_text('confirm_delete_settlement')}？\n\n"
        confirm_message += f"{self.get_text('settle_month')}: {self.selected_settlement.settle_month}\n"
        confirm_message += f"{self.get_text('settle_date')}: {self.selected_settlement.settle_date}\n"
        confirm_message += f"{self.get_text('total_amount')}: {self.selected_settlement.total_amount:.2f} {self.get_text('yuan')}\n"
        confirm_message += f"{self.get_text('cashier')}: {self.selected_settlement.cashier}\n"
        if self.selected_settlement.notes:
            confirm_message += f"{self.get_text('notes')}: {self.selected_settlement.notes}\n"
        confirm_message += f"\n{self.get_text('delete_irreversible')}"
        
        if messagebox.askyesno(self.get_text('confirm_delete_settlement'), confirm_message):
            try:
                # 更新UI状态，显示加载中
                self.parent.config(cursor="wait")
                self.parent.update()
                
                if self.selected_settlement.delete():
                    messagebox.showinfo(self.get_text('success'), self.get_text('settlement_deleted_successfully'))
                    self.load_settlement_list()
                    self.clear_form()
                else:
                    messagebox.showerror(self.get_text('error'), self.get_text('settlement_delete_failed'))
            finally:
                # 恢复UI状态
                self.parent.config(cursor="")
                self.parent.update()
    
    def calculate_monthly_received(self):
        """
        计算当月已收金额
        计算并显示指定月份的已收金额，包含详细的计算结果和用户反馈
        """
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            month = self.form_settle_month.get().strip()
            
            # 1. 验证月份格式
            if not month:
                raise ValueError(self.get_text('please_enter_settlement_month'))
            
            try:
                # 验证月份格式为YYYY-MM
                month_parts = month.split('-')
                if len(month_parts) != 2:
                    raise ValueError
                _, month_num = map(int, month_parts)
                if not (1 <= month_num <= 12):
                    raise ValueError
            except ValueError:
                raise ValueError(self.get_text('settle_month_format_error'))
            
            # 2. 获取当月所有已收金额
            payments = Payment.get_by_month(month)
            
            # 3. 计算总金额
            total_received = sum(payment.amount for payment in payments)
            
            # 4. 统计支付方式
            payment_methods = {}
            for payment in payments:
                method = payment.payment_method
                if method not in payment_methods:
                    payment_methods[method] = 0
                payment_methods[method] += 1
            
            # 5. 更新表单字段
            self.form_total_amount.delete(0, tk.END)
            self.form_total_amount.insert(0, f"{total_received:.2f}")
            
            # 6. 生成详细的计算结果说明
            result_message = f"{month} {self.get_text('month')} {self.get_text('monthly_received_calculation_result')}\n\n"
            result_message += f"{self.get_text('total_received_amount')}: {total_received:.2f} {self.get_text('yuan')}\n"
            result_message += f"{self.get_text('payment_count')}: {len(payments)} {self.get_text('records')}\n\n"
            
            if payment_methods:
                result_message += f"{self.get_text('payment_method_distribution')}:\n"
                for method, count in payment_methods.items():
                    result_message += f"  - {method}: {count} {self.get_text('records')}\n"
            else:
                result_message += f"{self.get_text('payment_method_distribution')}: {self.get_text('no_data')}\n"
            
            result_message += f"\n{self.get_text('calculation_result_filled')}\n{self.get_text('manual_adjustment_allowed')}"
            
            # 7. 显示详细的计算结果
            messagebox.showinfo(self.get_text('calculation_result'), result_message)
        except ValueError as e:
            # 显示验证错误信息
            messagebox.showwarning(self.get_text('warning'), str(e))
        except Exception as e:
            # 显示其他错误信息
            messagebox.showerror(self.get_text('error'), f"{self.get_text('calculate_monthly_received')} {self.get_text('failed')}: {str(e)}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def save_settlement(self):
        """
        保存结算记录
        实现创建新结算记录或修改已有记录的功能，包括表单验证、数据提交和成功/失败状态反馈
        """
        # 获取表单数据
        settle_month = self.form_settle_month.get().strip()
        total_amount_str = self.form_total_amount.get().strip()
        settle_date = self.form_settle_date.get().strip()
        cashier = self.form_cashier.get().strip()
        notes = self.form_notes.get(1.0, tk.END).strip()
        
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            # 验证数据
            # 1. 验证结算月份格式
            if not settle_month:
                raise ValueError(self.get_text('please_enter_settlement_month'))
            try:
                # 验证月份格式为YYYY-MM
                month_parts = settle_month.split('-')
                if len(month_parts) != 2:
                    raise ValueError
                year, month = map(int, month_parts)
                if not (1 <= month <= 12):
                    raise ValueError
            except ValueError:
                raise ValueError(self.get_text('settle_month_format_error'))
            
            # 2. 验证结算金额
            if not total_amount_str:
                raise ValueError(self.get_text('please_enter_settlement_amount'))
            try:
                total_amount = float(total_amount_str)
                if total_amount < 0:
                    raise ValueError(self.get_text('settlement_amount_cannot_be_negative'))
            except ValueError:
                raise ValueError(self.get_text('settlement_amount_must_be_valid_number'))
            
            # 3. 验证结算日期格式
            if not settle_date:
                raise ValueError(self.get_text('please_enter_settlement_date'))
            try:
                # 验证日期格式为YYYY-MM-DD
                date_obj = datetime.strptime(settle_date, "%Y-%m-%d")
            except ValueError:
                raise ValueError(self.get_text('settlement_date_format_error'))
            
            # 4. 验证结算日期合理性
            # 结算日期不能早于结算月份的第一天
            month_start = datetime.strptime(f"{settle_month}-01", "%Y-%m-%d")
            if date_obj < month_start:
                raise ValueError(f"{self.get_text('settlement_date_cannot_be_earlier_than')} {settle_month}-01")
            
            # 结算日期不能晚于当前日期
            current_date = datetime.now()
            if date_obj > current_date:
                raise ValueError(self.get_text('settlement_date_cannot_be_later_than_today'))
            
            # 5. 验证出纳姓名
            if not cashier:
                raise ValueError(self.get_text('please_enter_cashier_name'))
            if len(cashier) > 50:
                raise ValueError(self.get_text('cashier_name_cannot_exceed_50_characters'))
            
            # 6. 验证备注长度
            if len(notes) > 200:
                raise ValueError(self.get_text('notes_cannot_exceed_200_characters'))
            
            # 保存结算记录
            if self.selected_settlement:
                # 更新现有记录
                old_month = self.selected_settlement.settle_month
                self.selected_settlement.settle_month = settle_month
                self.selected_settlement.total_amount = total_amount
                self.selected_settlement.settle_date = settle_date
                self.selected_settlement.cashier = cashier
                self.selected_settlement.notes = notes
                
                # 检查是否修改了月份，且新月份已存在结算记录
                if old_month != settle_month:
                    existing_settlement = Settlement.get_by_month(settle_month)
                    if existing_settlement and existing_settlement.id != self.selected_settlement.id:
                        raise ValueError(f"{settle_month} {self.get_text('month_already_has_settlement_record_cannot_modify')}")
                
                if self.selected_settlement.save():
                    messagebox.showinfo(self.get_text('success'), f"{self.get_text('settlement_update_success')}".format(settle_month))
                    self.load_settlement_list()
                    self.clear_form()
                else:
                    raise Exception(self.get_text('database_operation_failed'))
            else:
                # 添加新记录
                # 检查该月份是否已经有结算记录
                existing_settlement = Settlement.get_by_month(settle_month)
                if existing_settlement:
                    raise ValueError(f"{settle_month} {self.get_text('month_already_has_settlement_record')}")
                
                settlement = Settlement(
                    settle_month=settle_month,
                    total_amount=total_amount,
                    settle_date=settle_date,
                    cashier=cashier,
                    notes=notes
                )
                
                if settlement.save():
                    messagebox.showinfo(self.get_text('success'), f"{self.get_text('settlement_add_success')}".format(settle_month))
                    self.load_settlement_list()
                    self.clear_form()
                else:
                    raise Exception(self.get_text('database_operation_failed'))
        except ValueError as e:
            # 显示验证错误信息
            messagebox.showwarning(self.get_text('warning'), str(e))
        except Exception as e:
            # 显示其他错误信息
            messagebox.showerror(self.get_text('error'), f"{self.get_text('failed_to_save_settlement_record')}: {str(e)}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def generate_settlement_report(self):
        """
        生成结算报表
        实现按指定条件生成结算报表的功能，支持Excel格式导出
        """
        # 让用户选择报表月份
        month = self.month_var.get()
        if not month:
            month = datetime.now().strftime("%Y-%m")
            self.month_var.set(month)
        
        # 询问用户是否确认生成该月份的报表
        confirm = messagebox.askyesno(self.get_text('confirm_generate_report'), f"{self.get_text('want_to_generate_settlement_report_for_month')} {month}？")
        if not confirm:
            return
        
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 1. 生成结算汇总表
            ws_summary = wb.active
            ws_summary.title = f"{month}结算汇总"
            
            # 设置报表标题
            ws_summary.merge_cells('A1:G1')
            title_cell = ws_summary['A1']
            title_cell.value = f"{month}月份结算报表"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 报表生成时间
            ws_summary['A2'] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_summary['A2'].font = Font(size=10, italic=True)
            
            # 生成月度汇总数据
            summary_data = self.calculate_monthly_summary(month)
            
            # 汇总表头
            summary_headers = ["项目", "金额/数量", "说明"]
            ws_summary.append(summary_headers)
            
            # 设置表头样式
            for cell in ws_summary[3]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.border = None
            
            # 汇总数据
            summary_rows = [
                ["总费用金额", f"{summary_data['total_charge']:.2f} 元", "当月所有租户总费用"],
                ["已收金额", f"{summary_data['total_received']:.2f} 元", "当月实际收到的金额"],
                ["欠费金额", f"{summary_data['total_arrears']:.2f} 元", "当月未收到的金额"],
                ["已缴户数", f"{summary_data['paid_count']} 户", "已缴纳费用的租户数量"],
                ["总租户数", f"{summary_data['total_count']} 户", "当月有费用记录的租户总数"],
                ["缴费率", f"{summary_data['payment_rate']:.2f}%", "已缴费租户占比"]
            ]
            
            for row_data in summary_rows:
                ws_summary.append(row_data)
            
            # 2. 生成结算明细 sheet
            ws_details = wb.create_sheet(title=f"{month}结算明细")
            
            # 明细表头
            detail_headers = ["序号", "结算月份", "结算日期", "结算金额", "出纳", "备注"]
            ws_details.append(detail_headers)
            
            # 设置表头样式
            for cell in ws_details[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.border = None
            
            # 获取结算记录
            settlement = Settlement.get_by_month(month)
            if settlement:
                ws_details.append([
                    1,
                    settlement.settle_month,
                    settlement.settle_date,
                    f"{settlement.total_amount:.2f}",
                    settlement.cashier,
                    settlement.notes
                ])
            
            # 3. 生成缴费明细 sheet
            ws_payments = wb.create_sheet(title=f"{month}缴费明细")
            
            # 缴费明细表头
            payment_headers = ["序号", "租户名称", "缴费金额", "缴费日期", "缴费方式", "收据号", "收费人"]
            ws_payments.append(payment_headers)
            
            # 设置表头样式
            for cell in ws_payments[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.border = None
            
            # 获取缴费记录
            payments = Payment.get_by_month(month)
            for idx, payment in enumerate(payments, 1):
                # 获取租户名称
                tenant_name = "未知租户"
                if payment.charge and payment.charge.tenant:
                    tenant_name = payment.charge.tenant.name
                elif payment.charge:
                    # 如果charge对象存在但tenant对象不存在，尝试加载
                    payment.charge.load_tenant_info()
                    tenant_name = payment.charge.tenant.name if payment.charge.tenant else "未知租户"
                
                # Payment对象没有receipt_no和collector字段，使用payer字段
                ws_payments.append([
                    idx,
                    tenant_name,
                    f"{payment.amount:.2f}",
                    payment.payment_date,
                    payment.payment_method,
                    "",  # receipt_no字段不存在，留空
                    payment.payer  # 使用payer字段代替collector
                ])
            
            # 设置列宽
            for ws in wb.worksheets:
                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    # 处理合并单元格的情况，使用列的索引来获取列字母
                    col_idx = column_cells[0].column
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, 30)
            
            # 让用户选择保存路径
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title=self.get_text('export_settlement_report'),
                initialfile=f"{month}_{self.get_text('settlement_report')}.xlsx"
            )
            
            if not file_path:
                return
            
            # 保存文件
            wb.save(file_path)
            
            messagebox.showinfo(self.get_text('success'), f"{self.get_text('settlement_report_successfully_exported_to')}\n{file_path}")
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('failed_to_generate_settlement_report')}：{str(e)}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def calculate_monthly_summary(self, month):
        """
        计算月度汇总数据
        :param month: 月份
        :return: 汇总数据字典
        """
        # 获取当月所有费用记录
        charges = Charge.get_by_month(month)
        
        # 计算总费用
        total_charge = sum(charge.total_charge for charge in charges)
        
        # 获取当月所有已收金额
        payments = Payment.get_by_month(month)
        total_received = sum(payment.amount for payment in payments)
        
        # 计算欠费金额
        total_arrears = total_charge - total_received
        
        # 统计缴费率
        paid_count = len([c for c in charges if c.status == "已缴"])
        total_count = len(charges)
        payment_rate = (paid_count / total_count * 100) if total_count > 0 else 0
        
        return {
            "month": month,
            "total_charge": total_charge,
            "total_received": total_received,
            "total_arrears": total_arrears,
            "paid_count": paid_count,
            "total_count": total_count,
            "payment_rate": payment_rate
        }