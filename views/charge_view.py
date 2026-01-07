#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
费用管理视图
负责处理费用计算、费用查询等功能
"""

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from models.tenant import Tenant
from models.charge import Charge
from models.reading import MeterReading
from models.meter import Meter
from models.payment import Payment
from utils.language_utils import LanguageUtils

class ChargeView:
    """费用管理视图类"""
    
    def __init__(self, parent, language_utils):
        """
        初始化费用管理视图
        :param parent: 父窗口组件
        :param language_utils: 语言工具类实例
        """
        self.parent = parent
        self.language_utils = language_utils
        self.charge_list = []
        self.selected_charge = None
        self.create_widgets()
        self.load_charge_list()
        
    def get_text(self, key):
        """
        获取翻译文本
        :param key: 文本键名
        :return: 翻译后的文本
        """
        return self.language_utils.get_text(key)
        
    def update_language(self):
        """
        更新界面语言
        """
        # 更新筛选框文本
        self.filter_labels['month_label']['text'] = self.get_text('month') + ':'
        self.filter_labels['tenant_type_label']['text'] = self.get_text('tenant_type') + ':'
        self.filter_labels['tenant_label']['text'] = self.get_text('tenant') + ':'
        self.filter_labels['status_label']['text'] = self.get_text('status') + ':'
        
        # 更新筛选按钮文本
        self.filter_buttons['search_btn']['text'] = self.get_text('button_search')
        self.filter_buttons['reset_btn']['text'] = self.get_text('button_reset')
        
        # 更新操作按钮文本
        self.action_buttons['calculate_btn']['text'] = self.get_text('calculate_charges')
        self.action_buttons['export_btn']['text'] = self.get_text('export_charge_sheet')
        self.action_buttons['delete_btn']['text'] = self.get_text('delete_records')
        
        # 更新全选文本
        self.action_labels['select_all_label']['text'] = self.get_text('select_all') + ':'
        
        # 更新状态选项
        self.status_combobox['values'] = ["", self.get_text('unpaid'), self.get_text('paid'), self.get_text('partially_paid')]
        
        # 更新租户类型选项
        self.load_tenant_types_to_combobox()
        
        # 更新列表列标题
        for col in self.charge_tree["columns"][1:]:  # 跳过select列
            if col == "serial":
                self.charge_tree.heading(col, text=self.get_text('serial'))  # 序号列不需要排序
            elif col == "tenant_type":
                self.charge_tree.heading(col, text=self.get_text('tenant_type'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "tenant_name":
                self.charge_tree.heading(col, text=self.get_text('tenant_name'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "month":
                self.charge_tree.heading(col, text=self.get_text('month'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "water_usage":
                self.charge_tree.heading(col, text=self.get_text('water_usage'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "water_price":
                self.charge_tree.heading(col, text=self.get_text('water_price'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "water_charge":
                self.charge_tree.heading(col, text=self.get_text('water_charge'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "electricity_usage":
                self.charge_tree.heading(col, text=self.get_text('electricity_usage'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "electricity_price":
                self.charge_tree.heading(col, text=self.get_text('electricity_price'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "electricity_charge":
                self.charge_tree.heading(col, text=self.get_text('electricity_charge'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "total_charge":
                self.charge_tree.heading(col, text=self.get_text('total_charge'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "paid_amount":
                self.charge_tree.heading(col, text=self.get_text('paid_amount'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "due_amount":
                self.charge_tree.heading(col, text=self.get_text('due_amount'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "status":
                self.charge_tree.heading(col, text=self.get_text('status'), command=lambda c=col: self.sort_column_clicked(c))
        
        # 更新统计标签
        self.stats_labels["total_count"].config(text=f"{self.get_text('total')} {len(self.charge_list)} {self.get_text('records')}")
        
        # 更新明细标题
        self.detail_title['text'] = self.get_text('charge_details')
        
        # 如果有选中的费用，重新显示明细
        if self.selected_charge:
            self.show_charge_detail()
        
        # 重新加载列表，确保状态文本正确
        self.load_charge_list()
    
    def create_widgets(self):
        """
        创建费用管理界面组件
        """
        # 创建主框架
        main_frame = ttk.Frame(self.parent)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 顶部：筛选和操作按钮
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        
        # 筛选条件
        filter_frame = ttk.Frame(top_frame)
        filter_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 筛选标签字典
        self.filter_labels = {}
        # 筛选按钮字典
        self.filter_buttons = {}
        # 操作按钮字典
        self.action_buttons = {}
        # 操作标签字典
        self.action_labels = {}
        
        # 月份标签
        self.filter_labels['month_label'] = ttk.Label(filter_frame, text=self.get_text('month') + ':')
        self.filter_labels['month_label'].pack(side=tk.LEFT, padx=5)
        
        self.month_var = tk.StringVar()
        self.month_var.set(datetime.now().strftime("%Y-%m"))
        # 将月份选择从Entry改为Combobox
        self.month_combobox = ttk.Combobox(filter_frame, textvariable=self.month_var, width=10)
        # 生成最近24个月的选项
        self.generate_month_options()
        self.month_combobox.pack(side=tk.LEFT, padx=5)
        
        # 租户类型筛选
        self.filter_labels['tenant_type_label'] = ttk.Label(filter_frame, text=self.get_text('tenant_type') + ':')
        self.filter_labels['tenant_type_label'].pack(side=tk.LEFT, padx=5)
        
        self.tenant_type_var = tk.StringVar()
        self.tenant_type_combobox = ttk.Combobox(filter_frame, textvariable=self.tenant_type_var, width=10)
        # 初始化租户类型选项
        self.load_tenant_types_to_combobox()
        self.tenant_type_combobox.pack(side=tk.LEFT, padx=5)
        
        # 租户标签
        self.filter_labels['tenant_label'] = ttk.Label(filter_frame, text=self.get_text('tenant') + ':')
        self.filter_labels['tenant_label'].pack(side=tk.LEFT, padx=5)
        
        self.tenant_var = tk.StringVar()
        self.tenant_combobox = ttk.Combobox(filter_frame, textvariable=self.tenant_var)
        # 初始化租户ID映射字典
        self.tenant_id_map = {}
        self.load_tenants_to_combobox()
        self.tenant_combobox.pack(side=tk.LEFT, padx=5)
        
        # 状态标签
        self.filter_labels['status_label'] = ttk.Label(filter_frame, text=self.get_text('status') + ':')
        self.filter_labels['status_label'].pack(side=tk.LEFT, padx=5)
        
        self.status_var = tk.StringVar()
        self.status_combobox = ttk.Combobox(filter_frame, textvariable=self.status_var, values=["", self.get_text('unpaid'), self.get_text('paid'), self.get_text('partially_paid')])
        self.status_combobox.pack(side=tk.LEFT, padx=5)
        
        # 查询按钮
        self.filter_buttons['search_btn'] = ttk.Button(filter_frame, text=self.get_text('button_search'), command=self.search_charges)
        self.filter_buttons['search_btn'].pack(side=tk.LEFT, padx=5)
        
        # 重置按钮
        self.filter_buttons['reset_btn'] = ttk.Button(filter_frame, text=self.get_text('button_reset'), command=self.reset_filter)
        self.filter_buttons['reset_btn'].pack(side=tk.LEFT, padx=5)
        
        # 操作按钮
        action_frame = ttk.Frame(top_frame)
        action_frame.pack(side=tk.RIGHT)
        
        # 计算费用按钮
        self.action_buttons['calculate_btn'] = ttk.Button(action_frame, text=self.get_text('calculate_charges'), command=self.calculate_charges)
        self.action_buttons['calculate_btn'].pack(side=tk.LEFT, padx=5)
        
        # 导出收费表按钮
        self.action_buttons['export_btn'] = ttk.Button(action_frame, text=self.get_text('export_charge_sheet'), command=self.export_charge_sheet)
        self.action_buttons['export_btn'].pack(side=tk.LEFT, padx=5)
        
        # 删除记录按钮
        self.action_buttons['delete_btn'] = ttk.Button(action_frame, text=self.get_text('delete_records'), command=self.delete_selected_charges)
        self.action_buttons['delete_btn'].pack(side=tk.LEFT, padx=5)
        
        # 初始化全选状态
        self.select_all_var = tk.BooleanVar()
        self.select_all_var.set(False)
        
        # 在删除记录按钮后方添加全选复选框
        self.action_labels['select_all_label'] = ttk.Label(action_frame, text=self.get_text('select_all') + ':')
        self.action_labels['select_all_label'].pack(side=tk.LEFT, padx=(15, 2))
        
        ttk.Checkbutton(action_frame, variable=self.select_all_var, command=self.toggle_select_all).pack(side=tk.LEFT, padx=(0, 5))
        
        # 中部：中间框架，用于水平排列费用列表和费用明细
        middle_frame = ttk.Frame(main_frame)
        middle_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 配置grid布局，使两列能够自适应宽度
        middle_frame.grid_columnconfigure(0, weight=1, minsize=400)
        middle_frame.grid_columnconfigure(1, weight=1, minsize=300)
        middle_frame.grid_rowconfigure(0, weight=1)
        
        # 左侧：费用列表
        list_frame = ttk.Frame(middle_frame)
        list_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        
        # 列表控件 - 添加复选框列、序号列，移除ID列显示
        columns = ("select", "serial", "tenant_type", "tenant_name", "month", "water_usage", "water_price", "water_charge",
                  "electricity_usage", "electricity_price", "electricity_charge", "total_charge", "paid_amount", "due_amount", "status")
        self.charge_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        
        # 保存排序状态
        self.sort_column = None
        self.sort_order = "asc"
        
        # 设置列标题和宽度
        self.charge_tree.heading("select", text="")
        self.charge_tree.column("select", width=40, anchor="center")
        
        # 为可排序列添加点击事件
        for col in columns[1:]:  # 跳过select列
            # 设置默认标题
            if col == "serial":
                self.charge_tree.heading(col, text=self.get_text('serial'))  # 序号列不需要排序
            elif col == "tenant_type":
                self.charge_tree.heading(col, text=self.get_text('tenant_type'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "tenant_name":
                self.charge_tree.heading(col, text=self.get_text('tenant_name'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "month":
                self.charge_tree.heading(col, text=self.get_text('month'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "water_usage":
                self.charge_tree.heading(col, text=self.get_text('water_usage'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "water_price":
                self.charge_tree.heading(col, text=self.get_text('water_price'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "water_charge":
                self.charge_tree.heading(col, text=self.get_text('water_charge'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "electricity_usage":
                self.charge_tree.heading(col, text=self.get_text('electricity_usage'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "electricity_price":
                self.charge_tree.heading(col, text=self.get_text('electricity_price'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "electricity_charge":
                self.charge_tree.heading(col, text=self.get_text('electricity_charge'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "total_charge":
                self.charge_tree.heading(col, text=self.get_text('total_charge'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "paid_amount":
                self.charge_tree.heading(col, text=self.get_text('paid_amount'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "due_amount":
                self.charge_tree.heading(col, text=self.get_text('due_amount'), command=lambda c=col: self.sort_column_clicked(c))
            elif col == "status":
                self.charge_tree.heading(col, text=self.get_text('status'), command=lambda c=col: self.sort_column_clicked(c))
        
        # 设置列宽
        self.charge_tree.column("serial", width=60, anchor="center")
        self.charge_tree.column("tenant_type", width=100, anchor="center")
        self.charge_tree.column("tenant_name", width=150, anchor="w")
        self.charge_tree.column("month", width=100, anchor="center")
        self.charge_tree.column("water_usage", width=80, anchor="e")
        self.charge_tree.column("water_price", width=80, anchor="e")
        self.charge_tree.column("water_charge", width=80, anchor="e")
        self.charge_tree.column("electricity_usage", width=80, anchor="e")
        self.charge_tree.column("electricity_price", width=80, anchor="e")
        self.charge_tree.column("electricity_charge", width=80, anchor="e")
        self.charge_tree.column("total_charge", width=100, anchor="e")
        self.charge_tree.column("paid_amount", width=100, anchor="e")
        self.charge_tree.column("due_amount", width=100, anchor="e")
        self.charge_tree.column("status", width=80, anchor="center")
        
        # 配置Treeview的tag，用于实现隔行背景色
        # 奇数行使用深灰色，偶数行使用白色，与reading_view.py保持一致
        self.charge_tree.tag_configure('odd', background='#c0c0c0')
        self.charge_tree.tag_configure('even', background='#ffffff')
        
        # 添加滚动条
        v_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.charge_tree.yview)
        h_scrollbar = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.charge_tree.xview)
        self.charge_tree.configure(yscroll=v_scrollbar.set, xscroll=h_scrollbar.set)
        
        # 配置列表框架的grid布局
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        self.charge_tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # 右侧：费用明细
        detail_frame = ttk.Frame(middle_frame)
        detail_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        
        # 配置明细框架的grid布局
        detail_frame.grid_rowconfigure(1, weight=1)
        detail_frame.grid_columnconfigure(0, weight=1)
        
        # 明细标题
        self.detail_title = ttk.Label(detail_frame, text=self.get_text('charge_details'), font=("Arial", 14))
        self.detail_title.grid(row=0, column=0, pady=10, sticky="n")
        
        # 明细内容区域
        self.detail_text = tk.Text(detail_frame, wrap=tk.WORD)
        self.detail_text.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        
        # 为明细文本添加垂直滚动条
        detail_scrollbar = ttk.Scrollbar(detail_frame, orient=tk.VERTICAL, command=self.detail_text.yview)
        self.detail_text.configure(yscroll=detail_scrollbar.set)
        detail_scrollbar.grid(row=1, column=1, sticky="ns", pady=5)
        
        # 绑定列表选择事件和复选框点击事件
        self.charge_tree.bind("<<TreeviewSelect>>", self.on_charge_select)
        self.charge_tree.bind("<Button-1>", self.on_tree_click)
        
        # 保存选中状态的字典
        self.selected_items = {}
        
        # 添加统计信息框架
        stats_frame = ttk.Frame(main_frame)
        stats_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=10)
        
        # 创建统计标签
        self.stats_labels = {
            "total_count": ttk.Label(stats_frame, text=f"{self.get_text('total')} 0 {self.get_text('records')}", font=("Arial", 10)),
            "total_water_charge": ttk.Label(stats_frame, text=f"{self.get_text('total_water_charge')}: 0.00 {self.get_text('yuan')}", font=("Arial", 10)),
            "total_electricity_charge": ttk.Label(stats_frame, text=f"{self.get_text('total_electricity_charge')}: 0.00 {self.get_text('yuan')}", font=("Arial", 10)),
            "total_charge": ttk.Label(stats_frame, text=f"{self.get_text('total_charge')}: 0.00 {self.get_text('yuan')}", font=("Arial", 10)),
            "total_paid": ttk.Label(stats_frame, text=f"{self.get_text('total_paid_amount')}: 0.00 {self.get_text('yuan')}", font=("Arial", 10)),
            "total_due": ttk.Label(stats_frame, text=f"{self.get_text('total_due_amount')}: 0.00 {self.get_text('yuan')}", font=("Arial", 10))
        }
        
        # 设置统计标签的布局
        for label in self.stats_labels.values():
            label.pack(side=tk.LEFT, padx=15, pady=5)
    
    def generate_month_options(self):
        """
        生成月份选项
        生成最近24个月的月份选项
        """
        import datetime
        months = []
        current_date = datetime.datetime.now()
        
        # 生成最近24个月
        for i in range(24):
            month_date = current_date - datetime.timedelta(days=i*30)
            months.append(month_date.strftime("%Y-%m"))
        
        # 去重并排序
        unique_months = sorted(list(set(months)), reverse=True)
        self.month_combobox['values'] = unique_months
    
    def update_stats_labels(self):
        """
        更新统计标签数据
        计算当前列表中记录的总条数、总水费金额、总电费金额、水电费总额、已收总费用金额以及应收总费用金额
        """
        # 计算总条数
        total_count = len(self.charge_list)
        
        # 计算总水费金额、总电费金额、水电费总额、已收总费用金额、应收总费用金额
        total_water_charge = 0.0
        total_electricity_charge = 0.0
        total_charge = 0.0
        total_paid = 0.0
        total_due = 0.0
        
        for charge in self.charge_list:
            total_water_charge += charge.water_charge
            total_electricity_charge += charge.electricity_charge
            total_charge += charge.total_charge
            # 计算已收费用
            paid_amount = self.calculate_paid_amount(charge.id)
            total_paid += paid_amount
            # 计算应收费用
            due_amount = charge.total_charge - paid_amount
            total_due += due_amount
        
        # 更新标签文本，保留两位小数
        self.stats_labels["total_count"].config(text=f"{self.get_text('total')} {total_count} {self.get_text('records')}")
        self.stats_labels["total_water_charge"].config(text=f"{self.get_text('total_water_charge')}: {total_water_charge:.2f} {self.get_text('yuan')}")
        self.stats_labels["total_electricity_charge"].config(text=f"{self.get_text('total_electricity_charge')}: {total_electricity_charge:.2f} {self.get_text('yuan')}")
        self.stats_labels["total_charge"].config(text=f"{self.get_text('total_charge')}: {total_charge:.2f} {self.get_text('yuan')}")
        self.stats_labels["total_paid"].config(text=f"{self.get_text('total_paid_amount')}: {total_paid:.2f} {self.get_text('yuan')}")
        self.stats_labels["total_due"].config(text=f"{self.get_text('total_due_amount')}: {total_due:.2f} {self.get_text('yuan')}")
    
    def load_tenant_types_to_combobox(self):
        """
        加载租户类型数据到筛选下拉框
        """
        tenants = Tenant.get_all()
        # 获取所有唯一的租户类型
        tenant_types = set()
        for t in tenants:
            tenant_types.add(t.type)
        
        # 转换为列表并排序
        tenant_type_list = sorted(list(tenant_types))
        
        # 中文到英文翻译键的映射
        type_mapping = {
            '办公室': 'office',
            '门面': 'storefront'
        }
        
        # 英文到中文的反向映射，用于筛选时转换
        reverse_type_mapping = {
            'office': '办公室',
            'storefront': '门面'
        }
        
        # 保存反向映射，用于搜索时使用
        self.reverse_type_mapping = reverse_type_mapping
        
        # 翻译租户类型
        translated_types = []
        for type in tenant_type_list:
            # 获取翻译键
            type_key = type_mapping.get(type, type)
            # 翻译为当前语言
            translated_types.append(self.get_text(type_key))
        
        # 添加空选项作为默认值
        translated_types.insert(0, "")
        # 设置下拉框值
        self.tenant_type_combobox['values'] = translated_types
    
    def load_tenants_to_combobox(self):
        """
        加载租户数据到筛选下拉框
        """
        tenants = Tenant.get_all()
        # 清空映射字典
        self.tenant_id_map.clear()
        # 添加空选项
        tenant_names = [""]
        self.tenant_id_map[""] = None
        # 添加租户名称和ID映射
        for t in tenants:
            tenant_names.append(t.name)
            self.tenant_id_map[t.name] = t.id
        # 设置下拉框值
        self.tenant_combobox['values'] = tenant_names
    
    def calculate_paid_amount(self, charge_id):
        """
        计算指定费用记录的已收费用总额
        :param charge_id: 费用ID
        :return: 已收费用总额
        """
        payments = Payment.get_by_charge(charge_id)
        return sum(payment.amount for payment in payments)
    
    def load_charge_list(self):
        """
        加载费用列表
        """
        # 清空现有列表
        for item in self.charge_tree.get_children():
            self.charge_tree.delete(item)
        
        # 清空选中状态
        self.selected_items.clear()
        self.select_all_var.set(False)
        
        # 获取费用列表
        self.charge_list = Charge.get_by_month(self.month_var.get())
        
        # 创建租户ID到名称和类型的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        tenant_type_map = {t.id: t.type for t in tenants}
        
        # 添加到列表控件 - 第一个值为复选框状态，第二个值为序号
        for idx, charge in enumerate(self.charge_list, 1):
            tenant_name = tenant_map.get(charge.tenant_id, "未知租户")
            tenant_type = tenant_type_map.get(charge.tenant_id, "未知类型")
            
            # 中文到英文翻译键的映射
            type_mapping = {
                '办公室': 'office',
                '门面': 'storefront'
            }
            
            # 翻译租户类型
            type_key = type_mapping.get(tenant_type, tenant_type)
            translated_tenant_type = self.get_text(type_key)
            
            # 计算已收费用
            paid_amount = self.calculate_paid_amount(charge.id)
            # 计算应收费用
            due_amount = round(charge.total_charge - paid_amount, 2)
            
            # 动态计算状态
            if due_amount == 0:
                status = self.get_text('paid')
            elif due_amount > 0 and due_amount < charge.total_charge:
                status = self.get_text('partially_paid')
            elif due_amount == charge.total_charge and paid_amount == 0:
                status = self.get_text('unpaid')
            else:
                # 默认状态，处理其他情况 - 确保翻译
                status = self.get_text(charge.status) if charge.status else self.get_text('unpaid')
            
            # 初始复选框状态为空，根据索引应用奇偶行标签，同时使用tags存储charge.id
            row_tag = 'odd' if idx % 2 != 0 else 'even'
            self.charge_tree.insert("", tk.END, values=("□", idx, translated_tenant_type, tenant_name, 
                                                      charge.month, 
                                                      round(charge.water_usage, 2), 
                                                      round(charge.water_price, 2), 
                                                      round(charge.water_charge, 2), 
                                                      round(charge.electricity_usage, 2), 
                                                      round(charge.electricity_price, 2), 
                                                      round(charge.electricity_charge, 2), 
                                                      round(charge.total_charge, 2), 
                                                      round(paid_amount, 2), 
                                                      due_amount, 
                                                      status), tags=(row_tag, str(charge.id),))
        
        # 更新统计标签
        self.update_stats_labels()
        
        # 更新统计标签
        self.update_stats_labels()
    
    def on_tree_click(self, event):
        """
        处理树形视图的点击事件，用于切换复选框状态
        :param event: 事件对象
        """
        # 获取点击的项目和列
        region = self.charge_tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.charge_tree.identify_column(event.x)
            item = self.charge_tree.identify_row(event.y)
            
            # 检查是否点击了复选框列
            if column == "#1":  # select列
                # 获取当前状态
                current_values = self.charge_tree.item(item, "values")
                current_status = current_values[0]
                
                # 切换复选框状态
                if current_status == "□":
                    new_status = "☑"
                    selected = True
                else:
                    new_status = "□"
                    selected = False
                
                # 更新显示
                new_values = list(current_values)
                new_values[0] = new_status
                self.charge_tree.item(item, values=new_values)
                
                # 从tags中获取charge_id，tags现在包含(row_tag, charge_id)
                tags = self.charge_tree.item(item, "tags")
                if tags and len(tags) >= 2:
                    charge_id = int(tags[1])
                    self.selected_items[charge_id] = selected
                    
                    # 更新全选状态
                    self.update_select_all_status()
    
    def toggle_select_all(self):
        """
        处理全选/取消全选功能
        """
        select_all = self.select_all_var.get()
        
        # 更新所有行的复选框状态
        for item in self.charge_tree.get_children():
            values = list(self.charge_tree.item(item, "values"))
            if select_all:
                values[0] = "☑"
                selected = True
            else:
                values[0] = "□"
                selected = False
            self.charge_tree.item(item, values=values)
            
            # 从tags中获取charge_id，tags现在包含(row_tag, charge_id)
            tags = self.charge_tree.item(item, "tags")
            if tags and len(tags) >= 2:
                    charge_id = int(tags[1])
                    self.selected_items[charge_id] = selected
    
    def update_select_all_status(self):
        """
        更新全选复选框的状态
        """
        # 获取所有行的总数
        total_rows = len(self.charge_tree.get_children())
        if total_rows == 0:
            self.select_all_var.set(False)
            return
        
        # 计算选中的行数
        selected_count = sum(1 for selected in self.selected_items.values() if selected)
        
        # 更新全选状态
        self.select_all_var.set(selected_count == total_rows)
    
    def delete_selected_charges(self):
        """
        删除选中的费用记录
        前置检查：根据业务规则判定记录是否可删除
        """
        # 获取选中的记录ID
        selected_ids = [charge_id for charge_id, selected in self.selected_items.items() if selected]
        
        if not selected_ids:
            messagebox.showinfo(self.get_text('info'), self.get_text('please_select_records_to_delete'))
            return
        
        # 建立业务规则判定机制，识别不同的不可删除场景
        error_messages = []
        for charge_id in selected_ids:
            charge = Charge.get_by_id(charge_id)
            if not charge:
                continue
            
            # 1. 状态限制检查
            if charge.status == self.get_text('paid'):
                error_messages.append(f"{self.get_text('record')}{charge.id}{self.get_text('status_is_paid_cannot_delete')}")
                continue
            
            if charge.status == self.get_text('partially_paid'):
                error_messages.append(f"{self.get_text('record')}{charge.id}{self.get_text('status_is_partially_paid_cannot_delete')}")
                continue
            
            # 2. 关联交易限制检查：检查是否存在关联的收费记录
            payments = Payment.get_by_charge(charge_id)
            if payments:
                error_messages.append(f"{self.get_text('record')}{charge.id}{self.get_text('has_related_payment_records_cannot_delete')}")
                continue
        
        # 设计动态提示文本映射系统，确保每种不可删除场景对应唯一且明确的提示内容
        if error_messages:
            # 构建动态提示内容，清晰、准确且具有针对性
            # 将相同类型的错误合并显示
            error_type_map = {}
            for msg in error_messages:
                # 提取错误类型关键字
                if self.get_text('paid') in msg:
                    error_type = "paid_status"
                elif self.get_text('partially_paid') in msg:
                    error_type = "partially_paid_status"
                elif self.get_text('related_payment_records') in msg:
                    error_type = "related_transaction"
                else:
                    error_type = "other_reason"
                
                if error_type not in error_type_map:
                    error_type_map[error_type] = []
                error_type_map[error_type].append(msg)
            
            # 实现场景识别与提示文本的自动适配功能
            dynamic_messages = []
            if "paid_status" in error_type_map:
                dynamic_messages.append(self.get_text('charge_record_status_is_paid_cannot_delete'))
            if "partially_paid_status" in error_type_map:
                dynamic_messages.append(self.get_text('charge_record_status_is_partially_paid_cannot_delete'))
            if "related_transaction" in error_type_map:
                dynamic_messages.append(self.get_text('record_has_related_payment_records_not_allowed_delete'))
            
            # 合并所有动态提示内容
            final_message = "\n".join(dynamic_messages)
            
            # 保证动态提示内容清晰、准确且具有针对性，帮助用户直观理解具体限制原因
            messagebox.showwarning(self.get_text('operation_tip'), final_message)
            return
        
        # 确认删除
        confirmed = messagebox.askyesno(self.get_text('confirm_delete'), f"{self.get_text('confirm_delete_selected_records')} {len(selected_ids)} {self.get_text('records')}?\n{self.get_text('this_operation_cannot_be_undone')}")
        if not confirmed:
            return
        
        # 执行删除操作
        deleted_count = 0
        for charge_id in selected_ids:
            # 直接通过ID获取费用对象，不依赖self.charge_list
            charge = Charge.get_by_id(charge_id)
            if charge and charge.delete():
                deleted_count += 1
        
        # 显示删除结果
        messagebox.showinfo(self.get_text('delete_result'), f"{self.get_text('delete_completed')}!\n{self.get_text('successfully_deleted')}: {deleted_count} {self.get_text('records')}\n{self.get_text('failed')}: {len(selected_ids) - deleted_count} {self.get_text('records')}")
        
        # 刷新费用列表
        self.load_charge_list()
        # 刷新列表后，统计标签会在load_charge_list中更新
    
    def search_charges(self):
        """
        搜索费用记录
        根据月份、租户和状态进行筛选
        """
        # 清空现有列表
        for item in self.charge_tree.get_children():
            self.charge_tree.delete(item)
        
        # 清空选中状态
        self.selected_items.clear()
        self.select_all_var.set(False)
        
        # 获取筛选条件
        month = self.month_var.get()
        tenant_type_filter = self.tenant_type_var.get()
        tenant_name = self.tenant_var.get()
        status_filter = self.status_var.get()
        
        # 获取所有费用记录
        all_charges = Charge.get_all()
        
        # 创建租户ID到名称和类型的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        tenant_type_map = {t.id: t.type for t in tenants}
        tenant_id_map = {t.name: t.id for t in tenants}
        # 创建租户类型到租户ID列表的映射
        tenant_type_to_ids = {}
        for t in tenants:
            if t.type not in tenant_type_to_ids:
                tenant_type_to_ids[t.type] = []
            tenant_type_to_ids[t.type].append(t.id)
        
        # 应用筛选条件
        filtered_charges = []
        for charge in all_charges:
            # 按月份筛选
            if month and charge.month != month:
                continue
            
            # 按租户类型筛选
            if tenant_type_filter:
                # 获取当前费用记录对应的租户类型
                current_tenant_type = tenant_type_map.get(charge.tenant_id, "")
                
                # 中文到英文翻译键的映射
                type_mapping = {
                    '办公室': 'office',
                    '门面': 'storefront'
                }
                
                # 英文到中文的反向映射，用于筛选时转换
                reverse_type_mapping = {
                    'office': '办公室',
                    'storefront': '门面'
                }
                
                # 将翻译后的租户类型转换回原始中文类型
                # 首先将当前记录的中文类型转换为翻译键，然后获取其在当前语言下的翻译
                current_type_key = type_mapping.get(current_tenant_type, current_tenant_type)
                translated_current_type = self.get_text(current_type_key)
                
                if translated_current_type != tenant_type_filter:
                    continue
            
            # 按租户筛选
            if tenant_name:
                tenant_id = tenant_id_map.get(tenant_name)
                if charge.tenant_id != tenant_id:
                    continue
            
            # 计算已收费用和应收费用，用于状态筛选
            paid_amount = self.calculate_paid_amount(charge.id)
            due_amount = round(charge.total_charge - paid_amount, 2)
            
            # 动态计算状态
            if due_amount == 0:
                dynamic_status = self.get_text('paid')
            elif due_amount > 0 and due_amount < charge.total_charge:
                dynamic_status = self.get_text('partially_paid')
            elif due_amount == charge.total_charge and paid_amount == 0:
                dynamic_status = self.get_text('unpaid')
            else:
                dynamic_status = charge.status
            
            # 按状态筛选（使用动态计算的状态）
            if status_filter and dynamic_status != status_filter:
                continue
            
            filtered_charges.append(charge)
        
        # 更新费用列表
        self.charge_list = filtered_charges
        
        # 添加到列表控件 - 第一个值为复选框状态，第二个值为序号
        for idx, charge in enumerate(filtered_charges, 1):
            tenant_name = tenant_map.get(charge.tenant_id, "未知租户")
            tenant_type = tenant_type_map.get(charge.tenant_id, "未知类型")
            
            # 中文到英文翻译键的映射
            type_mapping = {
                '办公室': 'office',
                '门面': 'storefront'
            }
            
            # 翻译租户类型
            type_key = type_mapping.get(tenant_type, tenant_type)
            translated_tenant_type = self.get_text(type_key)
            
            # 计算已收费用
            paid_amount = self.calculate_paid_amount(charge.id)
            # 计算应收费用
            due_amount = round(charge.total_charge - paid_amount, 2)
            
            # 动态计算状态
            if due_amount == 0:
                status = self.get_text('paid')
            elif due_amount > 0 and due_amount < charge.total_charge:
                status = self.get_text('partially_paid')
            elif due_amount == charge.total_charge and paid_amount == 0:
                status = self.get_text('unpaid')
            else:
                # 默认状态，处理其他情况
                status = charge.status
            
            # 初始复选框状态为空，根据索引应用奇偶行标签，同时使用tags存储charge.id
            row_tag = 'odd' if idx % 2 != 0 else 'even'
            self.charge_tree.insert("", tk.END, values=("□", idx, translated_tenant_type, tenant_name, 
                                                      charge.month, 
                                                      round(charge.water_usage, 2), 
                                                      round(charge.water_price, 2), 
                                                      round(charge.water_charge, 2), 
                                                      round(charge.electricity_usage, 2), 
                                                      round(charge.electricity_price, 2), 
                                                      round(charge.electricity_charge, 2), 
                                                      round(charge.total_charge, 2), 
                                                      round(paid_amount, 2), 
                                                      due_amount, 
                                                      status), tags=(row_tag, str(charge.id),))
        
        # 更新统计标签
        self.update_stats_labels()
    
    def reset_filter(self):
        """
        重置筛选条件
        """
        self.month_var.set(datetime.now().strftime("%Y-%m"))
        self.tenant_type_combobox.set("")
        self.tenant_combobox.set("")
        self.status_combobox.set("")
        self.load_charge_list()
    
    def sort_column_clicked(self, column):
        """
        处理列点击事件，实现排序功能
        :param column: 被点击的列名
        """
        # 切换排序顺序
        if self.sort_column == column:
            # 同一列再次点击，切换排序顺序
            self.sort_order = "desc" if self.sort_order == "asc" else "asc"
        else:
            # 不同列点击，重置排序顺序为升序
            self.sort_column = column
            self.sort_order = "asc"
        
        # 获取当前显示的数据
        current_items = []
        for item in self.charge_tree.get_children():
            values = self.charge_tree.item(item, "values")
            current_items.append((values, item))
        
        # 根据列索引获取要排序的值
        column_index = self.charge_tree["columns"].index(column)
        
        # 定义排序函数
        def sort_key(item):
            value = item[0][column_index]
            # 尝试将数值型数据转换为float进行排序
            try:
                return float(value)
            except (ValueError, TypeError):
                # 非数值型数据直接比较
                return value
        
        # 执行排序
        current_items.sort(key=sort_key, reverse=(self.sort_order == "desc"))
        
        # 重新插入排序后的数据
        for i, (values, item) in enumerate(current_items):
            # 将item移动到指定位置
            self.charge_tree.move(item, "", i)
        
        # 更新序号列
        for i, item in enumerate(self.charge_tree.get_children(), 1):
            values = list(self.charge_tree.item(item, "values"))
            values[1] = i  # 更新序号列
            self.charge_tree.item(item, values=values)
        
        # 更新列标题，显示排序方向
        for col in self.charge_tree["columns"][1:]:  # 跳过select列
            if col == self.sort_column:
                # 添加排序指示符号
                if self.sort_order == "asc":
                    self.charge_tree.heading(col, text=f"{self.charge_tree.heading(col, 'text').split()[0]} ↑", command=lambda c=col: self.sort_column_clicked(c))
                else:
                    self.charge_tree.heading(col, text=f"{self.charge_tree.heading(col, 'text').split()[0]} ↓", command=lambda c=col: self.sort_column_clicked(c))
            else:
                # 移除其他列的排序指示符号
                # 只保留原始列名，移除排序符号
                original_text = self.charge_tree.heading(col, 'text').split()[0]
                self.charge_tree.heading(col, text=original_text, command=lambda c=col: self.sort_column_clicked(c))
    
    def on_charge_select(self, _event=None):
        """
        列表选择事件处理
        """
        selected_items = self.charge_tree.selection()
        if selected_items:
            item_id = selected_items[0]
            
            # 立即显示加载状态，提升用户体验
            self.detail_text.delete(1.0, tk.END)
            self.detail_text.insert(tk.END, f"{self.get_text('loading')}...\n\n{self.get_text('please_wait_getting_charge_details')}...")
            
            # 从tags中获取费用ID，tags现在包含(row_tag, charge_id)
            tags = self.charge_tree.item(item_id, "tags")
            if tags and len(tags) >= 2:
                charge_id = int(tags[1])
                
                # 查找对应的费用对象
                self.selected_charge = next((c for c in self.charge_list if c.id == charge_id), None)
                if self.selected_charge:
                    self.show_charge_detail()
                else:
                    # 如果在当前列表中找不到，尝试直接从数据库获取
                    self.selected_charge = Charge.get_by_id(charge_id)
                    if self.selected_charge:
                        self.show_charge_detail()
    
    def show_charge_detail(self):
        """
        显示费用明细
        """
        if self.selected_charge:
            # 获取租户名称（从内存中的租户映射获取，避免重复查询数据库）
            # 构建租户ID到名称的映射（如果还没有的话）
            if not hasattr(self, 'tenant_cache'):
                self.tenant_cache = {t.id: t.name for t in Tenant.get_all()}
            
            tenant_name = self.tenant_cache.get(self.selected_charge.tenant_id, self.get_text('unknown_tenant'))
            
            # 使用列表拼接字符串，比多次字符串拼接更高效
            detail_parts = []
            detail_parts.append(f"{self.get_text('charge_details')}\n")
            detail_parts.append("=" * 30+ "\n")
            detail_parts.append(f"{self.get_text('tenant_name')}: {tenant_name}\n")
            detail_parts.append(f"{self.get_text('charge_month')}: {self.selected_charge.month}\n")
            detail_parts.append("\n")
            detail_parts.append(f"{self.get_text('water_charge_details')}:\n")
            detail_parts.append(f"  {self.get_text('usage')}: {round(self.selected_charge.water_usage, 2)} {self.get_text('units')}\n")
            detail_parts.append(f"  {self.get_text('price')}: {round(self.selected_charge.water_price, 2)} {self.get_text('yuan')}/{self.get_text('unit')}\n")
            detail_parts.append(f"  {self.get_text('amount')}: {round(self.selected_charge.water_charge, 2)} {self.get_text('yuan')}\n")
            detail_parts.append("\n")
            detail_parts.append(f"{self.get_text('electricity_charge_details')}:\n")
            detail_parts.append(f"  {self.get_text('usage')}: {round(self.selected_charge.electricity_usage, 2)} {self.get_text('units')}\n")
            detail_parts.append(f"  {self.get_text('price')}: {round(self.selected_charge.electricity_price, 2)} {self.get_text('yuan')}/{self.get_text('unit')}\n")
            detail_parts.append(f"  {self.get_text('amount')}: {round(self.selected_charge.electricity_charge, 2)} {self.get_text('yuan')}\n")
            detail_parts.append("\n")
            detail_parts.append("=" * 30+ "\n")
            detail_parts.append(f"{self.get_text('total_charge')}: {round(self.selected_charge.total_charge, 2)} {self.get_text('yuan')}\n")
            detail_parts.append(f"{self.get_text('charge_status')}: {self.selected_charge.status}\n")
            
            # 计算已收费用和应收费用，显示在明细中
            paid_amount = self.calculate_paid_amount(self.selected_charge.id)
            due_amount = round(self.selected_charge.total_charge - paid_amount, 2)
            detail_parts.append(f"{self.get_text('paid_amount')}: {round(paid_amount, 2)} {self.get_text('yuan')}\n")
            detail_parts.append(f"{self.get_text('due_amount')}: {due_amount} {self.get_text('yuan')}\n")
            
            # 合并列表并显示明细
            detail = ''.join(detail_parts)
            self.detail_text.delete(1.0, tk.END)
            self.detail_text.insert(tk.END, detail)
            
            # 更新UI，确保立即显示
            self.parent.update_idletasks()
    
    def calculate_charges(self):
        """
        计算费用
        根据抄表数据自动计算所有租户的水电费
        """
        # 获取当前月份
        month = self.month_var.get()
        
        # 获取所有租户
        tenants = Tenant.get_all()
        
        # 计算每个租户的费用
        success_count = 0
        update_count = 0
        fail_count = 0
        
        for tenant in tenants:
            # 检查该租户是否已存在该月份的费用记录
            existing_charge = Charge.get_by_tenant_and_month(tenant.id, month)
            
            if self.calculate_tenant_charge(tenant, month):
                if existing_charge:
                    update_count += 1
                else:
                    success_count += 1
            else:
                fail_count += 1
        
        # 构建结果消息
        result_message = f"{self.get_text('charge_calculation_completed')}!\n"
        if success_count > 0:
            result_message += f"{self.get_text('new_added')}: {success_count} {self.get_text('households')}\n"
        if update_count > 0:
            result_message += f"{self.get_text('updated')}: {update_count} {self.get_text('households')}\n"
        if fail_count > 0:
            result_message += f"{self.get_text('failed')}: {fail_count} {self.get_text('households')}\n"
        
        # 显示计算结果
        messagebox.showinfo(self.get_text('calculation_result'), result_message)
        
        # 刷新费用列表
        self.load_charge_list()
    
    def calculate_tenant_charge(self, tenant, month):
        """
        计算单个租户的费用
        :param tenant: 租户对象
        :param month: 费用月份
        :return: 是否计算成功
        """
        try:
            print(f"开始计算租户 {tenant.name} 的{month}月份费用")
            
            # 获取租户的水电表
            meters = Meter.get_by_tenant(tenant.id)
            
            # 按类型分组水电表
            water_meters = [m for m in meters if m.meter_type == "水"]
            electricity_meters = [m for m in meters if m.meter_type == "电"]
            
            print(f"租户 {tenant.name} 有 {len(water_meters)} 个水表，{len(electricity_meters)} 个电表")
            
            # 计算水费用量（所有水表用量之和）
            water_usage = 0
            for meter in water_meters:
                # 获取该水表当月的抄表记录
                readings = MeterReading.get_by_month(month)
                meter_readings = [r for r in readings if r.meter_id == meter.id]
                if meter_readings:
                    water_usage += meter_readings[0].usage
                    print(f"水表 {meter.id} 的当月用量: {meter_readings[0].usage} 吨")
                else:
                    print(f"水表 {meter.id} 没有当月的抄表记录")
            
            # 计算电费用量（所有电表用量之和）
            electricity_usage = 0
            for meter in electricity_meters:
                # 获取该电表当月的抄表记录
                readings = MeterReading.get_by_month(month)
                meter_readings = [r for r in readings if r.meter_id == meter.id]
                if meter_readings:
                    electricity_usage += meter_readings[0].usage
                    print(f"电表 {meter.id} 的当月用量: {meter_readings[0].usage} 度")
                else:
                    print(f"电表 {meter.id} 没有当月的抄表记录")
            
            print(f"租户 {tenant.name} 的总用量: 水 {water_usage} 吨, 电 {electricity_usage} 度")
            
            # 使用Charge类的calculate_charge方法计算费用
            # 该方法会自动获取最新的水价和电价，基于租户类型
            charge = Charge.calculate_charge(tenant.id, month, water_usage, electricity_usage)
            
            if charge:
                # 保存费用记录
                if charge.save():
                    print(f"成功保存租户 {tenant.name} 的费用记录")
                    return True
                else:
                    print(f"保存租户 {tenant.name} 的费用记录失败")
                    return False
            else:
                print(f"计算租户 {tenant.name} 的费用失败")
                return False
        except Exception as e:
            print(f"计算租户 {tenant.name} 费用失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def export_charge_sheet(self):
        """
        导出收费表到Excel文件
        实现收费表的Excel导出功能，包含完整的数据和样式
        """
        try:
            # 1. 验证当前是否存在已计算的费用数据
            if not self.charge_list:
                messagebox.showwarning(self.get_text('warning'), self.get_text('no_charge_data_to_export'))
                return
            
            # 2. 显示加载提示
            messagebox.showinfo(self.get_text('info'), self.get_text('exporting_data_please_wait'))
            
            # 3. 准备数据
            # 获取当前月份
            current_month = self.month_var.get()
            
            # 获取所有租户信息
            tenants = Tenant.get_all()
            tenant_map = {t.id: t for t in tenants}
            
            # 获取当前月份的抄表记录
            readings = MeterReading.get_by_month(current_month)
            # 按租户和类型分组抄表记录
            reading_map = {}
            for reading in readings:
                meter = Meter.get_by_id(reading.meter_id)
                if meter:
                    key = (meter.tenant_id, meter.meter_type)
                    if key not in reading_map:
                        reading_map[key] = []
                    reading_map[key].append(reading)
            
            # 4. 创建Excel工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = "收费表"
            
            # 5. 设置列宽
            column_widths = {
                'A': 8,    # 序号
                'B': 12,   # 租户类型
                'C': 20,   # 租户名称
                'D': 10,   # 月份
                'E': 12,   # 上次读数（水）
                'F': 12,   # 当前读数（水）
                'G': 15,   # 用水量(m³)
                'H': 12,   # 用水单价(元)
                'I': 15,   # 水费金额(元)
                'J': 12,   # 上次读数（电）
                'K': 12,   # 当前读数（电）
                'L': 15,   # 用电量(kWh)
                'M': 12,   # 用电单价(元)
                'N': 15,   # 电费金额(元)
                'O': 15    # 总金额(元)
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 6. 设置样式
            # 标题样式
            title_font = Font(bold=True, size=12)
            title_alignment = Alignment(horizontal="center", vertical="center")
            # 使用简单的RGB颜色格式，不带井号和alpha通道
            title_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # 数据样式
            data_font = Font(size=11)
            data_alignment = Alignment(horizontal="left", vertical="center")
            amount_alignment = Alignment(horizontal="right", vertical="center")
            
            # 用水信息背景色 - 使用简单的RGB颜色格式
            water_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            # 用电信息背景色 - 使用简单的RGB颜色格式
            electricity_fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
            
            # 边框样式
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # 7. 添加标题行
            # 生成标题：YYYY年MM月份水电费收费表
            current_month = self.month_var.get()
            year, month = current_month.split('-')
            title_text = f"{year}年{month}月份水电费收费表"
            
            # 合并单元格作为标题
            ws.merge_cells(f'A1:O1')
            title_cell = ws['A1']
            title_cell.value = title_text
            
            # 设置标题样式
            title_font_large = Font(bold=True, size=14)  # 比数据大3号
            title_cell.font = title_font_large
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            title_cell.border = border
            
            # 隐藏D列、E列和F列
            ws.column_dimensions['D'].hidden = True
            ws.column_dimensions['E'].hidden = True
            ws.column_dimensions['F'].hidden = True
            
            # 8. 写入表头
            headers = [
                "序号", "租户类型", "租户名称", "月份",
                "上次读数", "当前读数", "用水量(m³)", "用水单价(元)", "水费金额(元)",
                "上次读数", "当前读数", "用电量(kWh)", "用电单价(元)", "电费金额(元)",
                "总金额(元)"
            ]
            
            ws.append(headers)
            
            # 设置表头样式
            for cell in ws[2]:
                cell.font = title_font
                cell.alignment = title_alignment
                cell.fill = title_fill
                cell.border = border
            
            # 9. 写入数据行并计算合计
            total_water_charge = 0
            total_electricity_charge = 0
            total_charge = 0
            
            for idx, charge in enumerate(self.charge_list, 1):
                tenant = tenant_map.get(charge.tenant_id)
                if not tenant:
                    continue
                
                # 获取用水抄表记录
                water_readings = reading_map.get((tenant.id, "水"), [])
                water_last_reading = sum(r.previous_reading for r in water_readings) if water_readings else 0
                water_current_reading = sum(r.current_reading for r in water_readings) if water_readings else 0
                
                # 获取用电抄表记录
                electricity_readings = reading_map.get((tenant.id, "电"), [])
                electricity_last_reading = sum(r.previous_reading for r in electricity_readings) if electricity_readings else 0
                electricity_current_reading = sum(r.current_reading for r in electricity_readings) if electricity_readings else 0
                
                # 构造数据行
                water_charge = round(charge.water_charge, 2)
                electricity_charge = round(charge.electricity_charge, 2)
                total_charge_item = round(charge.total_charge, 2)
                
                # 累加合计值
                total_water_charge += water_charge
                total_electricity_charge += electricity_charge
                total_charge += total_charge_item
                
                data_row = [
                    idx,
                    tenant.type,
                    tenant.name,
                    charge.month,
                    round(water_last_reading, 2),
                    round(water_current_reading, 2),
                    round(charge.water_usage, 2),
                    round(charge.water_price, 2),
                    water_charge,
                    round(electricity_last_reading, 2),
                    round(electricity_current_reading, 2),
                    round(charge.electricity_usage, 2),
                    round(charge.electricity_price, 2),
                    electricity_charge,
                    total_charge_item
                ]
                
                ws.append(data_row)
                
                # 设置数据行样式
                row = idx + 2  # 标题行占了一行，所以从3行开始
                for col_idx, cell in enumerate(ws[row], 1):
                    cell.font = data_font
                    cell.border = border
                    
                    # 设置对齐方式
                    if col_idx in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
                        cell.alignment = amount_alignment
                    else:
                        cell.alignment = data_alignment
                    
                    # 设置背景色
                    col_letter = get_column_letter(col_idx)
                    if col_letter in ['E', 'F', 'G', 'H', 'I']:
                        cell.fill = water_fill
                    elif col_letter in ['J', 'K', 'L', 'M', 'N']:
                        cell.fill = electricity_fill
            
            # 10. 添加合计行
            # 计算合计行的行号
            total_row_num = len(self.charge_list) + 3  # 标题行+表头行+数据行数
            
            # 构造合计行数据：首列显示"合计"，其他列留空或填写合计金额
            total_row = [
                "合计", "", "", "",  # 只在首列显示"合计"
                "", "", "", "", total_water_charge,  # 用水合计
                "", "", "", "", total_electricity_charge,  # 用电合计
                total_charge  # 总金额合计
            ]
            
            ws.append(total_row)
            
            # 11. 合并合计行的首列至第3列（A-C列）
            ws.merge_cells(f'A{total_row_num}:C{total_row_num}')
            
            # 12. 设置合计行样式
            # 设置合并后的单元格样式
            merged_cell = ws.cell(row=total_row_num, column=1)
            merged_cell.value = "合计"
            merged_cell.font = title_font_large
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            merged_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            merged_cell.border = border
            
            # 为合计行的所有列设置样式，确保边框完整
            for col_idx in range(1, 16):  # 1-15列
                # 跳过合并的列（1-3列已在merged_cell中处理）
                if col_idx == 1:
                    continue
                
                cell = ws.cell(row=total_row_num, column=col_idx)
                
                # 设置单元格样式
                cell.font = title_font_large
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.border = border
                
                # 根据列类型设置对齐方式
                if col_idx == 4:  # D列
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in [9, 14, 15]:  # 金额列
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:  # 其他列
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 13. 隐藏D列、E列和F列
            ws.column_dimensions['D'].hidden = True
            ws.column_dimensions['E'].hidden = True
            ws.column_dimensions['F'].hidden = True
            
            # 14. 设置打印区域
            ws.page_setup.print_area = f'A1:{get_column_letter(15)}{total_row_num}'
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paper_size = 9  # A4纸张大小
            
            # 10. 保存文件
            import os
            from tkinter import filedialog
            
            # 生成文件名
            current_date = datetime.now().strftime("%Y%m%d")
            default_filename = f"收费表_{current_date}.xlsx"
            
            # 打开文件保存对话框
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                title=self.get_text("export_charge_sheet"),
                initialfile=default_filename
            )
            
            if not file_path:
                return  # 用户取消了保存
            
            # 保存文件
            wb.save(file_path)
            
            # 11. 显示成功提示
            messagebox.showinfo(self.get_text('success'), f"{self.get_text('charge_sheet_exported_successfully')}\n{file_path}")
            
        except Exception as e:
            # 12. 显示错误提示
            messagebox.showerror(self.get_text('error'), f"{self.get_text('export_failed')}: {str(e)}")
            print(f"导出收费表失败: {e}")
            import traceback
            traceback.print_exc()