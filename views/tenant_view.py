#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
租户管理视图
负责处理租户信息的添加、编辑、删除和查询
"""

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from models.tenant import Tenant
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import datetime
from utils.language_utils import LanguageUtils

class TenantView:
    """租户管理视图类"""
    
    def __init__(self, parent, main_window=None, language_utils=None):
        """
        初始化租户管理视图
        :param parent: 父窗口组件
        :param main_window: 主窗口实例，用于视图间通信
        :param language_utils: 语言工具类实例
        """
        self.parent = parent
        self.main_window = main_window
        self.tenant_list = []
        self.selected_tenant = None
        # 使用传入的语言工具或创建新实例
        self.language_utils = language_utils if language_utils else LanguageUtils()
        self.create_widgets()
        self.load_tenant_list()
    
    def get_text(self, key):
        """
        获取当前语言的文本
        :param key: 文本键名
        :return: 对应语言的文本
        """
        return self.language_utils.get_text(key)
    
    def update_language(self):
        """
        更新语言
        当系统语言切换时调用，更新界面上所有文本
        """
        # 更新搜索框标签和按钮
        search_frame = self.parent.winfo_children()[0].winfo_children()[0]
        search_children = search_frame.winfo_children()
        
        # 更新标签
        if len(search_children) > 0 and isinstance(search_children[0], ttk.Label):
            search_children[0].config(text=self.get_text("tenant_name")+":")
        if len(search_children) > 2 and isinstance(search_children[2], ttk.Label):
            search_children[2].config(text=self.get_text("tenant_type")+":")
        
        # 更新搜索角色下拉框选项
        self.search_type.config(values=["", self.get_text("office") if "办公室" in self.search_type['values'] else "", self.get_text("storefront") if "门面" in self.search_type['values'] else ""])
        
        # 更新按钮
        button_texts = ["button_search", "button_reset", "export_tenant", "import_tenant"]
        for i, btn_idx in enumerate([4, 5, 6, 7]):
            if i < len(button_texts) and btn_idx < len(search_children) and isinstance(search_children[btn_idx], ttk.Button):
                search_children[btn_idx].config(text=self.get_text(button_texts[i]))
        
        # 更新列表按钮
        left_frame = self.parent.winfo_children()[0].winfo_children()[1]
        button_frame = left_frame.winfo_children()[1] if len(left_frame.winfo_children()) > 1 else None
        if button_frame:
            button_children = button_frame.winfo_children()
            button_texts = ["add_tenant", "edit_tenant", "delete_tenant", "refresh_list"]
            for i, btn in enumerate(button_children):
                if i < len(button_texts) and isinstance(btn, ttk.Button):
                    btn.config(text=self.get_text(button_texts[i]))
        
        # 更新列标题
        column_texts = {
            "serial": "serial",
            "name": "tenant_name",
            "type": "tenant_type",
            "contact_person": "contact_person",
            "phone": "phone",
            "deactivated": "deactivated"
        }
        
        for col, key in column_texts.items():
            self.tenant_tree.heading(col, text=self.get_text(key))
        
        # 更新表单标题
        main_frame = self.parent.winfo_children()[0]
        right_frame = main_frame.winfo_children()[1] if len(main_frame.winfo_children()) > 1 else None
        if right_frame:
            form_title = right_frame.winfo_children()[0] if isinstance(right_frame.winfo_children()[0], ttk.Label) else None
            if form_title:
                form_title.config(text=self.get_text("tenant_details"))
        
        # 更新表单标签和选项
        if right_frame:
            form_frame = right_frame.winfo_children()[1] if len(right_frame.winfo_children()) > 1 else None
            if form_frame:
                form_children = form_frame.winfo_children()
                label_keys = ["tenant_name", "tenant_type", "address", "contact_person", "phone", "email", "deactivated"]
                for i, label_idx in enumerate([0, 2, 4, 6, 8, 10, 12]):
                    if i < len(label_keys) and label_idx < len(form_children) and isinstance(form_children[label_idx], ttk.Label):
                        form_children[label_idx].config(text=self.get_text(label_keys[i])+":")
                
                # 更新租户类型下拉框
                self.type_combobox.config(values=[self.get_text("office"), self.get_text("storefront")])
                
                # 更新按钮
                button_texts = ["button_save", "button_cancel"]
                for i, btn_idx in enumerate([14, 15]):
                    if i < len(button_texts) and btn_idx < len(form_children) and isinstance(form_children[btn_idx], ttk.Button):
                        form_children[btn_idx].config(text=self.get_text(button_texts[i]))
        
        # 重新加载数据，确保显示当前语言
        self.load_tenant_list()
    
    def create_widgets(self):
        """
        创建租户管理界面组件
        """
        # 创建主框架
        main_frame = ttk.Frame(self.parent)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 左侧：租户列表和搜索框
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 搜索框
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        
        ttk.Label(search_frame, text=self.get_text("tenant_name")+":").pack(side=tk.LEFT, padx=5)
        self.search_name = ttk.Entry(search_frame)
        self.search_name.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        ttk.Label(search_frame, text=self.get_text("tenant_type")+":").pack(side=tk.LEFT, padx=5)
        self.search_type = ttk.Combobox(search_frame, values=["", self.get_text("office"), self.get_text("storefront")])
        self.search_type.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(search_frame, text=self.get_text("button_search"), command=self.search_tenants).pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text=self.get_text("button_reset"), command=self.reset_search).pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text=self.get_text("export_tenant"), command=self.export_tenants).pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text=self.get_text("import_tenant"), command=self.import_tenants).pack(side=tk.LEFT, padx=5)
        
        # 租户列表
        list_frame = ttk.Frame(left_frame)
        list_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 列表控件 - 修改列定义，添加序号列，移除ID列显示
        columns = ("serial", "name", "type", "contact_person", "phone", "deactivated")
        self.tenant_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        
        # 设置列标题和宽度
        self.tenant_tree.heading("serial", text=self.get_text("serial"))
        self.tenant_tree.column("serial", width=80, anchor="center")
        
        self.tenant_tree.heading("name", text=self.get_text("tenant_name"))
        self.tenant_tree.column("name", width=150, anchor="w")
        
        self.tenant_tree.heading("type", text=self.get_text("tenant_type"))
        self.tenant_tree.column("type", width=100, anchor="center")
        
        self.tenant_tree.heading("contact_person", text=self.get_text("contact_person"))
        self.tenant_tree.column("contact_person", width=100, anchor="w")
        
        self.tenant_tree.heading("phone", text=self.get_text("phone"))
        self.tenant_tree.column("phone", width=120, anchor="w")
        
        self.tenant_tree.heading("deactivated", text=self.get_text("deactivated"))
        self.tenant_tree.column("deactivated", width=80, anchor="center")
        
        # 保存排序状态
        self.sort_column = None
        self.sort_order = "asc"
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tenant_tree.yview)
        self.tenant_tree.configure(yscroll=scrollbar.set)
        
        self.tenant_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 列表操作按钮
        button_frame = ttk.Frame(left_frame)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=5)
        
        ttk.Button(button_frame, text=self.get_text("add_tenant"), command=self.add_tenant).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.get_text("edit_tenant"), command=self.edit_tenant).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.get_text("delete_tenant"), command=self.delete_tenant).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.get_text("refresh_list"), command=self.load_tenant_list).pack(side=tk.LEFT, padx=5)
        
        # 右侧：租户详细信息
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 详细信息标题
        ttk.Label(right_frame, text=self.get_text("tenant_details"), font=("Arial", 14)).pack(side=tk.TOP, pady=10)
        
        # 详细信息表单
        form_frame = ttk.Frame(right_frame)
        form_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 表单控件
        row = 0
        
        # 租户名称
        ttk.Label(form_frame, text=self.get_text("tenant_name")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.name_entry = ttk.Entry(form_frame)
        self.name_entry.grid(row=row, column=1, columnspan=2, sticky=tk.W+tk.E, padx=5, pady=5)
        row += 1
        
        # 租户类型
        ttk.Label(form_frame, text=self.get_text("tenant_type")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.type_combobox = ttk.Combobox(form_frame, values=[self.get_text("office"), self.get_text("storefront")])
        self.type_combobox.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 地址
        ttk.Label(form_frame, text=self.get_text("address")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.address_entry = ttk.Entry(form_frame)
        self.address_entry.grid(row=row, column=1, columnspan=2, sticky=tk.W+tk.E, padx=5, pady=5)
        row += 1
        
        # 联系人
        ttk.Label(form_frame, text=self.get_text("contact_person")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.contact_person_entry = ttk.Entry(form_frame)
        self.contact_person_entry.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 联系电话
        ttk.Label(form_frame, text=self.get_text("phone")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.phone_entry = ttk.Entry(form_frame)
        self.phone_entry.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 邮箱
        ttk.Label(form_frame, text=self.get_text("email")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.email_entry = ttk.Entry(form_frame)
        self.email_entry.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 停用状态
        ttk.Label(form_frame, text=self.get_text("deactivated")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.deactivated_var = tk.BooleanVar()
        self.deactivated_check = ttk.Checkbutton(form_frame, variable=self.deactivated_var)
        self.deactivated_check.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 详细信息操作按钮
        self.save_button = ttk.Button(form_frame, text=self.get_text("button_save"), command=self.save_tenant)
        self.save_button.grid(row=row, column=1, sticky=tk.W, padx=5, pady=10)
        
        self.cancel_button = ttk.Button(form_frame, text=self.get_text("button_cancel"), command=self.clear_form)
        self.cancel_button.grid(row=row, column=2, sticky=tk.W, padx=5, pady=10)
        
        # 绑定列表选择事件
        self.tenant_tree.bind("<<TreeviewSelect>>", self.on_tenant_select)
    
    def load_tenant_list(self, filters=None):
        """
        加载租户列表
        :param filters: 过滤条件
        """
        # 清空现有列表
        for item in self.tenant_tree.get_children():
            self.tenant_tree.delete(item)
        
        try:
            # 获取租户列表
            self.tenant_list = Tenant.get_all(filters)
            
            # 应用默认排序规则：先按租户类型升序，再按租户名称升序
            self.tenant_list.sort(key=lambda x: (x.type, x.name))
            
            # 添加到列表控件，添加序号列
            for idx, tenant in enumerate(self.tenant_list, 1):
                # 保存ID到item的values中，但不显示在表格中
                deactivated_text = self.get_text("yes") if tenant.deactivated else self.get_text("no")
                tenant_type_text = self.get_text("office") if tenant.type == "办公室" else self.get_text("storefront")
                self.tenant_tree.insert("", tk.END, values=(idx, tenant.name, tenant_type_text, 
                                                      tenant.contact_person, tenant.phone, deactivated_text), tags=(str(tenant.id),))
        except Exception as e:
            # 添加错误处理，处理数据加载失败的情况
            messagebox.showerror(self.get_text("error"), f"{self.get_text('tenant_list_load_fail')}{str(e)}")
            # 恢复到空列表状态
            self.tenant_list = []
    
    def search_tenants(self):
        """
        搜索租户
        """
        # 获取搜索条件
        name = self.search_name.get()
        type = self.search_type.get()
        
        # 构建过滤条件
        filters = {}
        if name:
            filters['name'] = name
        if type:
            filters['type'] = type
        
        # 加载过滤后的列表
        self.load_tenant_list(filters)
    
    def reset_search(self):
        """
        重置搜索条件
        """
        self.search_name.delete(0, tk.END)
        self.search_type.set("")
        self.load_tenant_list()
    
    def on_tenant_select(self, event):
        """
        列表选择事件处理
        """
        # 获取选中的项目
        selected_items = self.tenant_tree.selection()
        if selected_items:
            # 获取选中项目的ID
            item_id = selected_items[0]
            
            # 从tags中获取租户ID
            tags = self.tenant_tree.item(item_id, "tags")
            if tags:
                tenant_id = int(tags[0])
            else:
                # 如果没有tags，尝试从列表中匹配
                values = self.tenant_tree.item(item_id, "values")
                if values:
                    name, type, _, _ = values[1:]
                    # 通过多字段匹配查找租户
                    self.selected_tenant = next((t for t in self.tenant_list if t.name == name and t.type == type), None)
                    return
                else:
                    messagebox.showwarning(self.get_text("warning"), self.get_text("cannot_get_tenant_id"))
                    return
            
            # 查找对应的租户对象
            # 首先尝试从当前租户列表中查找
            self.selected_tenant = next((t for t in self.tenant_list if t.id == tenant_id), None)
            
            # 如果当前列表中找不到（可能是因为搜索后的列表没有包含完整的租户对象信息），则重新从数据库获取
            if not self.selected_tenant:
                self.selected_tenant = Tenant.get_by_id(tenant_id)
            
            if self.selected_tenant:
                self.fill_form()
            else:
                messagebox.showwarning(self.get_text("warning"), self.get_text("cannot_find_tenant_info"))
        else:
            # 取消选择时，清空表单和选择状态
            self.clear_form()
    
    def fill_form(self):
        """
        填充表单
        """
        if self.selected_tenant:
            self.name_entry.delete(0, tk.END)
            self.name_entry.insert(0, self.selected_tenant.name or "")
            
            self.type_combobox.set(self.selected_tenant.type or "")
            
            self.address_entry.delete(0, tk.END)
            self.address_entry.insert(0, self.selected_tenant.address or "")
            
            self.contact_person_entry.delete(0, tk.END)
            self.contact_person_entry.insert(0, self.selected_tenant.contact_person or "")
            
            self.phone_entry.delete(0, tk.END)
            self.phone_entry.insert(0, self.selected_tenant.phone or "")
            
            self.email_entry.delete(0, tk.END)
            self.email_entry.insert(0, self.selected_tenant.email or "")
            
            self.deactivated_var.set(self.selected_tenant.deactivated)
    
    def clear_form(self):
        """
        清空表单
        """
        self.selected_tenant = None
        
        self.name_entry.delete(0, tk.END)
        self.type_combobox.set("")
        self.address_entry.delete(0, tk.END)
        self.contact_person_entry.delete(0, tk.END)
        self.phone_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.deactivated_var.set(False)
    
    def add_tenant(self):
        """
        添加租户
        """
        self.clear_form()
        # 可以添加一些提示信息
    
    def edit_tenant(self):
        """
        编辑租户
        """
        if not self.selected_tenant:
            messagebox.showwarning(self.get_text("warning"), self.get_text("please_select_tenant_edit"))
            return
        # 表单已经填充，直接编辑即可
    
    def delete_tenant(self):
        """
        删除租户
        """
        if not self.selected_tenant:
            messagebox.showwarning(self.get_text("warning"), self.get_text("please_select_tenant_delete"))
            return
        
        # 显示更详细的确认对话框，包含要删除的租户信息
        confirm_message = self.get_text("confirm_delete_tenant").format(
        self.selected_tenant.name, 
        self.selected_tenant.type, 
        self.selected_tenant.contact_person, 
        self.selected_tenant.phone
        )
        
        if messagebox.askyesno(self.get_text("system_confirm_delete"), confirm_message):
            try:
                # 执行删除操作
                if self.selected_tenant.delete():
                    messagebox.showinfo(self.get_text("success"), self.get_text("tenant_delete_success"))
                    # 刷新列表
                    self.load_tenant_list()
                    # 清空表单
                    self.clear_form()
                    # 通知主窗口刷新视图
                    if self.main_window:
                        self.main_window.refresh_view("tenant")
                else:
                    # 删除失败，检查是否是因为存在关联的抄表记录
                    # 1. 先查询该租户关联的所有水电表
                    from models.meter import Meter
                    meters = Meter.get_by_tenant(self.selected_tenant.id)
                    
                    if meters:
                        # 2. 检查是否有抄表记录
                        has_readings = False
                        from models.reading import MeterReading
                        for meter in meters:
                            # 获取该水电表的抄表记录
                            readings = MeterReading.get_by_meter(meter.id, limit=1)
                            if readings:
                                has_readings = True
                                break
                        
                        if has_readings:
                            messagebox.showerror(self.get_text("error"), self.get_text("tenant_has_readings"))
                        else:
                            messagebox.showerror(self.get_text("error"), self.get_text("tenant_delete_fail"))
                    else:
                        messagebox.showerror(self.get_text("error"), self.get_text("tenant_delete_fail"))
            except Exception as e:
                # 添加必要的错误处理机制
                messagebox.showerror(self.get_text("error"), f"{self.get_text('delete_exception').format(str(e))}")
    
    def save_tenant(self):
        """
        保存租户信息
        """
        # 获取表单数据
        name = self.name_entry.get().strip()
        type = self.type_combobox.get().strip()
        address = self.address_entry.get().strip()
        contact_person = self.contact_person_entry.get().strip()
        phone = self.phone_entry.get().strip()
        email = self.email_entry.get().strip()
        
        # 验证数据
        if not name:
            messagebox.showwarning(self.get_text("warning"), self.get_text("please_enter_tenant_name"))
            return
        
        if not type:
            messagebox.showwarning(self.get_text("warning"), self.get_text("please_select_tenant_type"))
            return
        
        if not contact_person:
            messagebox.showwarning(self.get_text("warning"), self.get_text("please_enter_contact_person"))
            return
        
        if not phone:
            messagebox.showwarning(self.get_text("warning"), self.get_text("please_enter_phone"))
            return
        
        # 保存租户信息
        try:
            if self.selected_tenant:
                # 更新现有租户
                self.selected_tenant.name = name
                self.selected_tenant.type = type
                self.selected_tenant.address = address
                self.selected_tenant.contact_person = contact_person
                self.selected_tenant.phone = phone
                self.selected_tenant.email = email
                self.selected_tenant.deactivated = self.deactivated_var.get()
                
                if self.selected_tenant.save():
                    messagebox.showinfo(self.get_text("success"), self.get_text("tenant_update_success"))
                    self.load_tenant_list()
                    self.clear_form()
                    # 通知主窗口刷新视图
                    if self.main_window:
                        self.main_window.refresh_view("tenant")
                else:
                    messagebox.showerror(self.get_text("error"), self.get_text("tenant_update_fail"))
            else:
                # 添加新租户
                tenant = Tenant(name=name, type=type, address=address, 
                               contact_person=contact_person, phone=phone, email=email, 
                               deactivated=self.deactivated_var.get())
                
                if tenant.save():
                    messagebox.showinfo(self.get_text("success"), self.get_text("tenant_add_success"))
                    self.load_tenant_list()
                    self.clear_form()
                    # 通知主窗口刷新视图
                    if self.main_window:
                        self.main_window.refresh_view("tenant")
                else:
                    messagebox.showerror(self.get_text("error"), self.get_text("tenant_add_fail"))
        except Exception as e:
            # 添加必要的错误处理机制
            messagebox.showerror(self.get_text("error"), f"{self.get_text('save_tenant_exception').format(str(e))}")
    
    def export_tenants(self):
        """
        导出租户信息
        支持导出当前筛选条件下的所有租户信息，文件格式为Excel（.xlsx）
        """
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = self.get_text("tenant_information")
            
            # 设置表头
            headers = [self.get_text("tenant_id"), self.get_text("tenant_name"), self.get_text("tenant_type"), self.get_text("contact_person"), self.get_text("phone"), self.get_text("email"), self.get_text("address"), self.get_text("create_time")]
            ws.append(headers)
            
            # 设置表头样式
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # 获取列字母
                # 设置表头字体和对齐方式
                ws[column + "1"].font = Font(bold=True, size=12)
                ws[column + "1"].alignment = Alignment(horizontal="center")
                # 计算列宽
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
                ws.column_dimensions[column].width = adjusted_width
            
            # 获取当前筛选条件下的租户列表
            name = self.search_name.get()
            type = self.search_type.get()
            filters = {}
            if name:
                filters['name'] = name
            if type:
                filters['type'] = type
            tenant_list = Tenant.get_all(filters)
            
            # 写入数据
            for tenant in tenant_list:
                ws.append([
                    tenant.id,
                    tenant.name,
                    tenant.type,
                    tenant.contact_person,
                    tenant.phone,
                    tenant.email,
                    tenant.address,
                    tenant.create_time if tenant.create_time else ""
                ])
            
            # 保存文件
            default_filename = f"租户信息_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                initialfile=default_filename
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo(self.get_text("success"), f"{self.get_text('export_tenant_success').format(file_path)}")
        except Exception as e:
            messagebox.showerror(self.get_text("error"), f"{self.get_text('export_tenant_fail').format(str(e))}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def import_tenants(self):
        """
        导入租户信息
        支持导入Excel（.xlsx）格式文件，包含数据验证和进度显示
        """
        try:
            # 选择文件
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                title=self.get_text("select_tenant_excel_file")
            )
            
            if not file_path:
                return
            
            # 加载文件
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 验证表头
            expected_headers = [self.get_text("tenant_name"), self.get_text("tenant_type"), self.get_text("contact_person"), self.get_text("phone"), self.get_text("email"), self.get_text("address")]
            actual_headers = [cell.value for cell in ws[1]]
            
            # 检查是否包含所有必填表头
            missing_headers = [header for header in expected_headers if header not in actual_headers]
            if missing_headers:
                messagebox.showerror(self.get_text("error"), f"{self.get_text('file_format_error').format(', '.join(missing_headers))}")
                return
            
            # 创建进度窗口
            progress_window = tk.Toplevel(self.parent)
            progress_window.title(self.get_text("import_progress"))
            progress_window.geometry("400x150")
            progress_window.transient(self.parent)
            progress_window.grab_set()
            
            # 进度标签
            progress_label = ttk.Label(progress_window, text=self.get_text("preparing_import"))
            progress_label.pack(pady=20)
            
            # 进度条
            progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate')
            progress_bar.pack(pady=10)
            
            # 结果标签
            result_label = ttk.Label(progress_window, text="")
            result_label.pack(pady=10)
            
            progress_window.update()
            
            # 初始化统计信息
            total_rows = ws.max_row - 1  # 减去表头
            success_count = 0
            failed_count = 0
            failed_reasons = []
            
            # 解析数据
            for row_idx in range(2, ws.max_row + 1):
                row_data = ws[row_idx]
                try:
                    # 获取数据
                    name = row_data[actual_headers.index("租户名称")].value
                    type = row_data[actual_headers.index("租户类型")].value
                    contact_person = row_data[actual_headers.index("联系人")].value
                    phone = row_data[actual_headers.index("联系电话")].value
                    email = row_data[actual_headers.index("邮箱")].value if "邮箱" in actual_headers else ""
                    address = row_data[actual_headers.index("地址")].value if "地址" in actual_headers else ""
                    
                    # 数据验证
                    if not name:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：租户名称不能为空")
                        continue
                    
                    if not type or type not in ["办公室", "门面"]:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：租户类型必须是'办公室'或'门面'")
                        continue
                    
                    if not contact_person:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：联系人不能为空")
                        continue
                    
                    if not phone:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：联系电话不能为空")
                        continue
                    
                    # 检查是否已存在相同名称的租户
                    existing_tenants = Tenant.get_all()
                    if any(t.name == name for t in existing_tenants):
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：租户 '{name}' 已存在")
                        continue
                    
                    # 创建并保存租户
                    tenant = Tenant(
                        name=name,
                        type=type,
                        contact_person=contact_person,
                        phone=phone,
                        email=email,
                        address=address
                    )
                    
                    if tenant.save():
                        success_count += 1
                    else:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：保存失败")
                        continue
                    
                except Exception as e:
                    failed_count += 1
                    failed_reasons.append(f"第{row_idx}行：{str(e)}")
                
                # 更新进度
            progress = int((row_idx - 1) / total_rows * 100)
            progress_bar['value'] = progress
            progress_label.config(text=f"{self.get_text('importing_row').format(row_idx - 1, total_rows)}")
            result_label.config(text=f"{self.get_text('success')}: {success_count}, {self.get_text('fail')}: {failed_count}")
            progress_window.update()
            
            # 关闭进度窗口
            progress_window.destroy()
            
            # 显示导入结果
            result_message = f"{self.get_text('import_completed').format(total_rows, success_count, failed_count)}"
            
            if failed_reasons:
                result_message += f"\n{self.get_text('failed_reasons')}:\n"
                result_message += "\n".join(failed_reasons[:10])  # 只显示前10条失败原因
                if len(failed_reasons) > 10:
                    result_message += f"\n{self.get_text('more_failed_reasons').format(len(failed_reasons) - 10)}"
            
            messagebox.showinfo(self.get_text('import_result'), result_message)
            
            # 刷新租户列表
            self.load_tenant_list()
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('import_tenant_fail').format(str(e))}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
