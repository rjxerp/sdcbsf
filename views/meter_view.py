#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
水电表管理视图
负责处理水电表的添加、编辑、删除和查询功能
"""

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from models.tenant import Tenant
from models.meter import Meter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import datetime
from utils.language_utils import LanguageUtils

class MeterView:
    """水电表管理视图类"""
    
    def __init__(self, parent, language_utils=None):
        """
        初始化水电表管理视图
        :param parent: 父窗口组件
        :param language_utils: 语言工具类实例
        """
        self.parent = parent
        self.meter_list = []
        self.selected_meter = None
        # 使用传入的语言工具或创建新实例
        self.language_utils = language_utils if language_utils else LanguageUtils()
        self.create_widgets()
        # 初始加载数据
        self.load_meter_list()
    
    def get_text(self, key):
        """
        获取当前语言的文本
        :param key: 文本键名
        :return: 对应语言的文本
        """
        return self.language_utils.get_text(key)
    
    def update_language(self):
        """
        更新语言
        当系统语言切换时调用，更新界面上所有文本
        """
        # 更新搜索框标签和按钮
        search_frame = self.parent.winfo_children()[0].winfo_children()[0]
        search_children = search_frame.winfo_children()
        
        # 更新标签
        if len(search_children) > 0 and isinstance(search_children[0], ttk.Label):
            search_children[0].config(text=self.get_text("meter_no")+":")
        if len(search_children) > 2 and isinstance(search_children[2], ttk.Label):
            search_children[2].config(text=self.get_text("meter_type")+":")
        
        # 更新搜索类型下拉框选项
        self.search_meter_type.config(values=["", self.get_text("water_meter"), self.get_text("electric_meter")])
        
        # 更新按钮
        button_texts = ["button_search", "button_reset", "export_meter", "import_meter"]
        for i, btn_idx in enumerate([4, 5, 6, 7]):
            if i < len(button_texts) and btn_idx < len(search_children) and isinstance(search_children[btn_idx], ttk.Button):
                search_children[btn_idx].config(text=self.get_text(button_texts[i]))
        
        # 更新列表按钮
        left_frame = self.parent.winfo_children()[0].winfo_children()[1]
        button_frame = left_frame.winfo_children()[1] if len(left_frame.winfo_children()) > 1 else None
        if button_frame:
            button_children = button_frame.winfo_children()
            button_texts = ["add_meter", "edit_meter", "delete_meter", "refresh_list"]
            for i, btn in enumerate(button_children):
                if i < len(button_texts) and isinstance(btn, ttk.Button):
                    btn.config(text=self.get_text(button_texts[i]))
        
        # 更新列标题
        column_texts = {
            "serial": "serial",
            "meter_no": "meter_no",
            "meter_type": "meter_type",
            "tenant_name": "tenant_name",
            "location": "location",
            "initial_reading": "initial_reading",
            "status": "status"
        }
        
        for col, key in column_texts.items():
            self.meter_tree.heading(col, text=self.get_text(key))
        
        # 更新表单标题
        main_frame = self.parent.winfo_children()[0]
        right_frame = main_frame.winfo_children()[1] if len(main_frame.winfo_children()) > 1 else None
        if right_frame:
            form_title = right_frame.winfo_children()[0] if isinstance(right_frame.winfo_children()[0], ttk.Label) else None
            if form_title:
                form_title.config(text=self.get_text("meter_details"))
        
        # 更新表单标签和选项
        if right_frame:
            form_frame = right_frame.winfo_children()[1] if len(right_frame.winfo_children()) > 1 else None
            if form_frame:
                form_children = form_frame.winfo_children()
                label_keys = ["meter_no", "meter_type", "tenant", "location", "initial_reading", "status"]
                for i, label_idx in enumerate([0, 2, 4, 6, 8, 10]):
                    if i < len(label_keys) and label_idx < len(form_children) and isinstance(form_children[label_idx], ttk.Label):
                        form_children[label_idx].config(text=self.get_text(label_keys[i])+":")
                
                # 更新表单下拉框选项
                self.form_meter_type.config(values=[self.get_text("water_meter"), self.get_text("electric_meter")])
                self.form_status.config(values=[self.get_text("normal"), self.get_text("damaged"), self.get_text("replaced")])
                
                # 更新按钮
                if len(form_children) > 12 and isinstance(form_children[12], ttk.Button):
                    form_children[12].config(text=self.get_text("button_save"))
                if len(form_children) > 13 and isinstance(form_children[13], ttk.Button):
                    form_children[13].config(text=self.get_text("button_cancel"))
        
        # 重新加载数据，确保显示当前语言
        self.load_meter_list()
    
    def create_widgets(self):
        """
        创建水电表管理界面组件
        """
        # 创建主框架
        main_frame = ttk.Frame(self.parent)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 左侧：水电表列表和搜索框
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 搜索框
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        
        ttk.Label(search_frame, text=self.get_text("meter_no")+":").pack(side=tk.LEFT, padx=5)
        self.search_meter_no = ttk.Entry(search_frame)
        self.search_meter_no.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        ttk.Label(search_frame, text=self.get_text("meter_type")+":").pack(side=tk.LEFT, padx=5)
        self.search_meter_type = ttk.Combobox(search_frame, values=["", self.get_text("water_meter"), self.get_text("electric_meter")])
        self.search_meter_type.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(search_frame, text=self.get_text("button_search"), command=self.search_meters).pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text=self.get_text("button_reset"), command=self.reset_search).pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text=self.get_text("export_meter"), command=self.export_meters).pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text=self.get_text("import_meter"), command=self.import_meters).pack(side=tk.LEFT, padx=5)
        
        # 水电表列表
        list_frame = ttk.Frame(left_frame)
        list_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 列表控件 - 修改列定义，添加序号列，移除ID列显示
        columns = ("serial", "meter_no", "meter_type", "tenant_name", "location", "initial_reading", "status")
        self.meter_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        
        # 设置列标题和宽度
        self.meter_tree.heading("serial", text=self.get_text("serial"))
        self.meter_tree.column("serial", width=80, anchor="center")
        
        self.meter_tree.heading("meter_no", text=self.get_text("meter_no"))
        self.meter_tree.column("meter_no", width=120, anchor="w")
        
        self.meter_tree.heading("meter_type", text=self.get_text("meter_type"))
        self.meter_tree.column("meter_type", width=80, anchor="center")
        
        self.meter_tree.heading("tenant_name", text=self.get_text("tenant_name"))
        self.meter_tree.column("tenant_name", width=150, anchor="w")
        
        self.meter_tree.heading("location", text=self.get_text("location"))
        self.meter_tree.column("location", width=150, anchor="w")
        
        self.meter_tree.heading("initial_reading", text=self.get_text("initial_reading"))
        self.meter_tree.column("initial_reading", width=100, anchor="e")
        
        self.meter_tree.heading("status", text=self.get_text("status"))
        self.meter_tree.column("status", width=80, anchor="center")
        
        # 保存排序状态
        self.sort_column = None
        self.sort_order = "asc"
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.meter_tree.yview)
        self.meter_tree.configure(yscroll=scrollbar.set)
        
        self.meter_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 列表操作按钮
        button_frame = ttk.Frame(left_frame)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=5)
        
        ttk.Button(button_frame, text=self.get_text("add_meter"), command=self.add_meter).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.get_text("edit_meter"), command=self.edit_meter).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.get_text("delete_meter"), command=self.delete_meter).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text=self.get_text("refresh_list"), command=self.load_meter_list).pack(side=tk.LEFT, padx=5)
        
        # 右侧：水电表明细表单
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 表单标题
        ttk.Label(right_frame, text=self.get_text("meter_details"), font=("Arial", 14)).pack(side=tk.TOP, pady=10)
        
        # 表单框架
        form_frame = ttk.Frame(right_frame)
        form_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 表单控件
        row = 0
        
        # 表编号
        ttk.Label(form_frame, text=self.get_text("meter_no")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.form_meter_no = ttk.Entry(form_frame)
        self.form_meter_no.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 表类型
        ttk.Label(form_frame, text=self.get_text("meter_type")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.form_meter_type = ttk.Combobox(form_frame, values=[self.get_text("water_meter"), self.get_text("electric_meter")])
        self.form_meter_type.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 所属租户
        ttk.Label(form_frame, text=self.get_text("tenant")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.form_tenant = ttk.Combobox(form_frame, values=[])
        self.load_tenants_to_form()
        self.form_tenant.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 安装位置
        ttk.Label(form_frame, text=self.get_text("location")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.form_location = ttk.Entry(form_frame)
        self.form_location.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        row += 1
        
        # 初始读数
        ttk.Label(form_frame, text=self.get_text("initial_reading")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.form_initial_reading = ttk.Entry(form_frame)
        self.form_initial_reading.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        self.form_initial_reading.insert(0, "0")
        row += 1
        
        # 状态
        ttk.Label(form_frame, text=self.get_text("status")+":").grid(row=row, column=0, sticky=tk.E, padx=5, pady=5)
        self.form_status = ttk.Combobox(form_frame, values=[self.get_text("normal"), self.get_text("damaged"), self.get_text("replaced")])
        self.form_status.grid(row=row, column=1, sticky=tk.W, padx=5, pady=5)
        self.form_status.set(self.get_text("normal"))
        row += 1
        
        # 表单操作按钮
        self.save_button = ttk.Button(form_frame, text=self.get_text("button_save"), command=self.save_meter)
        self.save_button.grid(row=row, column=0, sticky=tk.E, padx=5, pady=10)
        
        self.cancel_button = ttk.Button(form_frame, text=self.get_text("button_cancel"), command=self.clear_form)
        self.cancel_button.grid(row=row, column=1, sticky=tk.W, padx=5, pady=10)
        
        # 绑定列表选择事件
        self.meter_tree.bind("<<TreeviewSelect>>", self.on_meter_select)
    
    def load_tenants_to_form(self):
        """
        加载租户数据到表单下拉框
        每次调用都会重新从数据库获取最新的租户列表
        """
        try:
            # 从数据库获取最新的租户列表
            tenants = Tenant.get_all()
            # 更新租户字典，用于将租户名称转换为租户对象
            self.tenants_dict = {t.name: t for t in tenants}
            # 更新下拉框选项
            self.form_tenant['values'] = [t.name for t in tenants]
        except Exception as e:
            messagebox.showerror(self.get_text("error"), f"{self.get_text('load_tenant_list_exception').format(str(e))}")
    
    def search_meters(self):
        """
        搜索水电表
        """
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            # 获取搜索条件
            meter_no = self.search_meter_no.get()
            meter_type = self.search_meter_type.get()
            
            # 构建过滤条件
            filters = {}
            if meter_type:
                filters['meter_type'] = meter_type
            
            # 加载过滤后的列表
            self.meter_list = Meter.get_all(filters)
            
            # 清空现有列表
            for item in self.meter_tree.get_children():
                self.meter_tree.delete(item)
            
            # 创建租户ID到名称的映射
            tenants = Tenant.get_all()
            tenant_map = {t.id: t.name for t in tenants}
            
            # 应用默认排序规则：按所属租户、表编号、表类型升序排列
            # 创建租户ID到名称的映射，用于排序
            tenant_map_for_sort = {t.id: t.name for t in tenants}
            # 排序函数
            def meter_sort_key(meter):
                return (
                    tenant_map_for_sort.get(meter.tenant_id, "未知租户"),  # 首要：所属租户
                    meter.meter_no,  # 次要：表编号
                    meter.meter_type  # 第三：表类型
                )
            # 执行排序
            self.meter_list.sort(key=meter_sort_key)
            
            # 添加到列表控件，添加序号列
            search_count = 0
            for idx, meter in enumerate(self.meter_list, 1):
                # 如果指定了表编号，进行过滤
                if meter_no and meter_no not in meter.meter_no:
                    continue
                tenant_name = tenant_map.get(meter.tenant_id, "未知租户")
                self.meter_tree.insert("", tk.END, values=(idx, meter.meter_no, 
                                                        meter.meter_type, 
                                                        tenant_name, 
                                                        meter.location, 
                                                        meter.initial_reading, 
                                                        meter.status), tags=(str(meter.id),))
                search_count += 1
            
            # 显示搜索结果数量
            if meter_no or meter_type:
                self.show_temporary_message(f"{self.get_text('search_completed').format(search_count)}")
        except Exception as e:
            messagebox.showerror(self.get_text("error"), f"{self.get_text('search_meter_exception').format(str(e))}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def reset_search(self):
        """
        重置搜索条件
        """
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            self.search_meter_no.delete(0, tk.END)
            self.search_meter_type.set("")
            self.load_meter_list()
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def on_meter_select(self, _event=None):
        """
        列表选择事件处理
        """
        try:
            selected_items = self.meter_tree.selection()
            if not selected_items:
                return
            
            item_id = selected_items[0]
            
            # 从tags中获取水电表ID
            tags = self.meter_tree.item(item_id, "tags")
            if tags:
                meter_id = int(tags[0])
            else:
                # 如果没有tags，尝试从列表中匹配
                values = self.meter_tree.item(item_id, "values")
                if values and len(values) > 1:
                    meter_no, meter_type, _, _, _, _ = values[1:]
                    # 通过多字段匹配查找水电表
                    self.selected_meter = next((m for m in self.meter_list if m.meter_no == meter_no and m.meter_type == meter_type), None)
                    return
                else:
                    messagebox.showerror(self.get_text("error"), f"{self.get_text('failed_to_get_meter_id')}")
                    return
            
            # 查找对应的水电表对象
            self.selected_meter = next((m for m in self.meter_list if m.id == meter_id), None)
            if self.selected_meter:
                self.fill_form()
            else:
                # 如果找不到对应的水电表对象，可能是数据已过期，重新加载列表
                self.show_temporary_message(f"{self.get_text('data_synchronizing')}")
                # 使用after方法异步重新加载，避免阻塞UI
                self.parent.after(500, self.load_meter_list)
        except Exception as e:
            messagebox.showerror("错误", f"选择水电表时发生异常：{str(e)}")
    
    def fill_form(self):
        """
        填充表单
        """
        try:
            if self.selected_meter:
                # 表编号
                self.form_meter_no.delete(0, tk.END)
                self.form_meter_no.insert(0, self.selected_meter.meter_no)
                
                # 表类型
                self.form_meter_type.set(self.selected_meter.meter_type)
                
                # 所属租户
                tenant = Tenant.get_by_id(self.selected_meter.tenant_id)
                if tenant:
                    self.form_tenant.set(tenant.name)
                
                # 安装位置
                self.form_location.delete(0, tk.END)
                self.form_location.insert(0, self.selected_meter.location)
                
                # 初始读数
                self.form_initial_reading.delete(0, tk.END)
                self.form_initial_reading.insert(0, str(self.selected_meter.initial_reading))
                
                # 状态
                self.form_status.set(self.selected_meter.status)
        except Exception as e:
            messagebox.showerror(self.get_text("error"), f"{self.get_text('failed_to_fill_form').format(str(e))}")
    
    def clear_form(self):
        """
        清空表单
        """
        self.selected_meter = None
        self.form_meter_no.delete(0, tk.END)
        self.form_meter_type.set("")
        self.form_tenant.set("")
        self.form_location.delete(0, tk.END)
        self.form_initial_reading.delete(0, tk.END)
        self.form_initial_reading.insert(0, "0")
        self.form_status.set("正常")
    
    def add_meter(self):
        """
        添加水电表
        """
        self.clear_form()
        # 确保焦点在表编号输入框，方便用户直接输入
        self.form_meter_no.focus_set()
    
    def edit_meter(self):
        """
        编辑水电表
        """
        # 检查是否有选中的水电表
        selected_items = self.meter_tree.selection()
        if not selected_items:
            messagebox.showwarning(self.get_text("warning"), f"{self.get_text('please_select_meter_edit')}")
            return
        
        # 获取选中的水电表对象
        item_id = selected_items[0]
        
        # 从tags中获取水电表ID
        tags = self.meter_tree.item(item_id, "tags")
        if tags:
            meter_id = int(tags[0])
        else:
            messagebox.showerror(self.get_text("error"), f"{self.get_text('failed_to_get_meter_id')}")
            return
        
        # 查找对应的水电表对象
        self.selected_meter = next((m for m in self.meter_list if m.id == meter_id), None)
        if not self.selected_meter:
            messagebox.showwarning(self.get_text('warning'), f"{self.get_text('failed_to_find_meter')}")
            return
        
        # 填充表单
        self.fill_form()
        # 确保焦点在表编号输入框，方便用户直接编辑
        self.form_meter_no.focus_set()
    
    def delete_meter(self):
        """
        删除水电表
        """
        # 检查是否有选中的水电表
        selected_items = self.meter_tree.selection()
        if not selected_items:
            messagebox.showwarning(self.get_text('warning'), f"{self.get_text('please_select_meter_delete')}")
            return
        
        # 获取选中的水电表对象
        item_id = selected_items[0]
        values = self.meter_tree.item(item_id, "values")
        meter_no = values[1]  # 序号是values[0]，表编号是values[1]
        
        # 从tags中获取水电表ID
        tags = self.meter_tree.item(item_id, "tags")
        if tags:
            meter_id = int(tags[0])
        else:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('failed_to_get_meter_id')}")
            return
        
        # 查找对应的水电表对象
        meter_to_delete = next((m for m in self.meter_list if m.id == meter_id), None)
        if not meter_to_delete:
            messagebox.showwarning(self.get_text('warning'), f"{self.get_text('failed_to_find_meter')}")
            return
        
        try:
            if messagebox.askyesno(self.get_text('system_confirm_delete'), f"{self.get_text('confirm_delete_meter').format(meter_no)}"):
                # 更新UI状态，显示加载中
                self.parent.config(cursor="wait")
                self.parent.update()
                
                if meter_to_delete.delete():
                    messagebox.showinfo(self.get_text('success'), f"{self.get_text('meter_delete_success').format(meter_no)}")
                    self.load_meter_list()
                    self.clear_form()
                else:
                    messagebox.showerror(self.get_text('error'), f"{self.get_text('meter_delete_fail')}")
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('delete_meter_exception').format(str(e))}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def load_meter_list(self):
        """
        加载水电表列表
        确保页面状态的一致性和正确性
        """
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            # 清空现有列表
            for item in self.meter_tree.get_children():
                self.meter_tree.delete(item)
            
            # 清空选择状态和表单
            self.selected_meter = None
            self.clear_form()
            
            # 获取水电表列表
            self.meter_list = Meter.get_all()
            
            # 创建租户ID到名称的映射
            tenants = Tenant.get_all()
            tenant_map = {t.id: t.name for t in tenants}
            
            # 应用默认排序规则：按所属租户、表编号、表类型升序排列
            # 创建租户ID到名称的映射，用于排序
            tenant_map_for_sort = {t.id: t.name for t in tenants}
            # 排序函数
            def meter_sort_key(meter):
                return (
                    tenant_map_for_sort.get(meter.tenant_id, "未知租户"),  # 首要：所属租户
                    meter.meter_no,  # 次要：表编号
                    meter.meter_type  # 第三：表类型
                )
            # 执行排序
            self.meter_list.sort(key=meter_sort_key)
            
            # 添加到列表控件，添加序号列
            for idx, meter in enumerate(self.meter_list, 1):
                tenant_name = tenant_map.get(meter.tenant_id, "未知租户")
                # 翻译数据值
                translated_meter_type = self.get_text(meter.meter_type)
                translated_status = self.get_text(meter.status)
                self.meter_tree.insert("", tk.END, values=(idx, meter.meter_no, 
                                                        translated_meter_type, 
                                                        tenant_name, 
                                                        meter.location, 
                                                        meter.initial_reading, 
                                                        translated_status), tags=(str(meter.id),))
            
            # 清空选择状态
            self.clear_selection()
            
            # 短暂显示刷新成功提示
            # self.show_temporary_message(f"{self.get_text('list_refreshed')}")
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('load_meter_list_exception').format(str(e))}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def clear_selection(self):
        """
        清空列表选择状态
        """
        self.selected_meter = None
        for item in self.meter_tree.selection():
            self.meter_tree.selection_remove(item)
    
    def show_temporary_message(self, message, duration=1500):
        """
        显示临时消息提示
        :param message: 消息内容
        :param duration: 显示时长（毫秒）
        """
        # 创建临时消息窗口
        temp_window = tk.Toplevel(self.parent)
        temp_window.title(f"{self.get_text('temporary_message')}")
        temp_window.transient(self.parent)
        temp_window.attributes("-topmost", True)
        
        # 计算位置，居中显示
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        message_width = 200
        message_height = 60
        
        x = parent_x + (parent_width - message_width) // 2
        y = parent_y + (parent_height - message_height) // 2
        
        temp_window.geometry(f"{message_width}x{message_height}+{x}+{y}")
        
        # 添加消息标签
        ttk.Label(temp_window, text=message, font=("Arial", 12)).pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        
        # 设置自动关闭
        temp_window.after(duration, temp_window.destroy)
    
    def save_meter(self):
        """
        保存水电表信息
        """
        # 获取表单数据
        meter_no = self.form_meter_no.get().strip()
        meter_type = self.form_meter_type.get().strip()
        tenant_name = self.form_tenant.get().strip()
        location = self.form_location.get().strip()
        initial_reading_str = self.form_initial_reading.get().strip()
        status = self.form_status.get().strip()
        
        # 验证数据
        if not meter_no:
            messagebox.showwarning(self.get_text('warning'), f"{self.get_text('please_enter_meter_no')}")
            self.form_meter_no.focus_set()
            return
        
        if not meter_type:
            messagebox.showwarning(self.get_text('warning'), f"{self.get_text('please_select_meter_type')}")
            self.form_meter_type.focus_set()
            return
        
        if not tenant_name:
            messagebox.showwarning(self.get_text('warning'), f"{self.get_text('please_select_tenant')}")
            self.form_tenant.focus_set()
            return
        
        if not location:
            messagebox.showwarning(self.get_text('warning'), f"{self.get_text('please_enter_location')}")
            self.form_location.focus_set()
            return
        
        try:
            initial_reading = float(initial_reading_str)
        except ValueError:
            messagebox.showwarning(self.get_text('warning'), f"{self.get_text('initial_reading_must_be_number')}")
            self.form_initial_reading.focus_set()
            return
        
        if not status:
            messagebox.showwarning(self.get_text('warning'), f"{self.get_text('please_select_status')}")
            self.form_status.focus_set()
            return
        
        # 获取租户对象
        tenant = self.tenants_dict.get(tenant_name)
        if not tenant:
            # 可能是新添加的租户，重新加载租户列表
            self.load_tenants_to_form()
            tenant = self.tenants_dict.get(tenant_name)
            if not tenant:
                messagebox.showwarning(self.get_text('warning'), f"{self.get_text('please_select_valid_tenant')}")
                self.form_tenant.focus_set()
                return
        
        # 检查是否存在重复的表编号
        if self.selected_meter:
            # 编辑模式，排除当前ID
            if Meter.exists_by_meter_no(meter_no, self.selected_meter.id):
                messagebox.showwarning(self.get_text('warning'), f"{self.get_text('meter_no_exists').format(meter_no)}")
                self.form_meter_no.focus_set()
                return
        else:
            # 添加模式
            if Meter.exists_by_meter_no(meter_no):
                messagebox.showwarning(self.get_text('warning'), f"{self.get_text('meter_no_exists').format(meter_no)}")
                self.form_meter_no.focus_set()
                return
        
        # 保存水电表信息
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            if self.selected_meter:
                # 更新现有水电表
                self.selected_meter.meter_no = meter_no
                self.selected_meter.meter_type = meter_type
                self.selected_meter.tenant_id = tenant.id
                self.selected_meter.location = location
                self.selected_meter.initial_reading = initial_reading
                self.selected_meter.status = status
                
                if self.selected_meter.save():
                    messagebox.showinfo(self.get_text('success'), f"{self.get_text('meter_update_success').format(meter_no)}")
                    self.load_meter_list()
                    self.clear_form()
                else:
                    messagebox.showerror(self.get_text('error'), f"{self.get_text('meter_update_fail')}")
            else:
                # 添加新水电表
                meter = Meter(
                    meter_no=meter_no,
                    meter_type=meter_type,
                    tenant_id=tenant.id,
                    location=location,
                    initial_reading=initial_reading,
                    status=status
                )
                
                if meter.save():
                    messagebox.showinfo(self.get_text('success'), f"{self.get_text('meter_add_success').format(meter_no)}")
                    self.load_meter_list()
                    self.clear_form()
                else:
                    messagebox.showerror(self.get_text('error'), f"{self.get_text('meter_add_fail')}")
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('save_meter_exception').format(str(e))}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def export_meters(self):
        """
        导出水电表信息
        支持导出当前筛选条件下的所有水电表数据到Excel文件
        """
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.get_text('meter_information')}"
            
            # 设置表头
            headers = [f"{self.get_text('meter_id')}", f"{self.get_text('meter_no')}", f"{self.get_text('meter_type')}", f"{self.get_text('tenant_name')}", f"{self.get_text('location')}", f"{self.get_text('initial_reading')}", f"{self.get_text('status')}", f"{self.get_text('create_time')}"]
            ws.append(headers)
            
            # 设置表头样式
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # 获取列字母
                # 设置表头字体和对齐方式
                ws[column + "1"].font = Font(bold=True, size=12)
                ws[column + "1"].alignment = Alignment(horizontal="center")
                # 计算列宽
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
                ws.column_dimensions[column].width = adjusted_width
            
            # 获取当前筛选条件下的水电表列表
            meter_no = self.search_meter_no.get()
            meter_type = self.search_meter_type.get()
            filters = {}
            if meter_type:
                filters['meter_type'] = meter_type
            meter_list = Meter.get_all(filters)
            
            # 创建租户ID到名称的映射
            tenants = Tenant.get_all()
            tenant_map = {t.id: t.name for t in tenants}
            
            # 写入数据
            for meter in meter_list:
                # 如果指定了表编号，进行过滤
                if meter_no and meter_no not in meter.meter_no:
                    continue
                    
                ws.append([
                    meter.id,
                    meter.meter_no,
                    meter.meter_type,
                    tenant_map.get(meter.tenant_id, "未知租户"),
                    meter.location,
                    meter.initial_reading,
                    meter.status,
                    meter.create_time if meter.create_time else ""
                ])
            
            # 保存文件
            default_filename = f"水电表信息_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                initialfile=default_filename
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo(self.get_text('success'), f"{self.get_text('export_meter_success').format(file_path)}")
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('export_meter_fail').format(str(e))}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def import_meters(self):
        """
        导入水电表信息
        支持导入Excel文件中的水电表数据，包含数据验证和进度显示
        """
        try:
            # 选择文件
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                title=f"{self.get_text('select_meter_excel_file')}"
            )
            
            if not file_path:
                return
            
            # 加载文件
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 验证表头
            expected_headers = [f"{self.get_text('meter_no')}", f"{self.get_text('meter_type')}", f"{self.get_text('tenant_name')}", f"{self.get_text('location')}", f"{self.get_text('initial_reading')}", f"{self.get_text('status')}"]
            actual_headers = [cell.value for cell in ws[1]]
            
            # 检查是否包含所有必填表头
            missing_headers = [header for header in expected_headers if header not in actual_headers]
            if missing_headers:
                messagebox.showerror(self.get_text('error'), f"{self.get_text('file_format_error').format(', '.join(missing_headers))}")
                return
            
            # 创建进度窗口
            progress_window = tk.Toplevel(self.parent)
            progress_window.title(f"{self.get_text('import_progress')}")
            progress_window.geometry("400x150")
            progress_window.transient(self.parent)
            progress_window.grab_set()
            
            # 进度标签
            progress_label = ttk.Label(progress_window, text=f"{self.get_text('preparing_import')}")
            progress_label.pack(pady=20)
            
            # 进度条
            progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate')
            progress_bar.pack(pady=10)
            
            # 结果标签
            result_label = ttk.Label(progress_window, text="")
            result_label.pack(pady=10)
            
            progress_window.update()
            
            # 初始化统计信息
            total_rows = ws.max_row - 1  # 减去表头
            success_count = 0
            failed_count = 0
            failed_reasons = []
            
            # 创建租户名称到ID的映射
            tenants = Tenant.get_all()
            tenant_map = {t.name: t.id for t in tenants}
            
            # 解析数据
            for row_idx in range(2, ws.max_row + 1):
                row_data = ws[row_idx]
                try:
                    # 获取数据
                    meter_no = row_data[actual_headers.index("表编号")].value
                    meter_type = row_data[actual_headers.index("表类型")].value
                    tenant_name = row_data[actual_headers.index("所属租户")].value
                    location = row_data[actual_headers.index("安装位置")].value if "安装位置" in actual_headers else ""
                    initial_reading = row_data[actual_headers.index("初始读数")].value if "初始读数" in actual_headers else 0
                    status = row_data[actual_headers.index("状态")].value if "状态" in actual_headers else "正常"
                    
                    # 数据验证
                    if not meter_no:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：表编号不能为空")
                        continue
                    
                    if not meter_type or meter_type not in ["水", "电"]:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：表类型必须是'水'或'电'")
                        continue
                    
                    if not tenant_name:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：所属租户不能为空")
                        continue
                    
                    if tenant_name not in tenant_map:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：租户 '{tenant_name}' 不存在，请先添加该租户")
                        continue
                    
                    if not location:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：安装位置不能为空")
                        continue
                    
                    # 验证初始读数
                    try:
                        initial_reading = float(initial_reading) if initial_reading else 0
                    except ValueError:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：初始读数必须是数字")
                        continue
                    
                    if not status or status not in ["正常", "损坏", "更换"]:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：状态必须是'正常'、'损坏'或'更换'")
                        continue
                    
                    # 检查是否已存在相同表编号的水电表
                    if Meter.exists_by_meter_no(meter_no):
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：水电表 '{meter_no}' 已存在")
                        continue
                    
                    # 创建并保存水电表
                    meter = Meter(
                        meter_no=meter_no,
                        meter_type=meter_type,
                        tenant_id=tenant_map[tenant_name],
                        location=location,
                        initial_reading=initial_reading,
                        status=status
                    )
                    
                    if meter.save():
                        success_count += 1
                    else:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx}行：保存失败")
                        continue
                    
                except Exception as e:
                    failed_count += 1
                    failed_reasons.append(f"第{row_idx}行：{str(e)}")
                
                # 更新进度
                progress = int((row_idx - 1) / total_rows * 100)
                progress_bar['value'] = progress
                progress_label.config(text=f"{self.get_text('importing_row').format(row_idx - 1, total_rows)}")
                result_label.config(text=f"{self.get_text('success')}: {success_count}, {self.get_text('fail')}: {failed_count}")
                progress_window.update()
            
            # 关闭进度窗口
            progress_window.destroy()
            
            # 显示导入结果
            result_message = f"{self.get_text('import_completed').format(total_rows, success_count, failed_count)}"
            
            if failed_reasons:
                result_message += f"\n{self.get_text('failed_reasons')}:\n"
                result_message += "\n".join(failed_reasons[:10])  # 只显示前10条失败原因
                if len(failed_reasons) > 10:
                    result_message += f"\n{self.get_text('more_failed_reasons').format(len(failed_reasons) - 10)}"
            
            messagebox.showinfo(f"{self.get_text('import_result')}", result_message)
            
            # 刷新水电表列表
            self.load_meter_list()
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('import_meter_fail').format(str(e))}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
