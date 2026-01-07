#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
抄表管理视图
负责处理抄表数据的录入、编辑、删除和查询
"""

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import xlrd
import os
from models.tenant import Tenant
from models.meter import Meter
from models.reading import MeterReading
from models.charge import Charge
from utils.language_utils import LanguageUtils

class ReadingView:
    """抄表管理视图类"""
    
    def __init__(self, parent, language_utils=None):
        """
        初始化抄表管理视图
        :param parent: 父窗口组件
        :param language_utils: 语言工具类实例
        """
        self.parent = parent
        self.reading_list = []
        self.selected_reading = None
        # 使用传入的语言工具或创建新实例
        self.language_utils = language_utils if language_utils else LanguageUtils()
        self.create_widgets()
        self.load_reading_list()
    
    def get_text(self, key):
        """
        获取当前语言的文本
        :param key: 文本键名
        :return: 对应语言的文本
        """
        return self.language_utils.get_text(key)
    
    def update_language(self):
        """
        更新语言
        当系统语言切换时调用，更新界面上所有文本
        """
        # 更新筛选条件标签和按钮
        filter_frame = self.parent.winfo_children()[0].winfo_children()[0]  # 获取筛选框架
        filter_children = filter_frame.winfo_children()
        
        # 更新月份标签
        if len(filter_children) > 0 and isinstance(filter_children[0], ttk.Label):
            filter_children[0].config(text=self.get_text("month")+":")
        
        # 更新租户标签
        if len(filter_children) > 3 and isinstance(filter_children[3], ttk.Label):
            filter_children[3].config(text=self.get_text("tenant")+":")
        
        # 更新按钮文本
        button_texts = ["button_search", "button_reset", "make_hand_copy", "export_reading_form", "import_reading_form", "delete_reading", "select_all"]
        for i, btn_idx in enumerate([4, 5, 6, 7, 8, 9]):
            if i < len(button_texts) and btn_idx < len(filter_children) and isinstance(filter_children[btn_idx], ttk.Button):
                filter_children[btn_idx].config(text=self.get_text(button_texts[i]))
        
        # 更新全选按钮文本
        if self.select_all_var.get():
            self.select_all_btn.config(text=self.get_text("deselect_all"))
        else:
            self.select_all_btn.config(text=self.get_text("select_all"))
        
        # 更新记录总数标签
        self.total_records_label.config(text=self.get_text("total_records").format(len(self.reading_list)))
        
        # 更新列标题
        column_texts = {
            "month": "month",
            "tenant_name": "tenant_name",
            "meter_no": "meter_no",
            "meter_type": "meter_type",
            "previous_reading": "previous_reading",
            "current_reading": "current_reading",
            "adjustment": "adjustment",
            "usage": "usage",
            "reading_date": "reading_date",
            "charging_time": "charging_time",
            "reader": "reader"
        }
        
        for col, key in column_texts.items():
            # 保留排序符号
            current_text = self.reading_tree.heading(col, "text")
            sort_symbol = "" if len(current_text.split()) <= 1 else current_text.split()[1]
            self.reading_tree.heading(col, text=f"{self.get_text(key)} {sort_symbol}")
        
        # 更新表单标题
        main_frame = self.parent.winfo_children()[0]
        middle_frame = main_frame.winfo_children()[1] if len(main_frame.winfo_children()) > 1 else None
        if middle_frame:
            right_frame = middle_frame.winfo_children()[1] if len(middle_frame.winfo_children()) > 1 else None
            if right_frame:
                form_title = right_frame.winfo_children()[0] if isinstance(right_frame.winfo_children()[0], ttk.Label) else None
                if form_title:
                    form_title.config(text=self.get_text("reading_details"))
        
        # 更新表单控件标签
        if right_frame:
            form_frame = right_frame.winfo_children()[1] if len(right_frame.winfo_children()) > 1 else None
            if form_frame:
                form_children = form_frame.winfo_children()
                # 更新表单标签
                label_keys = ["month", "tenant", "meter", "previous_reading", "current_reading", "adjustment", "usage", "reading_date", "reader"]
                for i, label_idx in enumerate([0, 2, 4, 6, 8, 10, 12, 14, 16]):
                    if i < len(label_keys) and label_idx < len(form_children) and isinstance(form_children[label_idx], ttk.Label):
                        form_children[label_idx].config(text=self.get_text(label_keys[i])+":")
                
                # 更新按钮
                if len(form_children) > 18 and isinstance(form_children[18], ttk.Frame):
                    button_container = form_children[18]
                    button_children = button_container.winfo_children()
                    if len(button_children) > 0 and isinstance(button_children[0], ttk.Button):
                        button_children[0].config(text=self.get_text("button_save"))
                    if len(button_children) > 1 and isinstance(button_children[1], ttk.Button):
                        button_children[1].config(text=self.get_text("button_cancel"))
        
        # 重新加载数据，确保显示当前语言
        self.load_reading_list()
        # 更新统计信息，确保使用最新语言
        self.update_stats()
    
    def create_widgets(self):
        """
        创建抄表管理界面组件
        """
        # 创建主框架
        main_frame = ttk.Frame(self.parent)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 顶部：筛选条件
        filter_frame = ttk.Frame(main_frame)
        filter_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        
        # 月份选择 - 重构为具备自动完成功能的下拉列表
        ttk.Label(filter_frame, text=self.get_text("month")+"：").pack(side=tk.LEFT, padx=5)
        self.month_var = tk.StringVar()
        self.month_var.set(datetime.now().strftime("%Y-%m"))
        
        # 创建月份下拉列表组件
        self.month_combobox = ttk.Combobox(filter_frame, textvariable=self.month_var, width=10, state="normal", values=[])
        self.month_combobox.pack(side=tk.LEFT, padx=5)
        
        # 加载所有已存在的月份数据
        self.load_existing_months()
        
        # 绑定月份选择事件
        self.month_combobox.bind("<<ComboboxSelected>>", self.on_month_selected)
        # 绑定输入事件，实现自动匹配
        self.month_combobox.bind("<KeyRelease>", self.on_month_input)
        # 绑定回车键确认选择
        self.month_combobox.bind("<Return>", self.on_month_confirm)
        # 绑定ESC键关闭下拉面板
        self.month_combobox.bind("<Escape>", self.on_month_escape)
        
        # 租户筛选
        ttk.Label(filter_frame, text=self.get_text("tenant")+"：").pack(side=tk.LEFT, padx=5)
        self.tenant_var = tk.StringVar()
        self.tenant_combobox = ttk.Combobox(filter_frame, textvariable=self.tenant_var)
        self.load_tenants_to_combobox()
        self.tenant_combobox.pack(side=tk.LEFT, padx=5)
        
        # 操作按钮
        ttk.Button(filter_frame, text=self.get_text("button_search"), command=self.search_readings).pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_frame, text=self.get_text("button_reset"), command=self.reset_filter).pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_frame, text=self.get_text("make_hand_copy"), command=self.make_hand_copy).pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_frame, text=self.get_text("export_reading_form"), command=self.export_reading_form).pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_frame, text=self.get_text("import_reading_form"), command=self.import_reading_form).pack(side=tk.LEFT, padx=5)

        # 添加删除记录和全选按钮
        ttk.Button(filter_frame, text=self.get_text("delete_reading"), command=self.delete_reading).pack(side=tk.LEFT, padx=5)
        # 全选按钮
        self.select_all_var = tk.BooleanVar()
        self.select_all_btn = ttk.Button(filter_frame, text=self.get_text("select_all"), command=self.toggle_select_all)
        self.select_all_btn.pack(side=tk.LEFT, padx=5)
        
        # 添加记录总数标签
        self.total_records_label = ttk.Label(filter_frame, text=self.get_text("total_records").format(0), foreground="blue")
        self.total_records_label.pack(side=tk.LEFT, padx=10)
        
        # 中部：中间框架，用于水平排列表格和表单
        middle_frame = ttk.Frame(main_frame)
        middle_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 配置grid布局，使两列能够自适应宽度
        middle_frame.grid_columnconfigure(0, weight=1, minsize=500)
        middle_frame.grid_columnconfigure(1, weight=1, minsize=300)
        middle_frame.grid_rowconfigure(0, weight=1)
        
        # 左侧：抄表数据表格
        table_frame = ttk.Frame(middle_frame)
        table_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        
        # 表格控件 - 移除id列，添加序号列，添加计费时间列，添加调整列，添加备注列
        columns = ("serial", "month", "tenant_name", "meter_no", "meter_type", "previous_reading", 
                  "current_reading", "adjustment", "usage", "remark", "reading_date", "charging_time", "reader")
        self.reading_tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        
        # 保存排序状态
        self.sort_column = None
        self.sort_order = "asc"
        
        # 设置列标题和宽度
        # 添加序号列，不添加点击事件（不需要排序）
        self.reading_tree.heading("serial", text=self.get_text("serial"))
        self.reading_tree.column("serial", width=80, anchor="center")
        
        # 所属月份列
        self.reading_tree.heading("month", text=self.get_text("month"), command=lambda c="month": self.sort_column_clicked(c))
        self.reading_tree.column("month", width=90, anchor="center")
        
        self.reading_tree.heading("tenant_name", text=self.get_text("tenant_name"), command=lambda c="tenant_name": self.sort_column_clicked(c))
        self.reading_tree.column("tenant_name", width=150, anchor="w")
        
        self.reading_tree.heading("meter_no", text=self.get_text("meter_no"), command=lambda c="meter_no": self.sort_column_clicked(c))
        self.reading_tree.column("meter_no", width=100, anchor="w")
        
        self.reading_tree.heading("meter_type", text=self.get_text("meter_type"), command=lambda c="meter_type": self.sort_column_clicked(c))
        self.reading_tree.column("meter_type", width=80, anchor="center")
        
        self.reading_tree.heading("previous_reading", text=self.get_text("previous_reading"), command=lambda c="previous_reading": self.sort_column_clicked(c))
        self.reading_tree.column("previous_reading", width=100, anchor="e")
        
        self.reading_tree.heading("current_reading", text=self.get_text("current_reading"), command=lambda c="current_reading": self.sort_column_clicked(c))
        self.reading_tree.column("current_reading", width=100, anchor="e")
        
        self.reading_tree.heading("adjustment", text=self.get_text("adjustment"), command=lambda c="adjustment": self.sort_column_clicked(c))
        self.reading_tree.column("adjustment", width=80, anchor="e")
        
        self.reading_tree.heading("usage", text=self.get_text("usage"), command=lambda c="usage": self.sort_column_clicked(c))
        self.reading_tree.column("usage", width=80, anchor="e")
        
        # 备注列
        self.reading_tree.heading("remark", text=self.get_text("remark"), command=lambda c="remark": self.sort_column_clicked(c))
        self.reading_tree.column("remark", width=80, anchor="w")
        
        self.reading_tree.heading("reading_date", text=self.get_text("reading_date"), command=lambda c="reading_date": self.sort_column_clicked(c))
        self.reading_tree.column("reading_date", width=120, anchor="center")
        
        self.reading_tree.heading("charging_time", text=self.get_text("charging_time"), command=lambda c="charging_time": self.sort_column_clicked(c))
        self.reading_tree.column("charging_time", width=120, anchor="center")
        
        self.reading_tree.heading("reader", text=self.get_text("reader"), command=lambda c="reader": self.sort_column_clicked(c))
        self.reading_tree.column("reader", width=100, anchor="center")
        
        # 配置Treeview的tag，用于实现隔行背景色
        # 使用深灰色和白色的对比，奇数行使用深灰色，偶数行使用白色
        self.reading_tree.tag_configure('odd', background='#c0c0c0')
        self.reading_tree.tag_configure('even', background='#ffffff')
        
        # 添加滚动条
        v_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.reading_tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.reading_tree.xview)
        self.reading_tree.configure(yscroll=v_scrollbar.set, xscroll=h_scrollbar.set)
        
        # 配置表格框架的grid布局
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        self.reading_tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # 统计信息区域
        stats_frame = ttk.Frame(table_frame)
        stats_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=5, padx=10)
        
        # 水类型统计标签
        self.water_stats_label = ttk.Label(stats_frame, text="水表：用量总数: 0, 调整总数: 0", font=('Arial', 10))
        self.water_stats_label.pack(side=tk.LEFT, padx=20)
        
        # 电类型统计标签
        self.electricity_stats_label = ttk.Label(stats_frame, text="电表：用量总数: 0, 调整总数: 0", font=('Arial', 10))
        self.electricity_stats_label.pack(side=tk.LEFT, padx=20)
        
        # 绑定窗口大小变化事件，确保统计信息区域正确显示
        table_frame.bind("<Configure>", lambda e: stats_frame.update_idletasks())
        
        # 绑定列表选择事件
        self.reading_tree.bind("<<TreeviewSelect>>", self.on_reading_select)
        
        # 右侧：抄表明细表单
        right_frame = ttk.Frame(middle_frame)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        
        # 配置右侧框架的grid布局
        right_frame.grid_rowconfigure(1, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)
        
        # 表单标题
        ttk.Label(right_frame, text=self.get_text("reading_details"), font=('Arial', 14)).grid(row=0, column=0, pady=10, sticky="n")
        
        # 表单框架
        form_frame = ttk.Frame(right_frame)
        form_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        
        # 配置表单框架的grid布局
        form_frame.grid_columnconfigure(1, weight=1)
        
        # 表单字段 - 改为垂直排列，每行两列（标签+输入）
        
        # 所属月份 - 重构为具备自动完成功能的下拉列表
        ttk.Label(form_frame, text=self.get_text("month")+":").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.form_month = ttk.Combobox(form_frame, width=20, state="normal", values=[])
        self.form_month.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        # 设置默认值为当前月份，格式为YYYY-MM
        self.form_month.set(datetime.now().strftime("%Y-%m"))
        
        # 加载所有已存在的月份数据到表单中的月份选择下拉列表
        self.load_existing_months_to_form()
        
        # 租户
        ttk.Label(form_frame, text=self.get_text("tenant")+":").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.form_tenant = ttk.Combobox(form_frame, values=[], width=20)
        self.form_tenant.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        
        # 水电表
        ttk.Label(form_frame, text=self.get_text("meter")+":").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.form_meter = ttk.Combobox(form_frame, values=[], width=20)
        self.form_meter.grid(row=2, column=1, sticky="ew", padx=5, pady=5)
        
        # 上次读数
        ttk.Label(form_frame, text=self.get_text("previous_reading")+":").grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.form_previous_reading = ttk.Entry(form_frame, width=20)
        self.form_previous_reading.grid(row=3, column=1, sticky="ew", padx=5, pady=5)
        self.form_previous_reading.config(state="readonly")
        
        # 当前读数
        ttk.Label(form_frame, text=self.get_text("current_reading")+":").grid(row=4, column=0, sticky="e", padx=5, pady=5)
        self.form_current_reading = ttk.Entry(form_frame, width=20)
        self.form_current_reading.grid(row=4, column=1, sticky="ew", padx=5, pady=5)
        
        # 调整
        ttk.Label(form_frame, text=self.get_text("adjustment")+":").grid(row=5, column=0, sticky="e", padx=5, pady=5)
        self.form_adjustment = ttk.Entry(form_frame, width=20)
        self.form_adjustment.grid(row=5, column=1, sticky="ew", padx=5, pady=5)
        self.form_adjustment.insert(0, "0")
        # 用量
        ttk.Label(form_frame, text=self.get_text("usage")+":").grid(row=6, column=0, sticky="e", padx=5, pady=5)
        self.form_usage = ttk.Entry(form_frame, width=20)
        self.form_usage.grid(row=6, column=1, sticky="ew", padx=5, pady=5)
        self.form_usage.config(state="readonly")
        
        # 备注
        ttk.Label(form_frame, text=self.get_text("remark")+":").grid(row=7, column=0, sticky="ne", padx=5, pady=5)
        # 创建备注文本框，高度固定为60px，支持多行输入
        self.form_remark = tk.Text(form_frame, width=20, height=3, wrap=tk.WORD)
        self.form_remark.grid(row=7, column=1, sticky="ew", padx=5, pady=5)
        # 设置高度为60px
        self.form_remark.configure(height=3)
        # 添加字符数限制校验
        self.form_remark.bind("<KeyRelease>", self.on_remark_change)
        # 添加字符数显示标签
        self.remark_length_label = ttk.Label(form_frame, text="0/200", font=('Arial', 8), foreground="gray")
        self.remark_length_label.grid(row=7, column=2, sticky="n", padx=5, pady=5)
        
        # 抄表日期
        ttk.Label(form_frame, text=self.get_text("reading_date")+":").grid(row=8, column=0, sticky="e", padx=5, pady=5)
        self.form_reading_date = ttk.Entry(form_frame, width=20)
        self.form_reading_date.grid(row=8, column=1, sticky="ew", padx=5, pady=5)
        self.form_reading_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        # 抄表人
        ttk.Label(form_frame, text=self.get_text("reader")+":").grid(row=9, column=0, sticky="e", padx=5, pady=5)
        self.form_reader = ttk.Entry(form_frame, width=20)
        self.form_reader.grid(row=9, column=1, sticky="ew", padx=5, pady=5)
        self.form_reader.insert(0, "admin")  # 默认当前用户
        
        # 表单操作按钮容器
        buttons_container = ttk.Frame(form_frame)
        buttons_container.grid(row=10, column=0, columnspan=3, pady=10, sticky="ew")
        buttons_container.grid_columnconfigure(0, weight=1)
        buttons_container.grid_columnconfigure(1, weight=1)
        
        # 表单操作按钮 - 水平排列
        self.save_button = ttk.Button(buttons_container, text=self.get_text("button_save"), command=self.save_reading)
        self.save_button.pack(side=tk.LEFT, padx=5, pady=5)
        
        self.cancel_button = ttk.Button(buttons_container, text=self.get_text("button_cancel"), command=self.clear_form)
        self.cancel_button.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 绑定租户选择事件，动态加载水电表
        self.form_tenant.bind("<<ComboboxSelected>>", self.on_tenant_selected)
        
        # 绑定当前读数输入事件，自动计算用量
        self.form_current_reading.bind("<KeyRelease>", self.on_current_reading_change)
        
        # 绑定调整值变化事件，自动重新计算用量
        self.form_adjustment.bind("<KeyRelease>", self.on_adjustment_change)
        
        # 绑定自动填充上次读数的事件
        # 绑定所属月份输入事件
        self.form_month.bind("<KeyRelease>", self.auto_fill_previous_reading)
        # 绑定水电表选择事件
        self.form_meter.bind("<<ComboboxSelected>>", self.auto_fill_previous_reading)
        
        # 加载租户数据到表单下拉框
        self.load_tenants_to_form()
        
        # 初始加载数据并更新统计信息
        self.load_reading_list()
    
    def load_tenants_to_combobox(self):
        """
        加载租户数据到筛选下拉框
        """
        tenants = Tenant.get_all()
        tenant_names = [""] + [t.name for t in tenants]
        self.tenant_combobox['values'] = tenant_names
    
    def load_existing_months(self):
        """
        加载所有已存在的月份数据到月份选择下拉列表
        """
        try:
            # 获取所有已存在的月份数据
            from models.reading import MeterReading
            existing_months = MeterReading.get_all_months()
            
            # 添加当前月份（如果不存在）
            current_month = datetime.now().strftime("%Y-%m")
            if current_month not in existing_months:
                existing_months.insert(0, current_month)
            
            # 设置下拉列表的值
            self.month_combobox['values'] = existing_months
        except Exception as e:
            # 处理空状态
            print(f"加载月份数据失败: {str(e)}")
            self.month_combobox['values'] = [datetime.now().strftime("%Y-%m")]
    
    def load_existing_months_to_form(self):
        """
        加载所有已存在的月份数据到表单中的月份选择下拉列表
        """
        try:
            # 获取所有已存在的月份数据
            from models.reading import MeterReading
            existing_months = MeterReading.get_all_months()
            
            # 添加当前月份（如果不存在）
            current_month = datetime.now().strftime("%Y-%m")
            if current_month not in existing_months:
                existing_months.insert(0, current_month)
            
            # 设置下拉列表的值
            self.form_month['values'] = existing_months
            
            # 绑定输入事件，实现自动匹配
            self.form_month.bind("<KeyRelease>", self.on_month_input)
            # 绑定回车键确认选择
            self.form_month.bind("<Return>", self.on_month_confirm)
            # 绑定ESC键关闭下拉面板
            self.form_month.bind("<Escape>", self.on_month_escape)
        except Exception as e:
            # 处理空状态
            print(f"加载表单月份数据失败: {str(e)}")
            self.form_month['values'] = [datetime.now().strftime("%Y-%m")]
    
    def on_month_selected(self, _event=None):
        """
        月份选择事件处理
        """
        # 触发数据加载或筛选操作
        self.load_reading_list()
    
    def on_month_input(self, _event=None):
        """
        月份输入事件处理，实现自动匹配功能
        """
        # 获取当前输入值
        current_input = self.month_combobox.get().strip()
        
        if not current_input:
            # 输入为空，显示所有月份
            self.load_existing_months()
            return
        
        # 获取所有已存在的月份数据
        from models.reading import MeterReading
        existing_months = MeterReading.get_all_months()
        
        # 添加当前月份（如果不存在）
        current_month = datetime.now().strftime("%Y-%m")
        if current_month not in existing_months:
            existing_months.append(current_month)
        
        # 筛选包含输入内容的月份
        filtered_months = [month for month in existing_months if current_input in month]
        
        # 设置下拉列表的值
        self.month_combobox['values'] = filtered_months
        
        # 如果有匹配项，自动展开下拉列表
        if filtered_months:
            self.month_combobox.event_generate("<<ComboboxSelected>>")
    
    def on_month_confirm(self, _event=None):
        """
        回车键确认选择事件处理
        """
        # 验证输入的月份格式
        month_input = self.month_combobox.get().strip()
        try:
            datetime.strptime(month_input, "%Y-%m")
            # 格式正确，触发数据加载
            self.load_reading_list()
        except ValueError:
            # 格式错误，显示错误提示
            messagebox.showwarning(self.get_text("warning"), self.get_text("date_format_error"))
            # 重置为当前月份
            self.month_var.set(datetime.now().strftime("%Y-%m"))
            self.load_existing_months()
    
    def on_month_escape(self, _event=None):
        """
        ESC键关闭下拉面板事件处理
        """
        # 重置为当前月份
        self.month_var.set(datetime.now().strftime("%Y-%m"))
        self.load_existing_months()
        # 关闭下拉面板
        self.month_combobox.event_generate("<<ComboboxSelected>>")
    
    def load_tenants_to_form(self):
        """
        加载租户数据到表单下拉框
        """
        tenants = Tenant.get_all()
        self.tenants_dict = {t.name: t for t in tenants}
        self.form_tenant['values'] = [t.name for t in tenants]
    
    def on_tenant_selected(self, _event=None):
        """
        租户选择事件处理，动态加载水电表
        """
        tenant_name = self.form_tenant.get()
        if tenant_name in self.tenants_dict:
            tenant = self.tenants_dict[tenant_name]
            meters = Meter.get_by_tenant(tenant.id)
            self.meters_dict = {f"{m.meter_no} ({m.meter_type})": m for m in meters}
            self.form_meter['values'] = [f"{m.meter_no} ({m.meter_type})" for m in meters]
            # 清空水电表选择
            self.form_meter.set("")
            self.form_previous_reading.config(state="normal")
            self.form_previous_reading.delete(0, tk.END)
            self.form_previous_reading.config(state="readonly")
            self.form_current_reading.delete(0, tk.END)
            self.form_usage.config(state="normal")
            self.form_usage.delete(0, tk.END)
            self.form_usage.config(state="readonly")
        # 触发自动填充上次读数
        self.auto_fill_previous_reading()
    
    def calculate_previous_month(self, current_month):
        """
        计算上一月份
        处理跨年度的月份转换，如2025-01的上一月份为2024-12
        :param current_month: 当前月份，格式为YYYY-MM
        :return: 上一月份，格式为YYYY-MM
        """
        try:
            year, month = map(int, current_month.split('-'))
            if month == 1:
                # 跨年度，月份变为12，年份减1
                previous_month = f"{year-1}-12"
            else:
                # 正常情况，月份减1
                previous_month = f"{year}-{month-1:02d}"
            return previous_month
        except Exception as e:
            # 处理无效月份格式
            print(f"月份计算错误: {str(e)}")
            return None
    
    def auto_fill_previous_reading(self, _event=None):
        """
        自动填充上次读数
        当所属月份、租户和水电表都有效时，自动填充上一月份的历史读数
        """
        # 显示加载状态
        self.form_previous_reading.config(state="normal")
        self.form_previous_reading.delete(0, tk.END)
        self.form_previous_reading.insert(0, "加载中...")
        self.form_previous_reading.config(state="readonly")
        self.parent.update_idletasks()  # 更新UI，显示加载状态
        
        # 获取表单值
        month = self.form_month.get().strip()
        tenant_name = self.form_tenant.get().strip()
        meter_display = self.form_meter.get().strip()
        
        # 检查是否所有条件都满足
        if not month or not tenant_name or not meter_display:
            # 条件不满足，显示0
            self.form_previous_reading.config(state="normal")
            self.form_previous_reading.delete(0, tk.END)
            self.form_previous_reading.insert(0, "0")
            self.form_previous_reading.config(state="readonly")
            self.calculate_usage_with_adjustment()
            return
        
        # 检查租户是否有效
        if tenant_name not in self.tenants_dict:
            # 租户无效，显示0
            self.form_previous_reading.config(state="normal")
            self.form_previous_reading.delete(0, tk.END)
            self.form_previous_reading.insert(0, "0")
            self.form_previous_reading.config(state="readonly")
            self.calculate_usage_with_adjustment()
            return
        
        # 检查水电表是否有效
        if meter_display not in self.meters_dict:
            # 水电表无效，显示0
            self.form_previous_reading.config(state="normal")
            self.form_previous_reading.delete(0, tk.END)
            self.form_previous_reading.insert(0, "0")
            self.form_previous_reading.config(state="readonly")
            self.calculate_usage_with_adjustment()
            return
        
        # 计算上一月份
        previous_month = self.calculate_previous_month(month)
        if not previous_month:
            # 月份计算错误，显示0
            self.form_previous_reading.config(state="normal")
            self.form_previous_reading.delete(0, tk.END)
            self.form_previous_reading.insert(0, "0")
            self.form_previous_reading.config(state="readonly")
            self.calculate_usage_with_adjustment()
            return
        
        # 所有条件都满足，获取上一月份的历史读数
        meter = self.meters_dict[meter_display]
        tenant = self.tenants_dict[tenant_name]
        try:
            # 设计高效的数据查询：直接根据月份、租户ID和水表ID查询上一月份的抄表记录
            from models.reading import MeterReading
            
            # 查询上一月份的抄表记录
            # 1. 获取所有抄表记录
            all_readings = MeterReading.get_by_month(previous_month)
            
            # 2. 筛选匹配的记录
            matching_readings = [r for r in all_readings if r.meter_id == meter.id]
            
            if matching_readings:
                # 有匹配记录，使用最新的一条记录的当前读数作为上次读数
                # 按抄表日期倒序排序，取最新的记录
                latest_reading = max(matching_readings, key=lambda x: x.reading_date)
                last_reading = latest_reading.current_reading
            else:
                # 无匹配历史数据，显示0
                last_reading = 0
            
            # 将上次读数填充到输入框中
            self.form_previous_reading.config(state="normal")
            self.form_previous_reading.delete(0, tk.END)
            self.form_previous_reading.insert(0, str(last_reading))
            self.form_previous_reading.config(state="readonly")
            
            # 自动计算用量
            self.calculate_usage_with_adjustment()
        except Exception as e:
            # 发生错误，显示0
            print(f"自动填充上次读数失败: {str(e)}")
            self.form_previous_reading.config(state="normal")
            self.form_previous_reading.delete(0, tk.END)
            self.form_previous_reading.insert(0, "0")
            self.form_previous_reading.config(state="readonly")
            self.calculate_usage_with_adjustment()
    
    def calculate_usage_with_adjustment(self):
        """
        计算用量，包含调整值
        计算公式：用量 = 当前读数 - 上次读数 + 调整值
        """
        try:
            previous = float(self.form_previous_reading.get())
            current = float(self.form_current_reading.get())
            adjustment = float(self.form_adjustment.get())
            usage = current - previous + adjustment
            self.form_usage.config(state="normal")
            self.form_usage.delete(0, tk.END)
            self.form_usage.insert(0, f"{usage:.2f}")
            self.form_usage.config(state="readonly")
        except ValueError:
            # 输入不是数字，清空用量
            self.form_usage.config(state="normal")
            self.form_usage.delete(0, tk.END)
            self.form_usage.config(state="readonly")
    
    def on_current_reading_change(self, _event=None):
        """
        当前读数输入变化事件，自动计算用量（包含调整值）
        """
        self.calculate_usage_with_adjustment()
    
    def on_adjustment_change(self, _event=None):
        """
        调整值输入变化事件，自动重新计算用量
        """
        self.calculate_usage_with_adjustment()
    
    def on_remark_change(self, _event=None):
        """
        备注输入变化事件，处理字符数限制
        """
        # 获取当前文本，去除末尾的换行符但保留中间的空格
        current_text = self.form_remark.get("1.0", tk.END)[:-1]  # 去掉Tkinter Text组件自动添加的换行符
        # 检查字符数
        if len(current_text) > 200:
            # 超过限制，截断文本
            self.form_remark.delete("1.0", tk.END)
            self.form_remark.insert("1.0", current_text[:200])
            # 给予用户明确提示
            self.remark_length_label.configure(text="200/200", foreground="red")
            messagebox.showwarning(self.get_text("warning"), self.get_text("remark_length_limit"))
        else:
            # 更新字符数显示
            self.remark_length_label.configure(text=f"{len(current_text)}/200", foreground="gray" if len(current_text) < 200 else "red")
    
    def load_reading_list(self):
        """
        加载抄表记录列表
        """
        # 清空现有列表
        for item in self.reading_tree.get_children():
            self.reading_tree.delete(item)
        
        try:
            # 获取当前月份
            month = self.month_var.get()
            
            # 获取抄表记录
            self.reading_list = MeterReading.get_by_month(month)
            
            # 创建租户ID到名称的映射
            tenants = Tenant.get_all()
            tenant_map = {t.id: t.name for t in tenants}
            
            # 添加到列表控件，添加序号列
            for idx, reading in enumerate(self.reading_list, 1):
                if not reading.meter:
                    reading.load_meter_info()
                
                tenant_name = tenant_map.get(reading.meter.tenant_id, "未知租户")
                # 使用统一的日期格式化函数
                display_month, display_date = self.format_date(reading.reading_date)
                
                # 获取计费时间：根据租户ID和月份查找费用记录，使用费用记录的创建时间作为计费时间
                # 每次加载数据时重新查询，确保获取最新数据，避免缓存导致数据不一致
                charging_time = ""
                charges = Charge.get_by_tenant(reading.meter.tenant_id)
                for charge in charges:
                    if charge.month == display_month:
                        if charge.create_time:
                            charging_time = charge.create_time.split(' ')[0] if ' ' in charge.create_time else charge.create_time
                        break
                
                # 翻译表类型
                translated_meter_type = self.get_text(reading.meter.meter_type)
                # 处理备注字段，确保None值显示为空字符串
                remark_display = reading.remark if reading.remark is not None else ""
                # 根据索引应用奇偶行标签
                tag = 'odd' if idx % 2 != 0 else 'even'
                self.reading_tree.insert("", tk.END, values=(idx, display_month, tenant_name, 
                                                      reading.meter.meter_no, 
                                                      translated_meter_type, 
                                                      reading.previous_reading, 
                                                      reading.current_reading, 
                                                      reading.adjustment, 
                                                      reading.usage, 
                                                      remark_display, 
                                                      display_date, 
                                                      charging_time, 
                                                      reading.reader), tags=(tag,))
        except Exception as e:
            # 添加错误处理，处理数据加载失败的情况
            messagebox.showerror(self.get_text("error"), f"{self.get_text('failed_load_readings').format(str(e))}")
            # 恢复到空列表状态
            self.reading_list = []
        finally:
            # 更新记录总数标签
            self.update_total_records_label()
            # 更新统计信息
            self.update_stats()
    
    def format_date(self, date_value):
        """
        统一日期格式化函数，处理不同格式的日期值
        :param date_value: 日期值，可以是字符串、整数或日期对象
        :return: (formatted_month, formatted_date) - 格式化后的月份(YYYY-MM)和日期(YYYY-MM-DD)
        """
        from datetime import datetime
        
        # 默认返回当前月份和日期
        current_month = datetime.now().strftime("%Y-%m")
        current_date = datetime.now().strftime("%Y-%m-%d")
        
        if not date_value:
            return current_month, current_date
        
        # 将日期值转换为字符串
        date_str = str(date_value)
        
        try:
            # 处理Excel日期序列号（整数格式，如46038）
            if date_str.isdigit():
                excel_date = int(date_str)
                # Excel日期起点是1899-12-30，转换为Python日期
                python_date = datetime(1899, 12, 30) + timedelta(days=excel_date)
                return python_date.strftime("%Y-%m"), python_date.strftime("%Y-%m-%d")
            
            # 处理YYYYMMDD格式（如20231231）
            elif len(date_str) == 8 and date_str.isdigit():
                return f"{date_str[:4]}-{date_str[4:6]}", f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
            
            # 处理YYYY-MM-DD格式（如2023-12-31）
            elif len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
                return date_str[:7], date_str
            
            # 处理YYYY-MM格式（如2023-12）
            elif len(date_str) == 7 and date_str[4] == "-":
                return date_str, f"{date_str}-01"
            
            # 尝试解析其他日期格式
            else:
                # 尝试多种日期格式
                date_formats = [
                    "%Y/%m/%d",
                    "%d/%m/%Y",
                    "%m/%d/%Y",
                    "%Y.%m.%d",
                    "%d.%m.%Y",
                    "%m.%d.%Y"
                ]
                
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt)
                        return parsed_date.strftime("%Y-%m"), parsed_date.strftime("%Y-%m-%d")
                    except ValueError:
                        continue
        except Exception as e:
            # 任何解析错误都返回当前日期
            print(f"日期格式化错误: {e}, 日期值: {date_value}")
        
        # 无法解析时返回当前月份和日期
        return current_month, current_date
    
    def search_readings(self):
        """
        搜索抄表记录
        根据月份和租户进行筛选
        """
        # 清空现有列表
        for item in self.reading_tree.get_children():
            self.reading_tree.delete(item)
        
        # 获取筛选条件
        month = self.month_var.get()
        tenant_name = self.tenant_var.get()
        
        # 获取所有租户
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        tenant_id_map = {t.name: t.id for t in tenants}
        
        # 获取所有抄表记录
        all_readings = MeterReading.get_all()
        
        # 应用筛选条件
        filtered_readings = []
        for reading in all_readings:
            # 加载水电表信息（如果尚未加载）
            if not reading.meter:
                reading.load_meter_info()
            
            if not reading.meter:
                continue
            
            # 格式化日期，获取月份用于筛选
            reading_month, _ = self.format_date(reading.reading_date)
            
            # 按月份筛选
            if month and reading_month != month:
                continue
            
            # 按租户筛选
            if tenant_name:
                tenant_id = tenant_id_map.get(tenant_name)
                if reading.meter.tenant_id != tenant_id:
                    continue
            
            filtered_readings.append(reading)
        
        # 更新抄表记录列表
        self.reading_list = filtered_readings
        
        # 添加到列表控件，添加序号列
        for idx, reading in enumerate(filtered_readings, 1):
            if reading.meter:
                tenant_name = tenant_map.get(reading.meter.tenant_id, "未知租户")
                # 使用统一的日期格式化函数
                display_month, display_date = self.format_date(reading.reading_date)
                
                # 获取计费时间：根据租户ID和月份查找费用记录，使用费用记录的创建时间作为计费时间
                # 每次搜索时重新查询，确保获取最新数据，避免缓存导致数据不一致
                charging_time = ""
                charges = Charge.get_by_tenant(reading.meter.tenant_id)
                for charge in charges:
                    if charge.month == display_month:
                        if charge.create_time:
                            charging_time = charge.create_time.split(' ')[0] if ' ' in charge.create_time else charge.create_time
                        break
                
                # 处理备注字段，确保None值显示为空字符串
                remark_display = reading.remark if reading.remark is not None else ""
                # 根据索引应用奇偶行标签
                tag = 'odd' if idx % 2 != 0 else 'even'
                self.reading_tree.insert("", tk.END, values=(idx, display_month, tenant_name, 
                                                          reading.meter.meter_no, 
                                                          self.get_text(reading.meter.meter_type), 
                                                          reading.previous_reading, 
                                                          reading.current_reading, 
                                                          reading.adjustment, 
                                                          reading.usage, 
                                                          remark_display, 
                                                          display_date, 
                                                          charging_time, 
                                                          reading.reader), tags=(tag,))
        
        # 更新记录总数标签
        self.update_total_records_label()
        # 更新统计信息
        self.update_stats()
    
    def update_total_records_label(self):
        """
        更新记录总数标签
        实时显示当前列表中记录的总条数，格式为"共 X 条记录"
        """
        # 获取当前列表中的记录数量
        record_count = len(self.reading_list)
        # 更新标签文本
        self.total_records_label.config(text=self.get_text("total_records").format(record_count))
    
    def update_stats(self):
        """
        更新统计信息
        按表类型（水、电）分别统计用量总数和调整总数，并更新UI标签
        确保统计数据与当前显示的抄表记录保持一致
        """
        import logging
        logging.basicConfig(filename='statistics_update.log', level=logging.INFO, 
                            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        logger = logging.getLogger(__name__)
        
        # 记录统计更新开始
        logger.info("开始更新统计信息")
        
        # 初始化统计变量
        total_records = 0
        water_usage_total = 0.0
        water_adjustment_total = 0.0
        water_count = 0
        electricity_usage_total = 0.0
        electricity_adjustment_total = 0.0
        electricity_count = 0
        
        # 获取当前显示的数据用于统计，确保统计数据与显示数据一致
        current_display_data = []
        for item in self.reading_tree.get_children():
            values = self.reading_tree.item(item, "values")
            current_display_data.append(values)
        
        total_records = len(current_display_data)
        logger.info(f"当前显示记录数量: {total_records}")
        
        # 遍历当前显示的记录，按表类型统计数据
        for values in current_display_data:
            try:
                # 提取所需字段
                meter_type = values[4]  # 表类型
                usage = float(values[8])  # 用量
                adjustment = float(values[7])  # 调整
                
                # 根据表类型累加统计数据
                if meter_type in [self.get_text('water'), '水']:  # 支持中文和英文
                    water_usage_total += usage
                    water_adjustment_total += adjustment
                    water_count += 1
                    logger.debug(f"水表记录 - 用量: {usage}, 调整: {adjustment}, 累计用量: {water_usage_total}, 累计调整: {water_adjustment_total}")
                elif meter_type in [self.get_text('electricity'), '电']:  # 支持中文和英文
                    electricity_usage_total += usage
                    electricity_adjustment_total += adjustment
                    electricity_count += 1
                    logger.debug(f"电表记录 - 用量: {usage}, 调整: {adjustment}, 累计用量: {electricity_usage_total}, 累计调整: {electricity_adjustment_total}")
            except (ValueError, IndexError) as e:
                logger.error(f"统计数据解析错误: {str(e)}, 记录值: {values}")
                continue
        
        # 计算平均用量
        water_avg_usage = water_usage_total / water_count if water_count > 0 else 0
        electricity_avg_usage = electricity_usage_total / electricity_count if electricity_count > 0 else 0
        
        # 记录统计结果
        logger.info(f"统计结果 - 总记录数: {total_records}, 水用量: {water_usage_total:.2f}, 水调整: {water_adjustment_total:.2f}, 水平均用量: {water_avg_usage:.2f}, 电用量: {electricity_usage_total:.2f}, 电调整: {electricity_adjustment_total:.2f}, 电平均用量: {electricity_avg_usage:.2f}")
        
        # 更新统计标签文本，使用翻译键获取文本
        try:
            # 水表统计信息，包含总数、总用量、平均用量和调整总数
            water_text = f"{self.get_text('water')}表：{self.get_text('total_records').format(water_count)}, {self.get_text('usage_total')}: {water_usage_total:.2f}, {self.get_text('avg_usage')}: {water_avg_usage:.2f}, {self.get_text('adjustment_total')}: {water_adjustment_total:.2f}"
            # 电表统计信息，包含总数、总用量、平均用量和调整总数
            electricity_text = f"{self.get_text('electricity')}表：{self.get_text('total_records').format(electricity_count)}, {self.get_text('usage_total')}: {electricity_usage_total:.2f}, {self.get_text('avg_usage')}: {electricity_avg_usage:.2f}, {self.get_text('adjustment_total')}: {electricity_adjustment_total:.2f}"
            
            self.water_stats_label.config(text=water_text)
            self.electricity_stats_label.config(text=electricity_text)
            
            logger.info("统计信息更新成功")
        except Exception as e:
            logger.error(f"更新统计标签时出错: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def reset_filter(self):
        """
        重置筛选条件
        """
        self.month_var.set(datetime.now().strftime("%Y-%m"))
        self.tenant_combobox.set("")
        self.load_reading_list()
    
    def sort_column_clicked(self, column):
        """
        处理列点击事件，实现排序功能
        :param column: 被点击的列名
        """
        # 切换排序顺序
        if self.sort_column == column:
            # 同一列再次点击，切换排序顺序
            self.sort_order = "desc" if self.sort_order == "asc" else "asc"
        else:
            # 不同列点击，重置排序顺序为升序
            self.sort_column = column
            self.sort_order = "asc"
        
        # 获取当前显示的数据
        current_items = []
        for item in self.reading_tree.get_children():
            values = self.reading_tree.item(item, "values")
            current_items.append((values, item))
        
        # 根据列索引获取要排序的值
        column_index = self.reading_tree["columns"].index(column)
        
        # 定义排序函数
        def sort_key(item):
            value = item[0][column_index]
            # 尝试将数值型数据转换为float进行排序
            try:
                return float(value)
            except (ValueError, TypeError):
                # 非数值型数据直接比较
                return value
        
        # 执行排序
        current_items.sort(key=sort_key, reverse=(self.sort_order == "desc"))
        
        # 重新插入排序后的数据
        for i, (values, item) in enumerate(current_items):
            # 将item移动到指定位置
            self.reading_tree.move(item, "", i)
            # 更新序号列的值
            new_values = list(values)
            new_values[0] = i + 1  # 序号从1开始
            self.reading_tree.item(item, values=new_values)
        
        # 更新列标题，显示排序方向
        for col in self.reading_tree["columns"]:
            # 获取原始标题文本（移除之前的排序符号）
            original_text = self.reading_tree.heading(col, "text").split()[0]
            
            if col == self.sort_column:
                # 添加排序指示符号
                if self.sort_order == "asc":
                    new_text = f"{original_text} ↑"
                else:
                    new_text = f"{original_text} ↓"
            else:
                # 移除其他列的排序指示符号
                new_text = original_text
            
            # 为序号列设置标题时，不添加command参数
            if col == "serial":
                self.reading_tree.heading(col, text=new_text)
            else:
                self.reading_tree.heading(col, text=new_text, command=lambda c=col: self.sort_column_clicked(c))
        
        # 排序后更新统计信息，确保统计数据与显示数据一致
        self.update_stats()
    
    def on_reading_select(self, _event=None):
        """
        列表选择事件处理
        """
        selected_items = self.reading_tree.selection()
        if selected_items:
            item_id = selected_items[0]
            values = self.reading_tree.item(item_id, "values")
            
            # 获取当前行的所有值，用于匹配记录
            # values结构：[序号, 月份, 租户名称, 表编号, 表类型, 上次读数, 当前读数, 调整, 用量, 备注, 抄表日期, 计费时间, 抄表人]
            display_month, tenant_name, meter_no, meter_type, previous_reading, current_reading, adjustment, usage, remark, display_date, charging_time, reader = values[1:]
            
            # 查找对应的抄表记录 - 通过多字段匹配查找
            self.selected_reading = None
            for reading in self.reading_list:
                if not reading.meter:
                    reading.load_meter_info()
                
                # 使用多个字段匹配，而不是仅依赖ID
                if reading.meter and reading.meter.meter_no == meter_no and reading.meter.meter_type == meter_type:
                    # 格式化日期进行匹配
                    reading_month, reading_date = self.format_date(reading.reading_date)
                    if reading_month == display_month and reading_date == display_date:
                        self.selected_reading = reading
                        break
            
            # 如果当前列表中找不到，则通过数据库重新获取完整数据
            if not self.selected_reading:
                # 这里可以添加更复杂的数据库查询逻辑
                # 但当前场景下，我们假设self.reading_list已经包含所有需要的信息
                pass
            
            if self.selected_reading:
                self.fill_form()
            else:
                messagebox.showwarning(self.get_text("warning"), self.get_text("cannot_find_reading_info"))
        else:
            # 取消选择时，清空表单和选择状态
            self.clear_form()
    
    def fill_form(self):
        """
        填充表单
        """
        if self.selected_reading:
            # 加载关联的水电表信息
            if not self.selected_reading.meter:
                self.selected_reading.load_meter_info()
            
            # 获取租户名称
            tenant = Tenant.get_by_id(self.selected_reading.meter.tenant_id)
            if tenant:
                self.form_tenant.set(tenant.name)
                
                # 加载水电表数据
                meters = Meter.get_by_tenant(tenant.id)
                self.meters_dict = {f"{m.meter_no} ({m.meter_type})": m for m in meters}
                self.form_meter['values'] = [f"{m.meter_no} ({m.meter_type})" for m in meters]
                
                # 设置水电表
                meter_display = f"{self.selected_reading.meter.meter_no} ({self.selected_reading.meter.meter_type})"
                self.form_meter.set(meter_display)
                
                # 使用统一的日期格式化函数
                month, _ = self.format_date(self.selected_reading.reading_date)
                self.form_month.delete(0, tk.END)
                self.form_month.insert(0, month)
            
            # 填充其他字段
            self.form_previous_reading.config(state="normal")
            self.form_previous_reading.delete(0, tk.END)
            self.form_previous_reading.insert(0, self.selected_reading.previous_reading)
            self.form_previous_reading.config(state="readonly")
            
            self.form_current_reading.delete(0, tk.END)
            self.form_current_reading.insert(0, self.selected_reading.current_reading)
            
            # 填充调整值
            self.form_adjustment.delete(0, tk.END)
            self.form_adjustment.insert(0, self.selected_reading.adjustment)
            
            self.form_usage.config(state="normal")
            self.form_usage.delete(0, tk.END)
            self.form_usage.insert(0, self.selected_reading.usage)
            self.form_usage.config(state="readonly")
            
            # 使用统一的日期格式化函数处理抄表日期
            _, formatted_date = self.format_date(self.selected_reading.reading_date)
            self.form_reading_date.delete(0, tk.END)
            self.form_reading_date.insert(0, formatted_date)
            
            self.form_reader.delete(0, tk.END)
            self.form_reader.insert(0, self.selected_reading.reader)
            
            # 填充备注字段
            self.form_remark.delete("1.0", tk.END)
            # 处理备注可能为None的情况，转换为空字符串
            remark_text = self.selected_reading.remark if self.selected_reading.remark is not None else ""
            self.form_remark.insert("1.0", remark_text)
            # 更新字符数显示
            self.on_remark_change()
    
    def clear_form(self):
        """
        清空表单
        """
        self.selected_reading = None
        self.form_tenant.set("")
        self.form_meter.set("")
        self.form_meter['values'] = []
        self.form_previous_reading.config(state="normal")
        self.form_previous_reading.delete(0, tk.END)
        self.form_previous_reading.config(state="readonly")
        self.form_current_reading.delete(0, tk.END)
        self.form_adjustment.delete(0, tk.END)
        self.form_adjustment.insert(0, "0")
        self.form_usage.config(state="normal")
        self.form_usage.delete(0, tk.END)
        self.form_usage.config(state="readonly")
        self.form_reading_date.delete(0, tk.END)
        self.form_reading_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.form_reader.delete(0, tk.END)
        self.form_reader.insert(0, "admin")
        # 清空备注字段
        self.form_remark.delete("1.0", tk.END)
        # 更新字符数显示
        self.on_remark_change()
    
    def add_reading(self):
        """
        添加抄表记录
        """
        self.clear_form()
    

    
    def toggle_select_all(self):
        """
        切换全选/取消全选状态
        点击时选中或取消选中当前页面中所有可选择的抄表记录项
        """
        # 获取当前全选状态
        current_state = self.select_all_var.get()
        # 切换状态
        new_state = not current_state
        self.select_all_var.set(new_state)
        
        # 更新按钮文本
        self.select_all_btn.config(text=self.get_text("deselect_all") if new_state else self.get_text("select_all"))
        
        # 选中或取消选中所有项
        for item in self.reading_tree.get_children():
            if new_state:
                self.reading_tree.selection_add(item)
            else:
                self.reading_tree.selection_remove(item)
        
        # 处理选择事件
        if new_state:
            # 如果是全选状态，更新selected_reading为第一个选中项
            items = self.reading_tree.selection()
            if items:
                self.on_reading_select()
        else:
            # 如果是取消全选状态，清空selected_reading
            self.selected_reading = None
            self.clear_form()
    
    def delete_reading(self):
        """
        删除抄表记录
        支持单条或多条记录的批量删除
        前置检查：如果记录已计费，则禁止删除
        """
        # 获取选中的记录
        selected_items = self.reading_tree.selection()
        
        # 检查是否选择了记录
        if not selected_items:
            messagebox.showwarning("警告", "请先选择要删除的抄表记录")
            return
        
        # 收集选中记录的信息
        selected_readings = []
        cannot_delete_records = []
        
        # 检查每条选中记录的计费状态
        for item in selected_items:
            values = self.reading_tree.item(item, "values")
            
            # 查找对应的抄表记录
            for reading in self.reading_list:
                # 使用多个字段匹配查找对应的记录
                if not reading.meter:
                    reading.load_meter_info()
                
                if reading.meter:
                    # 获取显示月份
                    display_month, display_date = self.format_date(reading.reading_date)
                    
                    if (values[1] == display_month and 
                        values[3] == reading.meter.meter_no and 
                        values[4] == reading.meter.meter_type and 
                        values[10] == display_date):
                        
                        # 检查计费状态
                        is_charged = False
                        charges = Charge.get_by_tenant(reading.meter.tenant_id)
                        for charge in charges:
                            if charge.month == display_month and charge.create_time:
                                is_charged = True
                                break
                        
                        if is_charged:
                            # 已计费的记录，无法删除
                            cannot_delete_records.append(reading)
                        else:
                            # 未计费的记录，可以删除
                            selected_readings.append(reading)
                        break
        
        # 检查是否有不能删除的记录
        if cannot_delete_records:
            message = "以下记录已计费，无法删除：\n\n"
            for reading in cannot_delete_records:
                tenant = Tenant.get_by_id(reading.meter.tenant_id) if reading.meter else None
                tenant_name = tenant.name if tenant else "未知租户"
                message += f"- {tenant_name} ({reading.meter.meter_no}，{reading.meter.meter_type})\n"
            messagebox.showwarning("操作限制", message)
            
            # 如果所有选中的记录都不能删除，直接返回
            if not selected_readings:
                return
        
        # 显示批量删除确认对话框
        selected_count = len(selected_readings)
        if selected_count == 1:
            # 单条记录删除，显示详细信息
            reading = selected_readings[0]
            tenant = Tenant.get_by_id(reading.meter.tenant_id) if reading.meter else None
            tenant_name = tenant.name if tenant else "未知租户"
            
            confirm_message = f"确定要删除以下抄表记录吗？\n\n"
            confirm_message += f"租户名称: {tenant_name}\n"
            confirm_message += f"表编号: {reading.meter.meter_no}\n"
            confirm_message += f"表类型: {reading.meter.meter_type}\n"
            confirm_message += f"抄表日期: {reading.reading_date}\n"
            confirm_message += f"当前读数: {reading.current_reading}\n"
            confirm_message += f"用量: {reading.usage}\n\n"
            confirm_message += f"删除后将无法恢复，是否继续？"
        else:
            # 多条记录删除，显示数量信息
            confirm_message = f"确定要删除选中的 {selected_count} 条抄表记录吗？\n\n"
            confirm_message += "删除后将无法恢复，是否继续？"
        
        if messagebox.askyesno("确认删除", confirm_message):
            try:
                # 执行批量删除操作
                success_count = 0
                fail_count = 0
                
                for reading in selected_readings:
                    if reading.delete():
                        success_count += 1
                    else:
                        fail_count += 1
                
                # 显示操作结果
                if success_count > 0:
                    messagebox.showinfo("成功", f"成功删除 {success_count} 条抄表记录" + (f"，失败 {fail_count} 条" if fail_count > 0 else ""))
                    # 刷新列表
                    self.load_reading_list()
                    # 清空表单
                    self.clear_form()
                    # 重置全选状态
                    self.select_all_var.set(False)
                    self.select_all_btn.config(text=self.get_text("select_all"))
                else:
                    messagebox.showerror("错误", "抄表记录删除失败，可能是数据库操作错误")
            except Exception as e:
                # 添加必要的错误处理机制
                messagebox.showerror("错误", f"删除抄表记录时发生异常: {str(e)}")
    
    def save_reading(self):
        """
        保存抄表记录，实现完整的数据持久化逻辑
        """
        # 初始化操作类型，确保在所有错误情况下都有值
        operation_type = "新增"
        try:
            # 获取表单数据
            month = self.form_month.get().strip()
            tenant_name = self.form_tenant.get().strip()
            meter_display = self.form_meter.get().strip()
            current_reading_str = self.form_current_reading.get().strip()
            adjustment_str = self.form_adjustment.get().strip()
            reading_date = self.form_reading_date.get().strip()
            reader = self.form_reader.get().strip()
            # 获取备注，去除首尾空格
            remark = self.form_remark.get("1.0", tk.END).strip()
            
            # 1. 数据验证阶段：全面的数据验证
            validation_errors = []
            
            # 必填字段完整性检查
            if not month:
                validation_errors.append("请输入所属月份")
            else:
                # 月份格式验证
                try:
                    datetime.strptime(month, "%Y-%m")
                except ValueError:
                    validation_errors.append("所属月份格式不正确，应为YYYY-MM")
            
            if not tenant_name:
                validation_errors.append("请选择租户")
            
            if not meter_display:
                validation_errors.append("请选择水电表")
            
            if not current_reading_str:
                validation_errors.append("请输入当前读数")
            else:
                # 当前读数格式验证
                try:
                    current_reading = float(current_reading_str)
                    if current_reading < 0:
                        validation_errors.append("当前读数不能为负数")
                except ValueError:
                    validation_errors.append("当前读数必须是数字")
            
            # 调整值验证
            try:
                adjustment = float(adjustment_str)
            except ValueError:
                validation_errors.append("调整值必须是数字")
                adjustment = 0
            
            if not reading_date:
                validation_errors.append("请输入抄表日期")
            else:
                # 抄表日期格式验证
                try:
                    datetime.strptime(reading_date, "%Y-%m-%d")
                except ValueError:
                    validation_errors.append("抄表日期格式不正确，应为YYYY-MM-DD")
                else:
                    # 抄表日期合理性校验：不能早于上次抄表日期
                    if meter_display and meter_display in self.meters_dict:
                        meter = self.meters_dict[meter_display]
                        last_reading = MeterReading.get_last_reading_date(meter.id)
                        if last_reading and reading_date < last_reading:
                            validation_errors.append(f"抄表日期不能早于上次抄表日期({last_reading})")
                    
                    # 抄表日期合理性校验：不能晚于当前日期
                    if reading_date > datetime.now().strftime("%Y-%m-%d"):
                        validation_errors.append("抄表日期不能晚于当前日期")
            
            if not reader:
                validation_errors.append("请输入抄表人")
            
            # 验证水电表有效性
            if meter_display not in self.meters_dict:
                validation_errors.append("请选择有效的水电表")
            
            # 显示验证错误
            if validation_errors:
                error_message = "验证失败，请检查以下问题：\n\n" + "\n".join([f"• {err}" for err in validation_errors])
                messagebox.showerror("验证失败", error_message)
                return
            
            # 2. 业务规则验证
            # 获取水电表对象和上次读数
            meter = self.meters_dict[meter_display]
            tenant = self.tenants_dict[tenant_name]
            
            # 获取上次读数：使用表单中显示的上次读数，而非重新计算
            previous_reading_str = self.form_previous_reading.get().strip()
            previous_reading = float(previous_reading_str) if previous_reading_str.replace('.', '', 1).isdigit() else 0
            
            # 计算用量（包含调整值）
            current_reading = float(current_reading_str)
            usage = current_reading - previous_reading + adjustment
            
            # 获取历史用量数据，用于异常检测
            meter_readings = MeterReading.get_by_meter(meter.id, limit=6)  # 获取最近6个月的读数
            historical_usages = [r.usage for r in meter_readings if r.usage > 0]
            
            # 用量异常检测：如果历史用量稳定，当前用量与平均值相差太大则警告
            if historical_usages and len(historical_usages) >= 3:
                avg_usage = sum(historical_usages) / len(historical_usages)
                max_deviation = avg_usage * 2  # 允许两倍平均用量的偏差
                
                if usage > avg_usage + max_deviation:
                    result = messagebox.askyesno("警告", f"用量异常：当前用量({usage:.2f})远高于历史平均值({avg_usage:.2f})，是否继续保存？")
                    if not result:
                        return
                elif usage < avg_usage - max_deviation and usage > 0:
                    result = messagebox.askyesno("警告", f"用量异常：当前用量({usage:.2f})远低于历史平均值({avg_usage:.2f})，是否继续保存？")
                    if not result:
                        return
            
            # 检查用量是否异常（超过正常范围10倍）
            recent_readings = MeterReading.get_by_meter(meter.id, limit=3)
            if len(recent_readings) >= 2:
                # 计算平均用量，避免除零错误
                valid_usages = [r.usage for r in recent_readings if r.usage > 0]
                if valid_usages:  # 确保有有效的用量数据
                    avg_usage = sum(valid_usages) / len(valid_usages)
                    if avg_usage > 0 and usage > avg_usage * 10:
                        if not messagebox.askyesno("警告", f"当前用量({usage:.2f})远超过平均用量({avg_usage:.2f})的10倍，\n是否确定保存该记录？"):
                            return
            
            # 检查当前读数是否与上次读数相同
            if current_reading == previous_reading:
                if not messagebox.askyesno("警告", "当前读数与上次读数相同，\n是否确定保存该记录？"):
                    return
            
            # 3. 记录存在性判断阶段
            # 通过唯一标识符组合条件（所属月份、租户、水电表）检查记录是否存在
            # 先获取所有抄表记录
            all_readings = MeterReading.get_by_month(month)
            
            # 筛选匹配的记录
            matching_readings = []
            for reading in all_readings:
                if not reading.meter:
                    reading.load_meter_info()
                if reading.meter and reading.meter.id == meter.id:
                    matching_readings.append(reading)
            
            # 4. 数据库操作阶段
            operation_result = False
            operation_type = "新增"
            
            if self.selected_reading:
                # 更新现有记录
                operation_type = "编辑"
                
                # 更新记录字段
                self.selected_reading.meter_id = meter.id
                self.selected_reading.reading_date = reading_date
                self.selected_reading.previous_reading = previous_reading
                self.selected_reading.current_reading = current_reading
                self.selected_reading.adjustment = adjustment
                self.selected_reading.usage = usage
                self.selected_reading.reader = reader
                self.selected_reading.remark = remark
            
                if self.selected_reading.save():
                    operation_result = True
                else:
                    raise Exception("数据库更新操作失败")
            else:
                # 添加新记录
                if matching_readings:
                    # 记录已存在，更新现有记录
                    operation_type = "编辑"
                    
                    # 使用最新的一条记录
                    latest_reading = max(matching_readings, key=lambda x: x.reading_date)
                    latest_reading.meter_id = meter.id
                    latest_reading.reading_date = reading_date
                    latest_reading.previous_reading = previous_reading
                    latest_reading.current_reading = current_reading
                    latest_reading.adjustment = adjustment
                    latest_reading.usage = usage
                    latest_reading.reader = reader
                    latest_reading.remark = remark
                    
                    if latest_reading.save():
                        operation_result = True
                    else:
                        raise Exception("数据库更新操作失败")
                else:
                    # 记录不存在，执行插入操作
                    operation_type = "新增"
                    
                    # 创建新记录
                    reading = MeterReading(
                        meter_id=meter.id,
                        reading_date=reading_date,
                        current_reading=current_reading,
                        previous_reading=previous_reading,
                        usage=usage,
                        adjustment=adjustment,
                        reader=reader,
                        remark=remark
                    )
                
                    if reading.save():
                        operation_result = True
                    else:
                        raise Exception("数据库插入操作失败")
            
            # 5. 用户反馈阶段
            if operation_result:
                success_message = f"{operation_type}成功！"
                messagebox.showinfo("操作成功", success_message)
                
                # 6. 页面刷新阶段
                # 刷新抄表记录列表
                self.month_var.set(month)
                self.load_reading_list()
                
                # 重置表单为初始状态
                self.clear_form()
            else:
                # 操作失败
                error_message = f"{operation_type}失败！"
                messagebox.showerror("操作失败", error_message)
            
        except Exception as e:
            # 7. 错误处理机制：捕获并处理所有异常
            import logging
            logging.error(f"保存抄表记录失败: {str(e)}")
            
            # 向用户展示友好的错误提示
            error_message = f"{operation_type}失败！\n\n详细原因: {str(e)}"
            messagebox.showerror("操作失败", error_message)
    
    def calculate_charges(self):
        """
        计算费用
        这里简化实现，实际应该调用费用计算模块
        """
        messagebox.showinfo("提示", "费用计算功能将在后续实现")
    

    
    def export_reading_form(self):
        """
        导出抄表数据到Excel文件
        支持导出当前筛选条件下的所有抄表数据
        """
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "抄表数据"
            
            # 设置表头
            headers = ["ID", "租户名称", "表编号", "表类型", "上次读数", "当前读数", "调整值","用量", "抄表日期", "抄表人"]
            ws.append(headers)
            
            # 设置表头样式
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # 获取列字母
                # 设置表头字体和对齐方式
                ws[column + "1"].font = Font(bold=True, size=12)
                ws[column + "1"].alignment = Alignment(horizontal="center")
                # 计算列宽
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
                ws.column_dimensions[column].width = adjusted_width
            
            # 获取当前筛选条件
            month = self.month_var.get()
            tenant_name = self.tenant_var.get()
            
            # 获取所有租户
            tenants = Tenant.get_all()
            tenant_map = {t.id: t.name for t in tenants}
            tenant_id_map = {t.name: t.id for t in tenants}
            
            # 获取所有抄表记录
            all_readings = MeterReading.get_all()
            
            # 应用筛选条件
            filtered_readings = []
            for reading in all_readings:
                # 加载水电表信息（如果尚未加载）
                if not reading.meter:
                    reading.load_meter_info()
                
                if not reading.meter:
                    continue
                
                # 按月份筛选
                if month and reading.reading_date.startswith(month):
                    pass
                elif month and not reading.reading_date.startswith(month):
                    continue
                
                # 按租户筛选
                if tenant_name:
                    tenant_id = tenant_id_map.get(tenant_name)
                    if reading.meter.tenant_id != tenant_id:
                        continue
                
                filtered_readings.append(reading)
            
            # 写入数据
            for reading in filtered_readings:
                if reading.meter:
                    tenant_name = tenant_map.get(reading.meter.tenant_id, "未知租户")
                    ws.append([
                        reading.id,
                        tenant_name,
                        reading.meter.meter_no,
                        reading.meter.meter_type,
                        reading.previous_reading,
                        reading.current_reading,
                        reading.adjustment,
                        reading.usage,
                        reading.reading_date,
                        reading.reader
                    ])
            
            # 保存文件
            default_filename = f"抄表单_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                initialfile=default_filename
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("成功", f"导出成功！\n文件路径：{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出抄表数据失败: {str(e)}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def import_reading_form(self):
        """
        导入抄表数据
        支持导入Excel文件中的抄表数据，包含数据验证和进度显示
        """
        try:
            # 选择文件
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")],
                title="选择要导入的抄表Excel文件"
            )
            
            if not file_path:
                return
            
            # 根据文件扩展名选择处理方式
            if file_path.endswith('.xlsx'):
                # 使用openpyxl处理.xlsx文件
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 获取表头
                actual_headers = [cell.value for cell in ws[1]]
                
                # 获取所有数据行，忽略空行
                all_rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # 检查行是否全部为空
                    if any(cell is not None for cell in row):
                        all_rows.append(row)
                
                total_rows = len(all_rows)
                
                # 定义获取单元格值的函数
                def get_cell_value(row_idx, col_idx):
                    return all_rows[row_idx][col_idx]
            elif file_path.endswith('.xls'):
                # 使用xlrd处理.xls文件
                wb = xlrd.open_workbook(file_path)
                ws = wb.sheet_by_index(0)
                
                # 获取表头
                actual_headers = ws.row_values(0)
                
                # 获取所有数据行，忽略空行
                all_rows = []
                for row_idx in range(1, ws.nrows):  # 从第二行开始（索引1）
                    row = ws.row_values(row_idx)
                    # 检查行是否全部为空
                    if any(cell is not None and cell != '' for cell in row):
                        all_rows.append(row)
                
                total_rows = len(all_rows)
                
                # 定义获取单元格值的函数
                def get_cell_value(row_idx, col_idx):
                    return all_rows[row_idx][col_idx]
            else:
                messagebox.showerror("错误", "不支持的文件格式，请选择.xlsx或.xls文件")
                return
            
            # 验证表头
            expected_headers = ["租户名称", "表编号", "表类型", "上次读数", "当前读数","调整值", "用量", "抄表日期", "抄表人"]
            
            # 检查是否包含所有必填表头
            missing_headers = [header for header in expected_headers if header not in actual_headers]
            if missing_headers:
                messagebox.showerror("错误", f"文件格式不正确，缺少以下必填列：{', '.join(missing_headers)}")
                return
            
            # 创建进度窗口
            progress_window = tk.Toplevel(self.parent)
            progress_window.title("导入进度")
            progress_window.geometry("400x150")
            progress_window.transient(self.parent)
            progress_window.grab_set()
            
            # 进度标签
            progress_label = ttk.Label(progress_window, text=self.get_text("preparing_import"))
            progress_label.pack(pady=20)
            
            # 进度条
            progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate')
            progress_bar.pack(pady=10)
            
            # 结果标签
            result_label = ttk.Label(progress_window, text=self.get_text("import_result"))
            result_label.pack(pady=10)
            
            progress_window.update()
            
            # 初始化统计信息
            success_count = 0
            failed_count = 0
            failed_reasons = []
            
            # 获取所有租户和水电表数据
            meters = Meter.get_all()
            meter_map = {f"{m.meter_no} ({m.meter_type})": m for m in meters}
            
            # 1. 批量解析数据并进行初步验证，收集所有待导入记录
            import_records = []
            for row_idx in range(total_rows):
                try:
                    # 获取数据
                    tenant_name = str(get_cell_value(row_idx, actual_headers.index("租户名称"))).strip() if get_cell_value(row_idx, actual_headers.index("租户名称")) is not None else ""
                    meter_no = str(get_cell_value(row_idx, actual_headers.index("表编号"))).strip() if get_cell_value(row_idx, actual_headers.index("表编号")) is not None else ""
                    meter_type = get_cell_value(row_idx, actual_headers.index("表类型"))
                    previous_reading = get_cell_value(row_idx, actual_headers.index("上次读数"))
                    current_reading = get_cell_value(row_idx, actual_headers.index("当前读数"))
                    adjustment_value = get_cell_value(row_idx, actual_headers.index("调整值"))
                    reading_date = get_cell_value(row_idx, actual_headers.index("抄表日期"))
                    reader = get_cell_value(row_idx, actual_headers.index("抄表人"))
                    
                    # 数据验证
                    if not tenant_name:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx + 2}行：租户名称不能为空")
                        continue
                    
                    if not meter_no:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx + 2}行：表编号不能为空")
                        continue
                    
                    if not meter_type or meter_type not in ["水", "电"]:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx + 2}行：表类型必须是'水'或'电'")
                        continue
                    
                    # 查找水电表
                    meter_key = f"{meter_no} ({meter_type})"
                    if meter_key not in meter_map:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx + 2}行：水电表 '{meter_key}' 不存在")
                        continue
                    
                    meter = meter_map[meter_key]
                    
                    # 验证读数和调整值是否为数字
                    try:
                        previous_reading = float(previous_reading)
                        current_reading = float(current_reading)
                        adjustment_value = float(adjustment_value) if adjustment_value else 0
                    except ValueError:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx + 2}行：读数或调整值必须是数字")
                        continue
                    
                    # 计算用量（包含调整值）
                    usage = current_reading - previous_reading + adjustment_value
                    # 注意：允许负的调整值，所以不再检查 usage < 0
                    # 只检查当前读数是否小于上次读数
                    if current_reading < previous_reading:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx + 2}行：当前读数不能小于上次读数")
                        continue
                    
                    if not reading_date:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx + 2}行：抄表日期不能为空")
                        continue
                    
                    # 处理日期格式
                    if isinstance(reading_date, datetime):
                        reading_date_str = reading_date.strftime("%Y-%m-%d")
                    elif isinstance(reading_date, str):
                        reading_date_str = reading_date
                    elif hasattr(reading_date, 'value'):
                        # 处理xlrd的日期类型
                        reading_date_str = datetime(*xlrd.xldate_as_tuple(reading_date.value, wb.datemode)).strftime("%Y-%m-%d")
                    else:
                        try:
                            reading_date_str = str(reading_date)
                        except:
                            failed_count += 1
                            failed_reasons.append(f"第{row_idx + 2}行：无效的日期格式")
                            continue
                    
                    if not reader:
                        failed_count += 1
                        failed_reasons.append(f"第{row_idx + 2}行：抄表人不能为空")
                        continue
                    
                    # 获取所属月份（从抄表日期提取）
                    reading_month = reading_date_str[:7]  # 格式：YYYY-MM
                    
                    # 添加到待导入记录列表
                    import_records.append({
                        'row_idx': row_idx,
                        'tenant_name': tenant_name,
                        'meter_no': meter_no,
                        'meter_type': meter_type,
                        'reading_month': reading_month,
                        'reading_date_str': reading_date_str,
                        'previous_reading': previous_reading,
                        'current_reading': current_reading,
                        'adjustment_value': adjustment_value,
                        'usage': usage,
                        'reader': reader,
                        'meter': meter
                    })
                except Exception as e:
                    failed_count += 1
                    failed_reasons.append(f"第{row_idx + 2}行：{str(e)}")
            
            # 2. 重复导入校验
            if import_records:
                # 获取数据库中已存在的记录，构建唯一标识映射
                existing_mappings = {}  # 键：(reading_month, tenant_name, meter_no)，值：True
                
                # 获取所有租户ID到名称的映射
                tenants = Tenant.get_all()
                tenant_id_to_name = {t.id: t.name for t in tenants}
                
                # 获取所有已存在的抄表记录，按月份分组
                all_existing_readings = MeterReading.get_all()
                for reading in all_existing_readings:
                    if not reading.meter:
                        reading.load_meter_info()
                    if reading.meter:
                        # 获取月份
                        existing_month = reading.reading_date[:7]  # 格式：YYYY-MM
                        tenant_name = str(tenant_id_to_name.get(reading.meter.tenant_id, "未知租户")).strip()
                        meter_no = str(reading.meter.meter_no).strip()
                        key = (existing_month, tenant_name, meter_no)
                        existing_mappings[key] = True
                
                # 检查待导入记录是否重复
                duplicate_records = []
                for record in import_records:
                    key = (record['reading_month'], record['tenant_name'], record['meter_no'])
                    
                    if key in existing_mappings:
                        # 记录重复信息
                        duplicate_records.append({
                            'row_idx': record['row_idx'] + 2,  # 转换为Excel行号
                            'tenant_name': record['tenant_name'],
                            'meter_no': record['meter_no'],
                            'meter_type': record['meter_type'],
                            'reading_month': record['reading_month']
                        })
                
                # 如果发现重复记录，完全拒绝本次导入
                if duplicate_records:
                    # 构建错误信息
                    error_msg = "检测到重复数据，本次导入操作已取消。\n\n重复记录如下：\n"
                    error_msg += "行号 | 所属月份 | 租户名称 | 表编号 | 表类型\n"
                    error_msg += "-" * 60 + "\n"
                    for dup in duplicate_records:
                        error_msg += f"{dup['row_idx']:4d} | {dup['reading_month']} | {dup['tenant_name']} | {dup['meter_no']} | {dup['meter_type']}\n"
                    error_msg += "\n重复原因：以上记录的[所属月份]、[租户名称]和[表编号]组合已存在于系统中\n"
                    error_msg += "建议：请检查并修改数据后重新尝试导入。"
                    
                    # 关闭进度窗口
                    progress_window.destroy()
                    messagebox.showerror("导入失败", error_msg)
                    return
            
            # 3. 执行数据导入
            for record in import_records:
                try:
                    # 创建并保存抄表记录
                    reading = MeterReading(
                        meter_id=record['meter'].id,
                        reading_date=record['reading_date_str'],
                        previous_reading=record['previous_reading'],
                        current_reading=record['current_reading'],
                        adjustment=record['adjustment_value'],
                        usage=record['usage'],
                        reader=record['reader']
                    )
                    
                    if reading.save():
                        success_count += 1
                    else:
                        failed_count += 1
                        failed_reasons.append(f"第{record['row_idx'] + 2}行：保存失败")
                        continue
                except Exception as e:
                    failed_count += 1
                    failed_reasons.append(f"第{record['row_idx'] + 2}行：{str(e)}")
                
                # 更新进度
                progress = int((success_count + failed_count) / total_rows * 100)
                progress_bar['value'] = progress
                progress_label.config(text=f"正在导入第 {success_count + failed_count}/{total_rows} 行...")
                result_label.config(text=f"成功: {success_count}, 失败: {failed_count}")
                progress_window.update()
                
                # 更新进度
                progress = int((row_idx + 1) / total_rows * 100)
                progress_bar['value'] = progress
                progress_label.config(text=f"正在导入第 {row_idx + 1}/{total_rows} 行...")
                result_label.config(text=f"成功: {success_count}, 失败: {failed_count}")
                progress_window.update()
            
            # 关闭进度窗口
            progress_window.destroy()
            
            # 显示导入结果
            result_message = f"导入完成！\n\n总计: {total_rows} 行\n成功: {success_count} 行\n失败: {failed_count} 行\n"
            
            if failed_reasons:
                result_message += "\n失败原因:\n"
                result_message += "\n".join(failed_reasons[:10])  # 只显示前10条失败原因
                if len(failed_reasons) > 10:
                    result_message += f"\n... 还有 {len(failed_reasons) - 10} 条失败原因未显示\n"
            
            messagebox.showinfo("导入结果", result_message)
            
            # 刷新抄表记录列表
            self.load_reading_list()
        except Exception as e:
            messagebox.showerror("错误", f"导入抄表数据失败: {str(e)}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def export_manual_reading_form(self):
        """
        导出手抄单Excel文件
        基于上一月份的抄表数据生成手抄单
        """
        import os
        import subprocess
        from calendar import monthrange
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        try:
            # 更新UI状态，显示加载中
            self.parent.config(cursor="wait")
            self.parent.update()
            
            # 1. 月份逻辑处理
            current_month = self.month_var.get().strip()
            if not current_month:
                messagebox.showwarning("警告", "请先输入或选择月份")
                return
            
            # 解析当前月份
            try:
                current_date = datetime.strptime(current_month, "%Y-%m")
            except ValueError:
                messagebox.showwarning(self.get_text("warning"), self.get_text("date_format_error"))
                return
            
            # 计算上一月份
            if current_date.month == 1:
                # 如果是1月份，上一个月是去年12月份
                previous_month_date = current_date.replace(year=current_date.year - 1, month=12)
            else:
                # 其他月份，直接减1个月
                previous_month_date = current_date.replace(month=current_date.month - 1)
            
            previous_month = previous_month_date.strftime("%Y-%m")
            current_month_display = current_date.strftime("%Y年%m月")
            
            # 2. 获取上一月份的抄表数据
            previous_month_readings = MeterReading.get_by_month(previous_month)
            if not previous_month_readings:
                messagebox.showwarning("警告", f"未找到{previous_month}的抄表数据")
                return
            
            # 3. 数据预处理和排序
            # 获取所有租户
            tenants = Tenant.get_all()
            tenant_map = {t.id: t.name for t in tenants}
            
            # 处理抄表数据，添加租户名称
            processed_readings = []
            for reading in previous_month_readings:
                if not reading.meter:
                    reading.load_meter_info()
                
                if not reading.meter:
                    continue
                
                # 获取租户名称
                tenant_name = tenant_map.get(reading.meter.tenant_id, "未知租户")
                
                # 添加到处理列表
                processed_readings.append({
                    "tenant_name": tenant_name,
                    "meter_no": reading.meter.meter_no,
                    "meter_type": reading.meter.meter_type,
                    "previous_reading": reading.current_reading,  # 使用上一月份的当前读数
                    "current_reading": "",
                    "adjustment": "",
                    "usage": "",
                    "reading_date": "",
                    "reader": ""
                })
            
            # 按表类型和租户名称排序
            # 安装pinyin库进行中文排序
            try:
                from pypinyin import lazy_pinyin
                
                # 定义排序函数
                def sort_key(item):
                    return (lazy_pinyin(item["meter_type"]), lazy_pinyin(item["tenant_name"]))
                
                # 排序
                processed_readings.sort(key=sort_key)
            except ImportError:
                # 如果没有pinyin库，使用默认排序
                processed_readings.sort(key=lambda x: (x["meter_type"], x["tenant_name"]))
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "手抄单"
            
            # 设置纸张大小为A4竖向
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            
            # 设置页边距为最窄
            ws.page_margins.left = 0.75
            ws.page_margins.right = 0.75
            ws.page_margins.top = 0.75
            ws.page_margins.bottom = 0.75
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.3
            
            # 居中方式：居中
            ws.page_setup.horizontalCentered = True
            ws.page_setup.verticalCentered = True
            
            # 不显示页眉页脚，使用openpyxl的正确方式
            # 在openpyxl中，页眉页脚是通过PageSetup设置的，不需要单独调用setHeader和setFooter
            
            # 4. 表格标题行设置
            title = f"{current_month_display}水电费抄表单"
            # 合并单元格：A1-J1
            ws.merge_cells("A1:J1")
            # 设置标题
            title_cell = ws["A1"]
            title_cell.value = title
            # 设置标题样式
            title_cell.font = Font(bold=True, size=16, color="000000", underline="single")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            # 设置标题行高度
            ws.row_dimensions[1].height = 28
            
            # 5. 第二行单元格配置
            # A2: 抄表日期：
            ws["A2"].value = "抄表日期："
            # B2: 当前日期
            ws["B2"].value = datetime.now().strftime("%Y-%m-%d")
            # F2: 总水表读数：
            ws["F2"].value = "总水表读数："
            # 设置第二行样式
            gray_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
            second_row_font = Font(bold=True, size=11, color="000000")
            
            # A2和F2设置右对齐
            ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
            ws["A2"].font = second_row_font
            ws["A2"].fill = gray_fill
            
            # B2设置左对齐
            ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
            ws["B2"].font = second_row_font
            ws["B2"].fill = gray_fill
            
            # F2设置右对齐
            ws["F2"].alignment = Alignment(horizontal="right", vertical="center")
            ws["F2"].font = second_row_font
            ws["F2"].fill = gray_fill
            
            # 设置第二行高度
            ws.row_dimensions[2].height = 22
            
            # 6. 第三行行标题设置
            headers = ["序号", "租户名称", "表编号", "表类型", "上次读数", "当前读数", "调整值", "用量", "抄表日期", "抄表人"]
            # 写入行标题
            for col_idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                ws[f"{col_letter}3"].value = header
            
            # 设置行标题样式
            header_font = Font(bold=True, size=12, color="000000")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                                  top=Side(style="thin"), bottom=Side(style="thin"))
            
            for col_idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                cell = ws[f"{col_letter}3"]
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
                cell.fill = gray_fill
            
            # 设置第三行高度
            ws.row_dimensions[3].height = 24
            
            # 7. 隐藏列：表编号(3)、抄表日期(9)、抄表人(10)
            ws.column_dimensions["C"].hidden = True
            ws.column_dimensions["I"].hidden = True
            ws.column_dimensions["J"].hidden = True
            
            # 8. 填充数据内容
            # 设置数据行样式
            data_font = Font(size=11, color="000000")
            data_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                                top=Side(style="thin"), bottom=Side(style="thin"))
            
            # 填充数据
            for row_idx, reading in enumerate(processed_readings, 4):
                # 序号
                ws[f"A{row_idx}"].value = row_idx - 3
                ws[f"A{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
                
                # 租户名称
                ws[f"B{row_idx}"].value = reading["tenant_name"]
                ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
                
                # 表编号
                ws[f"C{row_idx}"].value = reading["meter_no"]
                ws[f"C{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
                
                # 表类型
                ws[f"D{row_idx}"].value = reading["meter_type"]
                ws[f"D{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
                
                # 上次读数
                ws[f"E{row_idx}"].value = int(reading["previous_reading"]) if reading["previous_reading"] != "" else ""
                ws[f"E{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
                
                # 当前读数
                ws[f"F{row_idx}"].value = reading["current_reading"]
                ws[f"F{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
                
                # 调整值
                ws[f"G{row_idx}"].value = reading["adjustment"]
                ws[f"G{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
                
                # 用量
                ws[f"H{row_idx}"].value = reading["usage"]
                ws[f"H{row_idx}"].alignment = Alignment(horizontal="right", vertical="center")
                
                # 抄表日期
                ws[f"I{row_idx}"].value = reading["reading_date"]
                
                # 抄表人
                ws[f"J{row_idx}"].value = reading["reader"]
                
                # 设置行高
                ws.row_dimensions[row_idx].height = 20
                
                # 设置字体和边框
                for col_idx in range(1, 11):
                    col_letter = get_column_letter(col_idx)
                    cell = ws[f"{col_letter}{row_idx}"]
                    cell.font = data_font
                    cell.border = data_border
            
            # 9. 列宽调整
            # 设置最小和最大列宽
            min_widths = {
                "A": 12,  # 序号
                "B": 16,  # 租户名称
                "D": 20,  # 表类型
                "E": 20,  # 上次读数
                "F": 20,  # 当前读数
                "G": 20,  # 调整值
                "H": 20   # 用量
            }
            max_width = 110
            
            # 计算并设置列宽
            for col_idx in range(1, 11):
                col_letter = get_column_letter(col_idx)
                
                # 跳过隐藏列
                if col_letter in ["C", "I", "J"]:
                    continue
                
                # 获取最小宽度
                min_width = min_widths.get(col_letter, 15)
                
                # 计算最大内容宽度
                max_content_width = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row_idx}"]
                    if cell.value:
                        cell_width = len(str(cell.value)) * 1.1  # 估计字符宽度
                        max_content_width = max(max_content_width, cell_width)
                
                # 设置列宽
                width = max(min_width, min(max_content_width, max_width))
                ws.column_dimensions[col_letter].width = width
            
            # 10. 文件处理
            # 创建导出文件目录
            export_dir = "导出文件"
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)
            
            # 生成文件名：YYYYMM_水电费抄表单_时间戳
            filename = f"{current_date.strftime('%Y%m')}_水电费抄表单_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            file_path = os.path.join(export_dir, filename)
            
            # 保存文件
            wb.save(file_path)
            
            # 11. 自动打开文件
            try:
                # 使用系统默认程序打开文件
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # macOS or Linux
                    subprocess.Popen(['open', file_path])
                
                messagebox.showinfo("成功", f"手抄单已生成并保存至：\n{file_path}\n\n文件已自动打开")
            except Exception as e:
                messagebox.showinfo("成功", f"手抄单已生成并保存至：\n{file_path}\n\n自动打开文件失败：{str(e)}")
            
        except Exception as e:
            messagebox.showerror("错误", f"生成手抄单失败：{str(e)}")
        finally:
            # 恢复UI状态
            self.parent.config(cursor="")
            self.parent.update()
    
    def make_hand_copy(self):
        """
        制作手抄单Excel文件
        基于上一月份的抄表数据生成手抄单
        """
        # 调用导出手抄单方法，保持向后兼容
        self.export_manual_reading_form()
