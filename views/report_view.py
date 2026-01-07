#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报表管理视图
负责处理报表生成、统计分析等功能
"""

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.fonts import addMapping
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

# 注册中文字体
def register_chinese_fonts():
    """
    注册中文字体到reportlab
    """
    # 尝试从Windows系统字体目录加载常用中文字体
    font_dirs = [
        "C:\\Windows\\Fonts",
        "C:\\Windows\\SysWOW64\\Fonts",
        "D:\\Windows\\Fonts",
        "D:\\Windows\\SysWOW64\\Fonts"
    ]
    
    font_files = {
        "SimHei": "simhei.ttf",  # 黑体
        "SimSun": "simsun.ttc",  # 宋体
        "Microsoft YaHei": "msyh.ttf"  # 微软雅黑
    }
    
    # 找到并注册第一个可用的中文字体
    for font_name, font_file in font_files.items():
        for font_dir in font_dirs:
            font_path = os.path.join(font_dir, font_file)
            if os.path.exists(font_path):
                try:
                    # 注册常规字体
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                    # 为了兼容粗体，我们将常规字体同时注册为粗体
                    pdfmetrics.registerFont(TTFont(f"{font_name}-Bold", font_path))
                    # 添加字体映射，使中文能正常显示
                    addMapping(font_name, 0, 0, font_name)  # 常规
                    addMapping(font_name, 0, 1, f"{font_name}-Bold")  # 粗体
                    return font_name
                except Exception:
                    continue
    
    # 如果没有找到系统字体，使用reportlab自带的字体或默认字体
    return "Helvetica"

# 注册中文字体并获取默认中文字体名称
DEFAULT_CHINESE_FONT = register_chinese_fonts()
from models.tenant import Tenant
from models.charge import Charge
from models.payment import Payment
from models.settlement import Settlement

# 导入matplotlib用于图表生成
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np

# 设置matplotlib使用中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']  # 设置中文显示
plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题

class ReportView:
    """报表管理视图类"""
    
    def __init__(self, parent, language_utils=None):
        """
        初始化报表管理视图
        :param parent: 父窗口组件
        :param language_utils: 语言工具类实例
        """
        self.parent = parent
        self.language_utils = language_utils
        self.create_widgets()
    
    def get_text(self, key):
        """
        获取翻译文本
        :param key: 文本键名
        :return: 翻译后的文本
        """
        return self.language_utils.get_text(key) if self.language_utils else key
        
    def update_language(self):
        """
        更新界面语言
        """
        # 更新左侧报表类型标题
        self.report_type_title.config(text=self.get_text('report_type'))
        
        # 更新报表类型选项
        self.report_type_radios['monthly'].config(text=self.get_text('monthly_report'))
        self.report_type_radios['tenant_detail'].config(text=self.get_text('tenant_detail_report'))
        self.report_type_radios['payment_stat'].config(text=self.get_text('payment_stat_report'))
        self.report_type_radios['settlement'].config(text=self.get_text('settlement_report'))
        
        # 更新查询条件框架标题
        self.query_frame.config(text=self.get_text('query_conditions'))
        
        # 更新查询条件标签
        self.query_labels['month_label'].config(text=self.get_text('month') + ':')
        self.query_labels['tenant_label'].config(text=self.get_text('tenant') + ':')
        self.query_labels['stat_type_label'].config(text=self.get_text('stat_type') + ':')
        
        # 更新统计方式选项
        self.stat_type_combobox.config(values=[self.get_text('by_tenant'), self.get_text('by_type')])
        current_stat_type = self.stat_type_var.get()
        if current_stat_type == "按租户" or current_stat_type == self.get_text('by_tenant'):
            self.stat_type_var.set(self.get_text('by_tenant'))
        elif current_stat_type == "按类型" or current_stat_type == self.get_text('by_type'):
            self.stat_type_var.set(self.get_text('by_type'))
        
        # 更新操作按钮
        self.action_buttons['generate_report_btn'].config(text=self.get_text('generate_report'))
        self.action_buttons['export_excel_btn'].config(text=self.get_text('export_excel'))
        self.action_buttons['export_pdf_btn'].config(text=self.get_text('export_pdf'))
        
        # 更新报表结果框架标题
        self.result_frame.config(text=self.get_text('report_result'))
        
        # 更新统计图表框架标题
        self.chart_frame.config(text=self.get_text('stat_chart'))
        
        # 更新图表占位符文本
        for widget in self.chart_frame.winfo_children():
            if isinstance(widget, ttk.Label):
                widget.config(text=self.get_text('chart_area'))
                break
        
        # 更新加载提示文本
        self.loading_label.config(text=self.get_text('generating_report_please_wait'))
        
        # 保存当前报表状态，用于判断是否需要重新生成图表
        has_report_content = self.report_text.get(1.0, tk.END).strip() != ""
        report_type = self.report_type.get()
        month = self.month_var.get()
        tenant_name = self.tenant_var.get()
        stat_type = self.stat_type_var.get()
        
        # 清空当前报表内容，以便用户重新生成正确语言的报表
        self.report_text.delete(1.0, tk.END)
        
        # 移除旧的图表
        for widget in self.chart_frame.winfo_children():
            widget.destroy()
        
        # 添加图表占位符
        self.chart_placeholder = ttk.Label(self.chart_frame, text=self.get_text('chart_area'))
        self.chart_placeholder.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 重新生成图表（如果当前有生成的报表）
        if has_report_content:
            # 如果报表文本不为空，重新生成图表
            self.generate_chart(report_type, month, tenant_name, stat_type)
    
    def create_widgets(self):
        """
        创建报表管理界面组件
        """
        # 先清除父容器中的所有现有组件，避免重复渲染
        for widget in self.parent.winfo_children():
            widget.destroy()
        
        # 创建主框架
        main_frame = ttk.Frame(self.parent)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 左侧：报表类型选择
        left_frame = ttk.Frame(main_frame, width=100)  # 宽度减少至当前的50%
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        # 移除pack_propagate(False)，确保文本内容能完整显示
        # left_frame.pack_propagate(False)
        
        # 确保左侧区域在不同屏幕尺寸下都能适应
        # 由于main_frame使用pack布局，我们需要调整右侧框架的expand参数
        # 通过修改右侧框架的pack配置来实现响应式布局
        
        # 报表类型标题
        self.report_type_title = ttk.Label(left_frame, text=self.get_text('report_type'), font=("Arial", 12, "bold"))
        self.report_type_title.pack(side=tk.TOP, pady=10)
        
        # 报表类型按钮组
        self.report_type = tk.StringVar()
        self.report_type_radios = {}
        
        self.report_type_radios['monthly'] = ttk.Radiobutton(left_frame, text=self.get_text('monthly_report'), variable=self.report_type, value="monthly", command=self.on_report_type_change)
        self.report_type_radios['monthly'].pack(side=tk.TOP, anchor=tk.W, padx=10, pady=5)
        
        self.report_type_radios['tenant_detail'] = ttk.Radiobutton(left_frame, text=self.get_text('tenant_detail_report'), variable=self.report_type, value="tenant_detail", command=self.on_report_type_change)
        self.report_type_radios['tenant_detail'].pack(side=tk.TOP, anchor=tk.W, padx=10, pady=5)
        
        self.report_type_radios['payment_stat'] = ttk.Radiobutton(left_frame, text=self.get_text('payment_stat_report'), variable=self.report_type, value="payment_stat", command=self.on_report_type_change)
        self.report_type_radios['payment_stat'].pack(side=tk.TOP, anchor=tk.W, padx=10, pady=5)
        
        self.report_type_radios['settlement'] = ttk.Radiobutton(left_frame, text=self.get_text('settlement_report'), variable=self.report_type, value="settlement", command=self.on_report_type_change)
        self.report_type_radios['settlement'].pack(side=tk.TOP, anchor=tk.W, padx=10, pady=5)
        
        # 设置默认选中
        self.report_type.set("monthly")
        
        # 右侧：报表查询条件和结果
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 查询条件框架
        self.query_frame = ttk.LabelFrame(right_frame, text=self.get_text('query_conditions'))
        self.query_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        
        # 存储查询条件标签
        self.query_labels = {}
        # 存储操作按钮
        self.action_buttons = {}
        
        # 月份选择
        self.query_labels['month_label'] = ttk.Label(self.query_frame, text=self.get_text('month') + ':')
        self.query_labels['month_label'].grid(row=0, column=0, sticky=tk.E, padx=5, pady=5)
        self.month_var = tk.StringVar()
        self.month_var.set(datetime.now().strftime("%Y-%m"))
        self.month_entry = ttk.Entry(self.query_frame, textvariable=self.month_var, width=10)
        self.month_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        
        # 租户选择
        self.query_labels['tenant_label'] = ttk.Label(self.query_frame, text=self.get_text('tenant') + ':')
        self.query_labels['tenant_label'].grid(row=0, column=2, sticky=tk.E, padx=5, pady=5)
        self.tenant_var = tk.StringVar()
        self.tenant_combobox = ttk.Combobox(self.query_frame, textvariable=self.tenant_var)
        self.load_tenants_to_combobox()
        self.tenant_combobox.grid(row=0, column=3, sticky=tk.W, padx=5, pady=5)
        
        # 统计方式（可选）
        self.query_labels['stat_type_label'] = ttk.Label(self.query_frame, text=self.get_text('stat_type') + ':')
        self.query_labels['stat_type_label'].grid(row=0, column=4, sticky=tk.E, padx=5, pady=5)
        self.stat_type_var = tk.StringVar()
        self.stat_type_combobox = ttk.Combobox(self.query_frame, textvariable=self.stat_type_var, values=[self.get_text('by_tenant'), self.get_text('by_type')])
        self.stat_type_combobox.grid(row=0, column=5, sticky=tk.W, padx=5, pady=5)
        self.stat_type_combobox.set(self.get_text('by_tenant'))
        # 移除统计方式变化事件绑定，改为手动触发更新
        # self.stat_type_combobox.bind("<<ComboboxSelected>>", self.on_stat_type_change)
        
        # 操作按钮 - 调整到统计方式右侧
        button_frame = ttk.Frame(self.query_frame)
        button_frame.grid(row=0, column=6, sticky=tk.E, padx=5, pady=5)
        
        self.action_buttons['generate_report_btn'] = ttk.Button(button_frame, text=self.get_text('generate_report'), command=self.generate_report)
        self.action_buttons['generate_report_btn'].pack(side=tk.LEFT, padx=5)
        
        self.action_buttons['export_excel_btn'] = ttk.Button(button_frame, text=self.get_text('export_excel'), command=self.export_excel)
        self.action_buttons['export_excel_btn'].pack(side=tk.LEFT, padx=5)
        
        self.action_buttons['export_pdf_btn'] = ttk.Button(button_frame, text=self.get_text('export_pdf'), command=self.export_pdf)
        self.action_buttons['export_pdf_btn'].pack(side=tk.LEFT, padx=5)
        
        # 确保第6列（按钮组所在列）可以伸缩，以适应不同屏幕尺寸
        self.query_frame.grid_columnconfigure(6, weight=1)
        
        # 主内容区域，用于动态布局
        self.content_frame = ttk.Frame(right_frame)
        self.content_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 报表结果框架
        self.result_frame = ttk.LabelFrame(self.content_frame, text=self.get_text('report_result'))
        
        # 报表内容区域
        self.report_text = tk.Text(self.result_frame, wrap=tk.WORD)
        self.report_text.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 统计图表区域
        self.chart_frame = ttk.LabelFrame(self.content_frame, text=self.get_text('stat_chart'))
        
        # 图表占位符
        self.chart_placeholder = ttk.Label(self.chart_frame, text=self.get_text('chart_area'))
        self.chart_placeholder.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 加载状态提示框架
        self.loading_frame = ttk.Frame(self.content_frame, relief=tk.RAISED, padding=20)
        self.loading_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.loading_frame.pack_forget()  # 初始隐藏
        
        # 加载提示标签
        self.loading_label = ttk.Label(self.loading_frame, text=self.get_text('generating_report_please_wait'), font=('Arial', 14))
        self.loading_label.pack(side=tk.TOP, pady=20)
        
        # 加载动画（简单的文本动画）
        self.loading_animation = ttk.Label(self.loading_frame, text="", font=('Arial', 12))
        self.loading_animation.pack(side=tk.TOP, pady=10)
        
        # 默认上下布局
        self.apply_layout()
        
        # 监听主内容框架大小变化，实现响应式布局
        # 使用content_frame的大小变化事件替代父窗口的Configure事件，避免影响主窗口状态栏
        self.content_frame.bind("<Configure>", self.on_window_resize)
    
    def load_tenants_to_combobox(self):
        """
        加载租户数据到筛选下拉框
        """
        tenants = Tenant.get_all()
        tenant_names = [""] + [t.name for t in tenants]
        self.tenant_combobox['values'] = tenant_names
    
    def apply_layout(self):
        """
        应用响应式布局
        根据当前内容框架宽度和报表类型决定布局方式
        """
        # 解绑所有之前的布局
        self.result_frame.pack_forget()
        self.chart_frame.pack_forget()
        
        # 获取当前内容框架宽度，而不是主窗口宽度，避免影响主窗口布局
        window_width = self.content_frame.winfo_width()
        current_report_type = self.report_type.get()
        
        # 只有收费统计报表使用左右布局
        if current_report_type == "payment_stat" and window_width >= 1024:
            # 桌面端：左右并排布局
            self.result_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
            self.chart_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        else:
            # 平板和移动端：上下堆叠布局
            # 使用fill=tk.BOTH和expand=True确保框架能适当扩展，但不会影响主窗口
            self.result_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
            self.chart_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
    
    def on_window_resize(self, event):
        """
        窗口大小变化事件处理
        优化：使用一个标志位避免频繁触发布局更新
        确保只影响当前内容框架，不影响主窗口布局
        """
        # 只响应内容框架的大小变化，不响应主窗口的大小变化
        if event.widget == self.content_frame:
            if not hasattr(self, '_resize_pending'):
                self._resize_pending = False
            
            if not self._resize_pending:
                self._resize_pending = True
                # 使用较长的延迟，避免频繁更新
                self.parent.after(500, self._apply_layout_delayed)
    
    def _apply_layout_delayed(self):
        """
        延迟应用布局
        """
        try:
            self.apply_layout()
        finally:
            self._resize_pending = False
    
    def on_report_type_change(self):
        """
        报表类型变化事件处理
        当用户点击切换报表类型时，立即清除当前页面中显示的所有报表数据和统计图表，仅保留并调整页面布局结构，不显示任何加载指示器
        """
        # 清空现有报表内容
        self.report_text.delete(1.0, tk.END)
        
        # 移除旧的图表
        for widget in self.chart_frame.winfo_children():
            widget.destroy()
        
        # 添加图表占位符
        self.chart_placeholder = ttk.Label(self.chart_frame, text=self.get_text('chart_area'))
        self.chart_placeholder.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 调整布局
        self.apply_layout()
    
    def _hide_loading(self):
        """
        隐藏加载状态
        """
        # 停止加载动画
        self.stop_loading_animation()
        # 隐藏加载状态
        self.loading_frame.pack_forget()
        # 显示结果和图表框架
        self.apply_layout()
    
    def on_stat_type_change(self):
        """
        统计方式变化事件处理
        """
        # 统计方式变化时，重新生成报表和图表
        self.generate_report()
    
    def start_loading_animation(self, counter=0):
        """
        开始加载动画
        :param counter: 动画计数器
        """
        # 显示不同数量的点，形成动画效果
        dots = "." * (counter % 4)
        self.loading_animation.config(text=dots)
        
        # 继续动画，每500毫秒更新一次
        if hasattr(self, '_loading_animation_id'):
            self.parent.after_cancel(self._loading_animation_id)
        self._loading_animation_id = self.parent.after(500, self.start_loading_animation, counter + 1)
    
    def stop_loading_animation(self):
        """
        停止加载动画
        """
        if hasattr(self, '_loading_animation_id'):
            self.parent.after_cancel(self._loading_animation_id)
            delattr(self, '_loading_animation_id')
        self.loading_animation.config(text="")
    
    def generate_report(self):
        """
        生成报表
        根据选择的报表类型和查询条件生成相应的报表
        """
        # 显示加载状态
        self.loading_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)
        # 隐藏结果和图表框架
        self.result_frame.pack_forget()
        self.chart_frame.pack_forget()
        
        # 开始加载动画
        self.start_loading_animation()
        
        # 清空现有报表内容
        self.report_text.delete(1.0, tk.END)
        
        try:
            # 延迟执行，模拟真实生成报表的耗时
            self.parent.after(500, lambda: self._generate_report())
        except Exception as e:
            # 停止加载动画
            self.stop_loading_animation()
            # 显示结果和图表框架
            self.apply_layout()
            # 记录错误
            print(f"生成报表失败: {str(e)}")
    
    def _generate_report(self):
        """
        实际生成报表的内部方法
        """
        try:
            report_type = self.report_type.get()
            month = self.month_var.get()
            tenant_name = self.tenant_var.get()
            stat_type = self.stat_type_var.get()
            
            # 根据报表类型生成不同的报表
            if report_type == "monthly":
                self.generate_monthly_report(month, tenant_name, stat_type)
            elif report_type == "tenant_detail":
                self.generate_tenant_detail_report(month, tenant_name, stat_type)
            elif report_type == "payment_stat":
                self.generate_payment_stat_report(month, tenant_name, stat_type)
            elif report_type == "settlement":
                self.generate_settlement_report(month, tenant_name, stat_type)
            
            # 生成图表
            self.generate_chart(report_type, month, tenant_name, stat_type)
        finally:
            # 停止加载动画
            self.stop_loading_animation()
            # 隐藏加载状态
            self.loading_frame.pack_forget()
            # 显示结果和图表框架
            self.apply_layout()
    
    def generate_monthly_report(self, month, tenant_name, stat_type=None):
        """
        生成月度报表
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式（按租户/按类型）
        """
        # 确保stat_type有默认值
        if stat_type is None:
            stat_type = self.get_text('by_tenant')
        
        # 获取费用数据
        charges = Charge.get_by_month(month)
        
        # 如果指定了租户，过滤数据
        if tenant_name:
            tenant = next((t for t in Tenant.get_all() if t.name == tenant_name), None)
            if tenant:
                charges = [c for c in charges if c.tenant_id == tenant.id]
        
        # 处理空数据情况
        if not charges:
            # 生成空数据报表内容
            report_content = f"{self.get_text('monthly_water_electricity_report')}\n"
            report_content += f"=" * 70+ "\n"
            report_content += f"{self.get_text('report_month')}: {month}\n"
            report_content += f"{self.get_text('stat_type')}: {stat_type}\n"
            report_content += f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            report_content += f"=" * 70+ "\n\n"
            report_content += f"{self.get_text('no_data')}\n\n"
            report_content += f"{self.get_text('water_charge')}合计: 0.00 {self.get_text('yuan')}\n"
            report_content += f"{self.get_text('electricity_charge')}合计: 0.00 {self.get_text('yuan')}\n"
            report_content += f"{self.get_text('total_charge')}合计: 0.00 {self.get_text('yuan')}\n"
            self.report_text.insert(tk.END, report_content)
            return
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        tenant_type_map = {t.id: t.type for t in tenants}
        
        # 根据统计方式分组数据
        if stat_type == self.get_text('by_type'):
            # 按租户类型分组
            type_charges = {}
            for charge in charges:
                tenant_type = tenant_type_map.get(charge.tenant_id, self.get_text('unknown_type'))
                if tenant_type not in type_charges:
                    type_charges[tenant_type] = []
                type_charges[tenant_type].append(charge)
            
            # 生成报表内容
            report_content = f"{self.get_text('monthly_water_electricity_report')}\n"
            report_content += f"=" * 70+ "\n"
            report_content += f"{self.get_text('report_month')}: {month}\n"
            report_content += f"{self.get_text('stat_type')}: {stat_type}\n"
            report_content += f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            report_content += f"=" * 70+ "\n\n"
            
            # 定义列宽和格式
            col_widths = {
                'type': 15,  # 租户类型列宽
                'count': 10,  # 户数列宽
                'water_charge': 10,  # 水费列宽
                'electricity_charge': 10,  # 电费列宽
                'total_charge': 12,  # 总费用列宽
            }
            
            # 表头格式
            header_formats = {
                'type': f'<{col_widths["type"]}',
                'count': f'>{col_widths["count"]}',
                'water_charge': f'>{col_widths["water_charge"]}',
                'electricity_charge': f'>{col_widths["electricity_charge"]}',
                'total_charge': f'>{col_widths["total_charge"]}',
            }
            
            # 数据格式
            data_formats = {
                'type': f'<{col_widths["type"]}',
                'count': f'>{col_widths["count"]}',
                'water_charge': f'>{col_widths["water_charge"]}.2f',
                'electricity_charge': f'>{col_widths["electricity_charge"]}.2f',
                'total_charge': f'>{col_widths["total_charge"]}.2f',
            }
            
            # 表头
            report_content += f"{self.get_text('tenant_type'):{header_formats['type']}}{self.get_text('households'):{header_formats['count']}}{self.get_text('water_fee')}({self.get_text('yuan')})          {self.get_text('electricity_fee')}({self.get_text('yuan')})          {self.get_text('total_charge')}({self.get_text('yuan')})\n"
            report_content += f"-" * 70+ "\n"
            
            # 统计数据
            total_water = 0
            total_electricity = 0
            total_charge = 0
            total_count = 0
            
            # 遍历分组数据
            for tenant_type, type_charge_list in type_charges.items():
                # 计算该类型的统计数据
                type_water = sum(charge.water_charge for charge in type_charge_list)
                type_electricity = sum(charge.electricity_charge for charge in type_charge_list)
                type_total = sum(charge.total_charge for charge in type_charge_list)
                type_count = len(type_charge_list)
                
                # 累加总统计数据
                total_water += type_water
                total_electricity += type_electricity
                total_charge += type_total
                total_count += type_count
                
                report_content += f"{tenant_type:{data_formats['type']}}{type_count:{data_formats['count']}}{type_water:{data_formats['water_charge']}}{type_electricity:{data_formats['electricity_charge']}}{type_total:{data_formats['total_charge']}}\n"
            
            # 统计汇总
            report_content += f"-" * 70+ "\n"
            report_content += f"{self.get_text('total'):{data_formats['type']}}{total_count:{data_formats['count']}}{total_water:{data_formats['water_charge']}}{total_electricity:{data_formats['electricity_charge']}}{total_charge:{data_formats['total_charge']}}\n\n"
            
            # 统计信息
            report_content += f"{self.get_text('statistical_information')}:\n"
            report_content += f"{self.get_text('total_tenants')}: {total_count} {self.get_text('households')}\n\n"
            
            # 使用total_charge变量存储总费用
            total_charge_amount = total_charge
        else:  # 按租户统计
            # 生成报表内容
            report_content = f"{self.get_text('monthly_water_electricity_report')}\n"
            report_content += f"=" * 65+ "\n"
            report_content += f"{self.get_text('report_month')}: {month}\n"
            report_content += f"{self.get_text('stat_type')}: {stat_type}\n"
            report_content += f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            report_content += f"=" * 65+ "\n\n"
            
            # 定义列宽和格式
            # 统一的列宽定义
            col_widths = {
                'tenant_name': 22,  # 租户名称列宽
                'water_charge': 10,  # 水费列宽
                'electricity_charge': 10,  # 电费列宽
                'total_charge': 12,  # 总费用列宽
                'status': 8  # 状态列宽
            }
            
            # 表头格式 - 使用与数据相同的对齐方式
            header_formats = {
                'tenant_name': f'<{col_widths["tenant_name"]}',  # 左对齐，与数据一致
                'water_charge': f'>{col_widths["water_charge"]}',  # 右对齐，与数据一致
                'electricity_charge': f'>{col_widths["electricity_charge"]}',  # 右对齐，与数据一致
                'total_charge': f'>{col_widths["total_charge"]}',  # 右对齐，与数据一致
                'status': f'<{col_widths["status"]}'  # 左对齐，与数据一致
            }
            
            # 数据格式（包含数值格式化）
            data_formats = {
                'tenant_name': f'<{col_widths["tenant_name"]}',  # 左对齐，宽度22
                'water_charge': f'>{col_widths["water_charge"]}.2f',  # 右对齐，宽度10，保留2位小数
                'electricity_charge': f'>{col_widths["electricity_charge"]}.2f',  # 右对齐，宽度10，保留2位小数
                'total_charge': f'>{col_widths["total_charge"]}.2f',  # 右对齐，宽度12，保留2位小数
                'status': f'<{col_widths["status"]}'  # 左对齐，宽度8
            }
            
            # 租户名称、水费、电费、总费用
            # 使用嵌套f-string确保正确的格式化
            tenant_name_header = f"{self.get_text('tenant_name')}"
            water_fee_header = f"{self.get_text('water_fee')}({self.get_text('yuan')})"
            electricity_fee_header = f"{self.get_text('electricity_fee')}({self.get_text('yuan')})"
            total_charge_header = f"{self.get_text('total_charge')}({self.get_text('yuan')})"
            status_header = f"{self.get_text('status')}"
            
            report_content += f"{tenant_name_header:{header_formats['tenant_name']}}{water_fee_header:{header_formats['water_charge']}}{electricity_fee_header:{header_formats['electricity_charge']}}{total_charge_header:{header_formats['total_charge']}}{status_header:{header_formats['status']}}\n"
            report_content += f"-" * 65+ "\n"
            
            # 统计数据
            total_water = 0
            total_electricity = 0
            total_charge_amount = 0
            paid_count = 0
            unpaid_count = 0
            
            # 遍历费用数据
            for charge in charges:
                tenant_name = tenant_map.get(charge.tenant_id, self.get_text('unknown_tenant'))
                
                # 计算已收费用
                paid_amount = sum(p.amount for p in Payment.get_by_charge(charge.id))
                # 计算应收费用
                due_amount = round(charge.total_charge - paid_amount, 2)
                
                # 动态计算状态
                if due_amount == 0:
                    status = self.get_text('paid')
                elif due_amount > 0 and due_amount < charge.total_charge:
                    status = self.get_text('partially_paid')
                elif due_amount == charge.total_charge and paid_amount == 0:
                    status = self.get_text('unpaid')
                else:
                    status = charge.status
                
                # 确保租户名称不超过列宽
                display_name = tenant_name[:20] + "..." if len(tenant_name) > 20 else tenant_name
                
                report_content += f"{display_name:{data_formats['tenant_name']}}{charge.water_charge:{data_formats['water_charge']}}{charge.electricity_charge:{data_formats['electricity_charge']}}{charge.total_charge:{data_formats['total_charge']}}{status:{data_formats['status']}}\n"
                
                # 累加统计数据
                total_water += charge.water_charge
                total_electricity += charge.electricity_charge
                total_charge_amount += charge.total_charge
                
                if status == self.get_text('paid'):
                    paid_count += 1
                else:
                    unpaid_count += 1
            
            # 统计汇总
            report_content += f"-" * 65+ "\n"
            report_content += f"{self.get_text('total'):{data_formats['tenant_name']}}{total_water:{data_formats['water_charge']}}{total_electricity:{data_formats['electricity_charge']}}{total_charge_amount:{data_formats['total_charge']}}{'':{data_formats['status']}}\n\n"
            
            # 统计信息
            report_content += f"{self.get_text('statistical_information')}:\n"
            report_content += f"{self.get_text('total_tenants')}: {len(charges)} {self.get_text('households')}\n"
            report_content += f"{self.get_text('paid_count')}: {paid_count} {self.get_text('households')}\n"
            report_content += f"{self.get_text('unpaid_count')}: {unpaid_count} {self.get_text('households')}\n"
        
        report_content += f"{self.get_text('water_charge')}合计: {total_water:.2f} {self.get_text('yuan')}\n"
        report_content += f"{self.get_text('electricity_charge')}合计: {total_electricity:.2f} {self.get_text('yuan')}\n"
        report_content += f"{self.get_text('total_charge')}合计: {total_charge_amount:.2f} {self.get_text('yuan')}\n"
        
        # 显示报表
        self.report_text.insert(tk.END, report_content)
    
    def generate_tenant_detail_report(self, month, tenant_name, stat_type=None):
        """
        生成租户明细报表
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式（按租户/按类型）
        """
        # 确保stat_type有默认值
        if stat_type is None:
            stat_type = self.get_text('by_tenant')
        
        # 获取费用数据
        charges = Charge.get_by_month(month)
        
        # 如果指定了租户，过滤数据
        if tenant_name:
            tenant = next((t for t in Tenant.get_all() if t.name == tenant_name), None)
            if tenant:
                charges = [c for c in charges if c.tenant_id == tenant.id]
        
        # 租户ID到名称和类型的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        # 租户类型使用翻译后的文本
        tenant_type_map = {t.id: self.get_text(t.type) for t in tenants}
        
        # 生成报表内容
        report_content = f"{self.get_text('tenant_water_electricity_detail_report')}\n"
        report_content += f"=" * 100 + "\n"
        report_content += f"{self.get_text('report_month')}: {month}\n"
        report_content += f"{self.get_text('stat_type')}: {stat_type}\n"
        report_content += f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        
        if tenant_name:
            report_content += f"{self.get_text('tenant_name')}: {tenant_name}\n"
        
        report_content += f"=" * 100 + "\n\n"
        
        # 根据统计方式生成不同的报表
        if stat_type == self.get_text('by_type'):
            # 按租户类型分组
            type_charges = {}
            for charge in charges:
                tenant_type = tenant_type_map.get(charge.tenant_id, self.get_text('unknown_type'))
                if tenant_type not in type_charges:
                    type_charges[tenant_type] = []
                type_charges[tenant_type].append(charge)
            
            # 定义列宽和格式
            col_widths = {
                'type': 15,  # 租户类型列宽
                'count': 10,  # 户数列宽
                'water_usage': 10,  # 水用量列宽
                'electricity_usage': 10,  # 电用量列宽
                'water_charge': 10,  # 水费列宽
                'electricity_charge': 10,  # 电费列宽
                'total_charge': 12,  # 总费用列宽
            }
            
            # 表头格式
            header_formats = {
                'type': f'<{col_widths["type"]}',
                'count': f'>{col_widths["count"]}',
                'water_usage': f'>{col_widths["water_usage"]}',
                'electricity_usage': f'>{col_widths["electricity_usage"]}',
                'water_charge': f'>{col_widths["water_charge"]}',
                'electricity_charge': f'>{col_widths["electricity_charge"]}',
                'total_charge': f'>{col_widths["total_charge"]}',
            }
            
            # 数据格式
            data_formats = {
                'type': f'<{col_widths["type"]}',
                'count': f'>{col_widths["count"]}',
                'water_usage': f'>{col_widths["water_usage"]}.2f',
                'electricity_usage': f'>{col_widths["electricity_usage"]}.2f',
                'water_charge': f'>{col_widths["water_charge"]}.2f',
                'electricity_charge': f'>{col_widths["electricity_charge"]}.2f',
                'total_charge': f'>{col_widths["total_charge"]}.2f',
            }
            
            # 表头
            water_fee_header = f"{self.get_text('water_fee')}({self.get_text('yuan')})"
            electricity_fee_header = f"{self.get_text('electricity_fee')}({self.get_text('yuan')})"
            total_charge_header = f"{self.get_text('total_charge')}({self.get_text('yuan')})"
            
            report_content += f"{self.get_text('tenant_type'):{header_formats['type']}}{self.get_text('households'):{header_formats['count']}}{self.get_text('water_usage'):{header_formats['water_usage']}}{self.get_text('electricity_usage'):{header_formats['electricity_usage']}}{water_fee_header:{header_formats['water_charge']}}{electricity_fee_header:{header_formats['electricity_charge']}}{total_charge_header:{header_formats['total_charge']}}\n"
            report_content += f"-" * 100 + "\n"
            
            # 统计数据
            total_water_usage = 0
            total_electricity_usage = 0
            total_water_charge = 0
            total_electricity_charge = 0
            total_charge = 0
            total_count = 0
            
            # 遍历分组数据
            for tenant_type, type_charge_list in type_charges.items():
                # 计算该类型的统计数据
                type_water_usage = sum(charge.water_usage for charge in type_charge_list)
                type_electricity_usage = sum(charge.electricity_usage for charge in type_charge_list)
                type_water_charge = sum(charge.water_charge for charge in type_charge_list)
                type_electricity_charge = sum(charge.electricity_charge for charge in type_charge_list)
                type_total_charge = sum(charge.total_charge for charge in type_charge_list)
                type_count = len(type_charge_list)
                
                # 累加总统计数据
                total_water_usage += type_water_usage
                total_electricity_usage += type_electricity_usage
                total_water_charge += type_water_charge
                total_electricity_charge += type_electricity_charge
                total_charge += type_total_charge
                total_count += type_count
                
                report_content += f"{tenant_type:{data_formats['type']}}{type_count:{data_formats['count']}}{type_water_usage:{data_formats['water_usage']}}{type_electricity_usage:{data_formats['electricity_usage']}}{type_water_charge:{data_formats['water_charge']}}{type_electricity_charge:{data_formats['electricity_charge']}}{type_total_charge:{data_formats['total_charge']}}\n"
            
            # 统计汇总
            report_content += f"-" * 100 + "\n"
            report_content += f"{self.get_text('total'):{data_formats['type']}}{total_count:{data_formats['count']}}{total_water_usage:{data_formats['water_usage']}}{total_electricity_usage:{data_formats['electricity_usage']}}{total_water_charge:{data_formats['water_charge']}}{total_electricity_charge:{data_formats['electricity_charge']}}{total_charge:{data_formats['total_charge']}}\n\n"
        else:  # 按租户统计
            # 定义列宽和格式
            # 统一的列宽定义
            col_widths = {
                'tenant_name': 22,  # 租户名称列宽
                'month': 8,  # 月份列宽
                'water_usage': 10,  # 水用量列宽
                'water_price': 8,  # 水单价列宽
                'water_charge': 10,  # 水费列宽
                'electricity_usage': 10,  # 电用量列宽
                'electricity_price': 8,  # 电单价列宽
                'electricity_charge': 10,  # 电费列宽
                'total_charge': 10,  # 总费用列宽
                'status': 8  # 状态列宽
            }
            
            # 表头格式 - 使用与数据相同的对齐方式
            header_formats = {
                'tenant_name': f'<{col_widths["tenant_name"]}',  # 左对齐，与数据一致
                'month': f'<{col_widths["month"]}',  # 左对齐，与数据一致
                'water_usage': f'>{col_widths["water_usage"]}',  # 右对齐，与数据一致
                'water_price': f'>{col_widths["water_price"]}',  # 右对齐，与数据一致
                'water_charge': f'>{col_widths["water_charge"]}',  # 右对齐，与数据一致
                'electricity_usage': f'>{col_widths["electricity_usage"]}',  # 右对齐，与数据一致
                'electricity_price': f'>{col_widths["electricity_price"]}',  # 右对齐，与数据一致
                'electricity_charge': f'>{col_widths["electricity_charge"]}',  # 右对齐，与数据一致
                'total_charge': f'>{col_widths["total_charge"]}',  # 右对齐，与数据一致
                'status': f'<{col_widths["status"]}'  # 左对齐，与数据一致
            }
            
            # 数据格式（包含数值格式化）
            data_formats = {
                'tenant_name': f'<{col_widths["tenant_name"]}',  # 左对齐，宽度22
                'month': f'<{col_widths["month"]}',  # 左对齐，宽度8
                'water_usage': f'>{col_widths["water_usage"]}.2f',  # 右对齐，宽度10，保留2位小数
                'water_price': f'>{col_widths["water_price"]}.2f',  # 右对齐，宽度8，保留2位小数
                'water_charge': f'>{col_widths["water_charge"]}.2f',  # 右对齐，宽度10，保留2位小数
                'electricity_usage': f'>{col_widths["electricity_usage"]}.2f',  # 右对齐，宽度10，保留2位小数
                'electricity_price': f'>{col_widths["electricity_price"]}.2f',  # 右对齐，宽度8，保留2位小数
                'electricity_charge': f'>{col_widths["electricity_charge"]}.2f',  # 右对齐，宽度10，保留2位小数
                'total_charge': f'>{col_widths["total_charge"]}.2f',  # 右对齐，宽度10，保留2位小数
                'status': f'<{col_widths["status"]}'  # 左对齐，宽度8
            }
            
            # 表头
            report_content += f"{self.get_text('tenant_name'):{header_formats['tenant_name']}}{self.get_text('month'):{header_formats['month']}}{self.get_text('water_usage'):{header_formats['water_usage']}}{self.get_text('water_price'):{header_formats['water_price']}}{self.get_text('water_charge'):{header_formats['water_charge']}}{self.get_text('electricity_usage'):{header_formats['electricity_usage']}}{self.get_text('electricity_price'):{header_formats['electricity_price']}}{self.get_text('electricity_charge'):{header_formats['electricity_charge']}}{self.get_text('total_charge'):{header_formats['total_charge']}}{self.get_text('status'):{header_formats['status']}}\n"
            report_content += f"-" * 100 + "\n"
            
            # 遍历费用数据
            for charge in charges:
                tenant_name = tenant_map.get(charge.tenant_id, "未知租户")
                
                # 计算已收费用
                paid_amount = sum(p.amount for p in Payment.get_by_charge(charge.id))
                # 计算应收费用
                due_amount = round(charge.total_charge - paid_amount, 2)
                
                # 动态计算状态
                if due_amount == 0:
                    status = self.get_text('paid')
                elif due_amount > 0 and due_amount < charge.total_charge:
                    status = self.get_text('partially_paid')
                elif due_amount == charge.total_charge and paid_amount == 0:
                    status = self.get_text('unpaid')
                else:
                    status = charge.status
                
                # 确保租户名称不超过列宽
                display_name = tenant_name[:20] + "..." if len(tenant_name) > 20 else tenant_name
                
                report_content += f"{display_name:{data_formats['tenant_name']}}{charge.month:{data_formats['month']}}{charge.water_usage:{data_formats['water_usage']}}{charge.water_price:{data_formats['water_price']}}{charge.water_charge:{data_formats['water_charge']}}"
                report_content += f"{charge.electricity_usage:{data_formats['electricity_usage']}}{charge.electricity_price:{data_formats['electricity_price']}}{charge.electricity_charge:{data_formats['electricity_charge']}}{charge.total_charge:{data_formats['total_charge']}}{status:{data_formats['status']}}\n"
        
        # 显示报表
        self.report_text.insert(tk.END, report_content)
    
    def generate_payment_stat_report(self, month, tenant_name, stat_type):
        """
        生成收费统计报表
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式
        """
        report_content = f"{self.get_text('payment_stat_report_title')}\n"
        report_content += f"=" * 60 + "\n"
        report_content += f"{self.get_text('report_month')}: {month}\n"
        report_content += f"{self.get_text('stat_type')}: {stat_type}\n"
        report_content += f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        
        if tenant_name:
            report_content += f"{self.get_text('tenant_name')}: {tenant_name}\n"
        
        report_content += f"=" * 60 + "\n\n"
        
        # 获取收费数据
        payments = Payment.get_by_month(month)
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 获取租户ID到类型的映射
        
        # 租户ID到类型的映射，使用翻译后的类型名称
        tenant_type_map = {t.id: self.get_text(t.type) for t in tenants}
        
        # 统计数据
        stat_data = {}
        total_amount = 0
        
        for payment in payments:
            # 获取支付记录关联的费用记录
            charge = payment.charge
            if not charge:
                payment.load_charge_info()
                charge = payment.charge
            
            # 直接从费用记录中获取租户ID，确保准确性
            tenant_id = charge.tenant_id if charge else 0
            
            # 获取租户名称
            current_tenant_name = tenant_map.get(tenant_id, self.get_text('unknown_tenant'))
            
            # 如果指定了租户，过滤数据
            if tenant_name and current_tenant_name != tenant_name:
                continue
            
            # 按统计方式分组
            if stat_type == self.get_text('by_tenant'):
                key = current_tenant_name
            elif stat_type == self.get_text('by_type'):
                key = tenant_type_map.get(tenant_id, self.get_text('unknown_type'))
            else:
                key = current_tenant_name
            
            # 累加金额
            stat_data[key] = stat_data.get(key, 0) + payment.amount
            total_amount += payment.amount
        
        # 生成报表内容
        report_content += f"{self.get_text('statistical_item'):<20}{self.get_text('amount_yuan'):<15}{self.get_text('percentage'):<10}\n"
        report_content += f"-" * 50 + "\n"
        
        for key, amount in stat_data.items():
            percentage = (amount / total_amount * 100) if total_amount > 0 else 0
            report_content += f"{key:<20}{amount:<15.2f}{percentage:<10.2f}%\n"
        
        report_content += f"-" * 50 + "\n"
        report_content += f"{self.get_text('total'):<20}{total_amount:<15.2f}{100.00:<10.2f}%\n"
        
        # 显示报表
        self.report_text.insert(tk.END, report_content)
    
    def generate_settlement_report(self, month, tenant_name, stat_type=None):
        """
        生成结算报表
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式（按租户/按类型）
        """
        # 确保stat_type有默认值
        if stat_type is None:
            stat_type = self.get_text('by_tenant')
        
        # 获取结算数据
        settlement = Settlement.get_by_month(month)
        
        # 获取收费数据
        payments = Payment.get_by_month(month)
        total_payment = Payment.get_total_by_month(month)
        
        # 租户ID到名称和类型的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        tenant_type_map = {t.id: t.type for t in tenants}
        
        # 根据统计方式分组收费数据
        stat_data = {}
        for payment in payments:
            # 获取支付记录关联的费用记录
            charge = payment.charge
            if not charge:
                payment.load_charge_info()
                charge = payment.charge
            
            # 直接从费用记录中获取租户ID，确保准确性
            tenant_id = charge.tenant_id if charge else 0
            
            # 获取租户名称和类型
            current_tenant_name = tenant_map.get(tenant_id, self.get_text('unknown_tenant'))
            current_tenant_type = tenant_type_map.get(tenant_id, self.get_text('unknown_type'))
            
            # 如果指定了租户，过滤数据
            if tenant_name and current_tenant_name != tenant_name:
                continue
            
            # 按统计方式分组
            if stat_type == self.get_text('by_tenant'):
                key = current_tenant_name
            elif stat_type == self.get_text('by_type'):
                key = current_tenant_type
            else:
                key = current_tenant_name
            
            # 累加金额
            stat_data[key] = stat_data.get(key, 0) + payment.amount
        
        # 生成报表内容
        report_content = f"{self.get_text('water_electricity_settlement_report')}\n"
        report_content += f"=" * 60 + "\n"
        report_content += f"{self.get_text('report_month')}: {month}\n"
        report_content += f"{self.get_text('stat_type')}: {stat_type}\n"
        report_content += f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        report_content += f"=" * 60 + "\n\n"
        
        # 收费统计
        report_content += f"{self.get_text('payment_statistics')}:\n"
        report_content += f"-" * 30 + "\n"
        report_content += f"{self.get_text('monthly_total_payment')}: {total_payment:.2f} {self.get_text('yuan')}\n"
        report_content += f"{self.get_text('payment_record_count')}: {len(payments)} {self.get_text('records')}\n\n"
        
        # 按统计方式显示详细收费情况
        if stat_data:
            report_content += f"{self.get_text('statistics_by')}{stat_type}{self.get_text('colon')}:\n"
            report_content += f"-" * 50 + "\n"
            
            # 定义列宽和格式
            col_widths = {
                'stat_item': 20,  # 统计项列宽
                'amount': 15,  # 金额列宽
                'percentage': 10  # 占比列宽
            }
            
            # 表头格式
            header_formats = {
                'stat_item': f'<{col_widths["stat_item"]}',
                'amount': f'>{col_widths["amount"]}',
                'percentage': f'>{col_widths["percentage"]}'
            }
            
            # 数据格式
            data_formats = {
                'stat_item': f'<{col_widths["stat_item"]}',
                'amount': f'>{col_widths["amount"]}.2f',
                'percentage': f'>{col_widths["percentage"]}.2f'
            }
            
            # 表头
            report_content += f"{self.get_text('statistical_item'):{header_formats['stat_item']}}{self.get_text('amount_yuan'):{header_formats['amount']}}{self.get_text('percentage'):{header_formats['percentage']}}\n"
            report_content += f"-" * 50 + "\n"
            
            # 遍历统计数据
            for key, amount in stat_data.items():
                percentage = (amount / total_payment * 100) if total_payment > 0 else 0
                report_content += f"{key:{data_formats['stat_item']}}{amount:{data_formats['amount']}}{percentage:{data_formats['percentage']}}%\n"
            
            # 汇总行
            report_content += f"-" * 50 + "\n"
            report_content += f"{self.get_text('total'):{data_formats['stat_item']}}{total_payment:{data_formats['amount']}}{100.00:{data_formats['percentage']}}%\n\n"
        
        # 结算信息
        report_content += f"{self.get_text('settlement_info')}:\n"
        report_content += f"-" * 30 + "\n"
        
        if settlement:
            report_content += f"{self.get_text('settlement_date')}: {settlement.settle_date}\n"
            report_content += f"{self.get_text('settlement_amount')}: {settlement.total_amount:.2f} {self.get_text('yuan')}\n"
            report_content += f"{self.get_text('cashier')}: {settlement.cashier}\n"
            report_content += f"{self.get_text('notes')}: {settlement.notes if settlement.notes else self.get_text('none')}\n"
        else:
            report_content += f"{self.get_text('no_settlement_this_month')}\n"
            report_content += f"{self.get_text('suggested_settlement_amount')}: {total_payment:.2f} {self.get_text('yuan')}\n"
        
        # 显示报表
        self.report_text.insert(tk.END, report_content)
    
    def export_excel(self):
        """
        导出Excel报表
        根据当前生成的报表类型和数据导出为Excel文件
        """
        report_type = self.report_type.get()
        month = self.month_var.get()
        tenant_name = self.tenant_var.get()
        stat_type = self.stat_type_var.get()
        
        # 让用户选择保存路径
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title=self.get_text('export_excel_report')
        )
        
        if not file_path:
            return
        
        try:
            # 创建Workbook对象
            wb = Workbook()
            ws = wb.active
            
            # 根据报表类型生成不同的Excel内容
            if report_type == "monthly":
                ws.title = f"{self.get_text('monthly_report')}"
                self.export_monthly_excel(ws, month, tenant_name)
            elif report_type == "tenant_detail":
                ws.title = f"{self.get_text('tenant_detail_report')}"
                self.export_tenant_detail_excel(ws, month, tenant_name)
            elif report_type == "payment_stat":
                ws.title = f"{self.get_text('payment_stat_report_title')}"
                self.export_payment_stat_excel(ws, month, tenant_name, stat_type)
            elif report_type == "settlement":
                ws.title = f"{self.get_text('water_electricity_settlement_report')}"
                self.export_settlement_excel(ws, month, tenant_name)
            
            # 保存Excel文件
            wb.save(file_path)
            messagebox.showinfo(self.get_text('success'), f"{self.get_text('excel_report_successfully_exported_to')}\n{file_path}")
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('failed_to_export_excel_report')}：{str(e)}")
    
    def export_monthly_excel(self, ws, month, tenant_name):
        """
        导出月度报表到Excel
        :param ws: Worksheet对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        """
        # 设置工作表标题
        ws.title = f"月度报表_{month}"
        
        # 表头
        header = [self.get_text('tenant_name'), f"{self.get_text('water_fee')}({self.get_text('yuan')})", f"{self.get_text('electricity_fee')}({self.get_text('yuan')})", f"{self.get_text('total_charge')}({self.get_text('yuan')})", self.get_text('status')]
        ws.append(header)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 获取费用数据
        charges = Charge.get_by_month(month)
        
        # 如果指定了租户，过滤数据
        if tenant_name:
            tenant = next((t for t in Tenant.get_all() if t.name == tenant_name), None)
            if tenant:
                charges = [c for c in charges if c.tenant_id == tenant.id]
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 填充数据
        for row_num, charge in enumerate(charges, start=2):
            # 计算已收费用
            paid_amount = sum(p.amount for p in Payment.get_by_charge(charge.id))
            # 计算应收费用
            due_amount = round(charge.total_charge - paid_amount, 2)
            
            # 动态计算状态
            if due_amount == 0:
                status = "已缴"
            elif due_amount > 0 and due_amount < charge.total_charge:
                status = "部分缴纳"
            elif due_amount == charge.total_charge and paid_amount == 0:
                status = "未缴"
            else:
                status = charge.status
            
            row_data = [
                tenant_map.get(charge.tenant_id, "未知租户"),
                charge.water_charge,
                charge.electricity_charge,
                charge.total_charge,
                status
            ]
            
            # 填充数据并设置样式
            for col_num, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加总计行
        ws.append([])
        total_row = ws.max_row + 1
        # 先创建单元格，然后设置font属性
        total_cell = ws.cell(row=total_row, column=1, value=self.get_text('total'))
        total_cell.font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{len(charges)+1})")
        ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{len(charges)+1})")
        ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{len(charges)+1})")
        
        # 设置列宽
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 12
        
        # 设置表格边框
        for row in ws.iter_rows(min_row=1, max_row=total_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = Border(left=Side(style="thin"), 
                                   right=Side(style="thin"), 
                                   top=Side(style="thin"), 
                                   bottom=Side(style="thin"))
    
    def export_tenant_detail_excel(self, ws, month, tenant_name):
        """
        导出租户明细报表到Excel
        :param ws: Worksheet对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        """
        # 设置工作表标题
        ws.title = f"租户明细报表_{month}"
        
        # 表头
        header = ["租户名称", "月份", "水用量", "水单价", "水费", "电用量", "电单价", "电费", "总费用", "状态"]
        ws.append(header)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 获取费用数据
        charges = Charge.get_by_month(month)
        
        # 如果指定了租户，过滤数据
        if tenant_name:
            tenant = next((t for t in Tenant.get_all() if t.name == tenant_name), None)
            if tenant:
                charges = [c for c in charges if c.tenant_id == tenant.id]
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 填充数据
        for row_num, charge in enumerate(charges, start=2):
            # 计算已收费用
            paid_amount = sum(p.amount for p in Payment.get_by_charge(charge.id))
            # 计算应收费用
            due_amount = round(charge.total_charge - paid_amount, 2)
            
            # 动态计算状态
            if due_amount == 0:
                status = "已缴"
            elif due_amount > 0 and due_amount < charge.total_charge:
                status = "部分缴纳"
            elif due_amount == charge.total_charge and paid_amount == 0:
                status = "未缴"
            else:
                status = charge.status
            
            row_data = [
                tenant_map.get(charge.tenant_id, "未知租户"),
                charge.month,
                charge.water_usage,
                charge.water_price,
                charge.water_charge,
                charge.electricity_usage,
                charge.electricity_price,
                charge.electricity_charge,
                charge.total_charge,
                status
            ]
            
            # 填充数据并设置样式
            for col_num, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置列宽
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 12
        
        # 设置表格边框
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10):
            for cell in row:
                cell.border = Border(left=Side(style="thin"), 
                                   right=Side(style="thin"), 
                                   top=Side(style="thin"), 
                                   bottom=Side(style="thin"))
    
    def export_payment_stat_excel(self, ws, month, tenant_name, stat_type):
        """
        导出收费统计报表到Excel
        :param ws: Worksheet对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式
        """
        # 设置工作表标题
        ws.title = f"收费统计报表_{month}"
        
        # 表头
        header = ["统计项", "金额(元)", "占比"]
        ws.append(header)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 获取收费数据
        payments = Payment.get_by_month(month)
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 费用ID到租户ID的映射
        charges = Charge.get_by_month(month)
        charge_tenant_map = {c.id: c.tenant_id for c in charges}
        
        # 租户ID到类型的映射
        tenant_type_map = {t.id: t.type for t in tenants}
        
        # 统计数据
        stat_data = {}
        total_amount = 0
        
        for payment in payments:
            tenant_id = charge_tenant_map.get(payment.charge_id, 0)
            current_tenant_name = tenant_map.get(tenant_id, "未知租户")
            
            # 如果指定了租户，过滤数据
            if tenant_name and current_tenant_name != tenant_name:
                continue
            
            # 按统计方式分组
            if stat_type == "按租户":
                key = current_tenant_name
            elif stat_type == "按类型":
                key = tenant_type_map.get(tenant_id, "未知类型")
            else:
                key = current_tenant_name
            
            stat_data[key] = stat_data.get(key, 0) + payment.amount
            total_amount += payment.amount
        
        # 填充数据
        for key, amount in stat_data.items():
            percentage = (amount / total_amount * 100) if total_amount > 0 else 0
            row_data = [key, amount, f"{percentage:.2f}%"]
            ws.append(row_data)
        
        # 添加总计行
        ws.append([])
        ws.append(["总计", total_amount, "100.00%"])
        
        # 设置A列宽度
        ws.column_dimensions["A"].width = 20
    
    def export_settlement_excel(self, ws, month, tenant_name=None):
        """
        导出结算报表到Excel
        :param ws: Worksheet对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        """
        # 设置工作表标题
        ws.title = f"结算报表_{month}"
        
        # 获取结算数据
        settlement = Settlement.get_by_month(month)
        
        # 获取收费数据
        payments = Payment.get_by_month(month)
        total_payment = Payment.get_total_by_month(month)
        
        # 填充结算信息
        ws.append(["结算报表"])
        ws.append(["报表月份", month])
        ws.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        
        ws.append(["收费统计"])
        ws.append(["当月收费总金额", f"{total_payment:.2f} {self.get_text('yuan')}"])
        ws.append(["收费记录条数", f"{len(payments)} 条"])
        ws.append([])
        
        ws.append(["结算信息"])
        if settlement:
            ws.append(["结算日期", settlement.settle_date])
            ws.append(["结算金额", f"{settlement.total_amount:.2f} {self.get_text('yuan')}"])
            ws.append(["出纳姓名", settlement.cashier])
            ws.append(["备注", settlement.notes if settlement.notes else "无"])
        else:
            ws.append(["本月尚未结算"])
            ws.append(["建议结算金额", f"{total_payment:.2f} {self.get_text('yuan')}"])
    
    def export_pdf(self):
        """
        导出PDF报表
        根据当前生成的报表类型和数据导出为PDF文件
        实现自动生成、按指定格式命名、保存到指定目录、优化样式格式以及自动打开PDF文件的功能
        """
        report_type = self.report_type.get()
        month = self.month_var.get()
        tenant_name = self.tenant_var.get()
        stat_type = self.stat_type_var.get()
        
        # 定义报表类型映射，用于文件命名
        report_type_map = {
            "monthly": self.get_text('monthly_report'),
            "tenant_detail": self.get_text('tenant_detail_report'),
            "payment_stat": self.get_text('payment_stat_report'),
            "settlement": self.get_text('settlement_report')
        }
        
        # 获取报表类型名称
        report_type_name = report_type_map.get(report_type, self.get_text('report'))
        
        # 生成时间戳
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        
        # 构建文件名：报表类型_月份_统计方式_时间戳
        filename = f"{report_type_name}_{month}_{stat_type}_{timestamp}.pdf"
        
        # 创建导出文件目录
        export_dir = os.path.join(os.getcwd(), self.get_text('export_files'))
        os.makedirs(export_dir, exist_ok=True)
        
        # 构建完整文件路径
        file_path = os.path.join(export_dir, filename)
        
        try:
            # 创建canvas对象，设置页边距和字体
            c = canvas.Canvas(file_path, pagesize=A4)
            
            # 根据报表类型生成不同的PDF内容
            if report_type == "monthly":
                self.export_monthly_pdf(c, month, tenant_name)
            elif report_type == "tenant_detail":
                self.export_tenant_detail_pdf(c, month, tenant_name)
            elif report_type == "payment_stat":
                self.export_payment_stat_pdf(c, month, tenant_name, stat_type)
            elif report_type == "settlement":
                self.export_settlement_pdf(c, month, tenant_name)
            
            # 保存PDF文件
            c.save()
            
            # 自动打开PDF文件
            os.startfile(file_path)
            
            messagebox.showinfo(self.get_text('success'), f"{self.get_text('pdf_report_successfully_exported_to')}\n{file_path}")
        except Exception as e:
            messagebox.showerror(self.get_text('error'), f"{self.get_text('failed_to_export_pdf_report')}：{str(e)}")
    
    def export_monthly_pdf(self, c, month, tenant_name):
        """
        导出月度报表到PDF
        :param c: Canvas对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        """
        # 页面设置
        page_width, page_height = A4
        margin = 50
        
        # 设置标题
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 16)
        title = self.get_text('monthly_water_electricity_report')
        title_width = c.stringWidth(title, f"{DEFAULT_CHINESE_FONT}-Bold", 16)
        c.drawString((page_width - title_width) / 2, page_height - margin - 20, title)
        
        # 设置基本信息
        c.setFont(DEFAULT_CHINESE_FONT, 12)
        y = page_height - margin - 40
        c.drawString(margin, y, f"{self.get_text('report_month')}: {month}")
        
        y -= 15
        c.drawString(margin, y, f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        if tenant_name:
            y -= 15
            c.drawString(margin, y, f"{self.get_text('tenant')}: {tenant_name}")
        
        # 画分隔线
        y -= 15
        c.line(margin, y, page_width - margin, y)
        
        # 设置表头
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
        headers = [self.get_text('tenant_name'), f"{self.get_text('water_fee')}({self.get_text('yuan')})", f"{self.get_text('electricity_fee')}({self.get_text('yuan')})", f"{self.get_text('total_charge')}({self.get_text('yuan')})", self.get_text('status')]
        # 调整列宽，使布局更合理
        col_widths = [150, 80, 80, 80, 80]
        x_positions = [margin]
        for i in range(1, len(col_widths)):
            x_positions.append(x_positions[i-1] + col_widths[i-1])
        
        y -= 20
        for i, header in enumerate(headers):
            # 计算列标题居中位置
            header_width = c.stringWidth(header, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
            c.drawString(x_positions[i] + (col_widths[i] - header_width) / 2, y, header)
        
        y -= 15
        # 画分隔线
        c.line(margin, y, page_width - margin, y)
        
        # 获取费用数据
        charges = Charge.get_by_month(month)
        
        # 如果指定了租户，过滤数据
        if tenant_name:
            tenant = next((t for t in Tenant.get_all() if t.name == tenant_name), None)
            if tenant:
                charges = [c for c in charges if c.tenant_id == tenant.id]
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 设置数据字体
        c.setFont(DEFAULT_CHINESE_FONT, 10)
        
        # 填充数据
        for charge in charges:
            y -= 15
            # 确保不超过页边距
            if y < margin + 50:
                # 新建页面
                c.showPage()
                # 重新设置字体
                c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
                # 重新绘制表头
                for i, header in enumerate(headers):
                    header_width = c.stringWidth(header, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
                    c.drawString(x_positions[i] + (col_widths[i] - header_width) / 2, page_height - margin - 40, header)
                
                y = page_height - margin - 55
                c.line(margin, y, page_width - margin, y)
                y -= 15
                c.setFont(DEFAULT_CHINESE_FONT, 10)
            
            tenant_name = tenant_map.get(charge.tenant_id, "未知租户")
            
            # 计算已收费用
            paid_amount = sum(p.amount for p in Payment.get_by_charge(charge.id))
            # 计算应收费用
            due_amount = round(charge.total_charge - paid_amount, 2)
            
            # 动态计算状态
            if due_amount == 0:
                status = "已缴"
            elif due_amount > 0 and due_amount < charge.total_charge:
                status = "部分缴纳"
            elif due_amount == charge.total_charge and paid_amount == 0:
                status = "未缴"
            else:
                status = charge.status
            
            # 租户名称（左对齐）
            c.drawString(x_positions[0], y, tenant_name[:20] + "..." if len(tenant_name) > 20 else tenant_name)
            
            # 水费（右对齐）
            water_str = f"{charge.water_charge:.2f}"
            water_width = c.stringWidth(water_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[1] + col_widths[1] - water_width - 5, y, water_str)
            
            # 电费（右对齐）
            electricity_str = f"{charge.electricity_charge:.2f}"
            electricity_width = c.stringWidth(electricity_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[2] + col_widths[2] - electricity_width - 5, y, electricity_str)
            
            # 总费用（右对齐）
            total_str = f"{charge.total_charge:.2f}"
            total_width = c.stringWidth(total_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[3] + col_widths[3] - total_width - 5, y, total_str)
            
            # 状态（居中）
            status_width = c.stringWidth(status, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[4] + (col_widths[4] - status_width) / 2, y, status)
        
        # 画分隔线
        y -= 15
        c.line(margin, y, page_width - margin, y)
        
        # 计算总计
        total_water = sum([c.water_charge for c in charges])
        total_electricity = sum([c.electricity_charge for c in charges])
        total_charge = sum([c.total_charge for c in charges])
        
        # 绘制总计行
        y -= 15
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
        
        c.drawString(x_positions[0], y, "合计")
        
        total_water_str = f"{total_water:.2f}"
        total_water_width = c.stringWidth(total_water_str, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
        c.drawString(x_positions[1] + col_widths[1] - total_water_width - 5, y, total_water_str)
        
        total_electricity_str = f"{total_electricity:.2f}"
        total_electricity_width = c.stringWidth(total_electricity_str, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
        c.drawString(x_positions[2] + col_widths[2] - total_electricity_width - 5, y, total_electricity_str)
        
        total_charge_str = f"{total_charge:.2f}"
        total_charge_width = c.stringWidth(total_charge_str, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
        c.drawString(x_positions[3] + col_widths[3] - total_charge_width - 5, y, total_charge_str)
    
    def export_tenant_detail_pdf(self, c, month, tenant_name):
        """
        导出租户明细报表到PDF
        :param c: Canvas对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        """
        # 页面设置
        page_width, page_height = A4
        margin = 50
        
        # 设置标题
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 16)
        title = self.get_text('tenant_water_electricity_detail_report')
        title_width = c.stringWidth(title, f"{DEFAULT_CHINESE_FONT}-Bold", 16)
        c.drawString((page_width - title_width) / 2, page_height - margin - 20, title)
        
        # 设置基本信息
        c.setFont(DEFAULT_CHINESE_FONT, 12)
        y = page_height - margin - 40
        c.drawString(margin, y, f"{self.get_text('report_month')}: {month}")
        
        y -= 15
        c.drawString(margin, y, f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        if tenant_name:
            y -= 15
            c.drawString(margin, y, f"{self.get_text('tenant')}: {tenant_name}")
        
        # 画分隔线
        y -= 15
        c.line(margin, y, page_width - margin, y)
        
        # 设置表头
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
        headers = [self.get_text('tenant_name'), self.get_text('month'), self.get_text('water_usage'), self.get_text('water_unit_price'), self.get_text('water_fee'), self.get_text('electricity_usage'), self.get_text('electricity_unit_price'), self.get_text('electricity_fee'), self.get_text('total_charge'), self.get_text('status')]
        # 调整列宽，使布局更合理
        col_widths = [120, 60, 60, 60, 60, 60, 60, 60, 80, 60]
        x_positions = [margin]
        for i in range(1, len(col_widths)):
            x_positions.append(x_positions[i-1] + col_widths[i-1])
        
        y -= 20
        for i, header in enumerate(headers):
            # 计算列标题居中位置
            header_width = c.stringWidth(header, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
            c.drawString(x_positions[i] + (col_widths[i] - header_width) / 2, y, header)
        
        y -= 15
        # 画分隔线
        c.line(margin, y, page_width - margin, y)
        
        # 获取费用数据
        charges = Charge.get_by_month(month)
        
        # 如果指定了租户，过滤数据
        if tenant_name:
            tenant = next((t for t in Tenant.get_all() if t.name == tenant_name), None)
            if tenant:
                charges = [c for c in charges if c.tenant_id == tenant.id]
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 设置数据字体
        c.setFont(DEFAULT_CHINESE_FONT, 10)
        
        # 填充数据
        for charge in charges:
            y -= 15
            # 确保不超过页边距
            if y < margin + 50:
                # 新建页面
                c.showPage()
                # 重新设置字体
                c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
                # 重新绘制表头
                for i, header in enumerate(headers):
                    header_width = c.stringWidth(header, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
                    c.drawString(x_positions[i] + (col_widths[i] - header_width) / 2, page_height - margin - 40, header)
                
                y = page_height - margin - 55
                c.line(margin, y, page_width - margin, y)
                y -= 15
                c.setFont(DEFAULT_CHINESE_FONT, 10)
            
            tenant_name = tenant_map.get(charge.tenant_id, "未知租户")
            
            # 计算已收费用
            paid_amount = sum(p.amount for p in Payment.get_by_charge(charge.id))
            # 计算应收费用
            due_amount = round(charge.total_charge - paid_amount, 2)
            
            # 动态计算状态
            if due_amount == 0:
                status = "已缴"
            elif due_amount > 0 and due_amount < charge.total_charge:
                status = "部分缴纳"
            elif due_amount == charge.total_charge and paid_amount == 0:
                status = "未缴"
            else:
                status = charge.status
            
            # 租户名称（左对齐）
            c.drawString(x_positions[0], y, tenant_name[:15] + "..." if len(tenant_name) > 15 else tenant_name)
            
            # 月份（居中）
            month_width = c.stringWidth(charge.month, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[1] + (col_widths[1] - month_width) / 2, y, charge.month)
            
            # 水用量（右对齐）
            water_usage_str = f"{charge.water_usage:.2f}"
            water_usage_width = c.stringWidth(water_usage_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[2] + col_widths[2] - water_usage_width - 5, y, water_usage_str)
            
            # 水单价（右对齐）
            water_price_str = f"{charge.water_price:.2f}"
            water_price_width = c.stringWidth(water_price_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[3] + col_widths[3] - water_price_width - 5, y, water_price_str)
            
            # 水费（右对齐）
            water_charge_str = f"{charge.water_charge:.2f}"
            water_charge_width = c.stringWidth(water_charge_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[4] + col_widths[4] - water_charge_width - 5, y, water_charge_str)
            
            # 电用量（右对齐）
            electricity_usage_str = f"{charge.electricity_usage:.2f}"
            electricity_usage_width = c.stringWidth(electricity_usage_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[5] + col_widths[5] - electricity_usage_width - 5, y, electricity_usage_str)
            
            # 电单价（右对齐）
            electricity_price_str = f"{charge.electricity_price:.2f}"
            electricity_price_width = c.stringWidth(electricity_price_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[6] + col_widths[6] - electricity_price_width - 5, y, electricity_price_str)
            
            # 电费（右对齐）
            electricity_charge_str = f"{charge.electricity_charge:.2f}"
            electricity_charge_width = c.stringWidth(electricity_charge_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[7] + col_widths[7] - electricity_charge_width - 5, y, electricity_charge_str)
            
            # 总费用（右对齐）
            total_charge_str = f"{charge.total_charge:.2f}"
            total_charge_width = c.stringWidth(total_charge_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[8] + col_widths[8] - total_charge_width - 5, y, total_charge_str)
            
            # 状态（居中）
            status_width = c.stringWidth(status, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[9] + (col_widths[9] - status_width) / 2, y, status)
    
    def export_payment_stat_pdf(self, c, month, tenant_name, stat_type):
        """
        导出收费统计报表到PDF
        :param c: Canvas对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式
        """
        # 页面设置
        page_width, page_height = A4
        margin = 50
        
        # 设置标题
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 16)
        title = self.get_text('payment_stat_report')
        title_width = c.stringWidth(title, f"{DEFAULT_CHINESE_FONT}-Bold", 16)
        c.drawString((page_width - title_width) / 2, page_height - margin - 20, title)
        
        # 设置基本信息
        c.setFont(DEFAULT_CHINESE_FONT, 12)
        y = page_height - margin - 40
        c.drawString(margin, y, f"{self.get_text('report_month')}: {month}")
        
        y -= 15
        c.drawString(margin, y, f"{self.get_text('stat_type')}: {stat_type}")
        
        y -= 15
        c.drawString(margin, y, f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        if tenant_name:
            y -= 15
            c.drawString(margin, y, f"{self.get_text('tenant')}: {tenant_name}")
        
        # 画分隔线
        y -= 15
        c.line(margin, y, page_width - margin, y)
        
        # 获取收费数据
        payments = Payment.get_by_month(month)
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 费用ID到租户ID的映射
        charges = Charge.get_by_month(month)
        charge_tenant_map = {c.id: c.tenant_id for c in charges}
        
        # 租户ID到类型的映射
        tenant_type_map = {t.id: t.type for t in tenants}
        
        # 统计数据
        stat_data = {}
        total_amount = 0
        
        for payment in payments:
            tenant_id = charge_tenant_map.get(payment.charge_id, 0)
            current_tenant_name = tenant_map.get(tenant_id, "未知租户")
            
            # 如果指定了租户，过滤数据
            if tenant_name and current_tenant_name != tenant_name:
                continue
            
            # 按统计方式分组
            if stat_type == "按租户":
                key = current_tenant_name
            elif stat_type == "按类型":
                key = tenant_type_map.get(tenant_id, "未知类型")
            else:
                key = current_tenant_name
            
            stat_data[key] = stat_data.get(key, 0) + payment.amount
            total_amount += payment.amount
        
        # 设置表头
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
        headers = [self.get_text('stat_item'), f"{self.get_text('amount')}({self.get_text('yuan')})"]
        # 调整列宽，使布局更合理
        col_widths = [200, 100, 80]
        x_positions = [margin]
        for i in range(1, len(col_widths)):
            x_positions.append(x_positions[i-1] + col_widths[i-1])
        
        y -= 20
        for i, header in enumerate(headers):
            # 计算列标题居中位置
            header_width = c.stringWidth(header, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
            c.drawString(x_positions[i] + (col_widths[i] - header_width) / 2, y, header)
        
        y -= 15
        # 画分隔线
        c.line(margin, y, page_width - margin, y)
        
        # 设置数据字体
        c.setFont(DEFAULT_CHINESE_FONT, 10)
        
        # 填充数据
        for key, amount in stat_data.items():
            y -= 15
            # 确保不超过页边距
            if y < margin + 50:
                # 新建页面
                c.showPage()
                # 重新设置字体
                c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
                # 重新绘制表头
                for i, header in enumerate(headers):
                    header_width = c.stringWidth(header, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
                    c.drawString(x_positions[i] + (col_widths[i] - header_width) / 2, page_height - margin - 40, header)
                
                y = page_height - margin - 55
                c.line(margin, y, page_width - margin, y)
                y -= 15
                c.setFont(DEFAULT_CHINESE_FONT, 10)
            
            percentage = (amount / total_amount * 100) if total_amount > 0 else 0
            
            # 统计项（左对齐）
            c.drawString(x_positions[0], y, key[:30] + "..." if len(key) > 30 else key)
            
            # 金额（右对齐）
            amount_str = f"{amount:.2f}"
            amount_width = c.stringWidth(amount_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[1] + col_widths[1] - amount_width - 5, y, amount_str)
            
            # 占比（右对齐）
            percentage_str = f"{percentage:.2f}%"
            percentage_width = c.stringWidth(percentage_str, DEFAULT_CHINESE_FONT, 10)
            c.drawString(x_positions[2] + col_widths[2] - percentage_width - 5, y, percentage_str)
        
        # 画分隔线
        y -= 15
        c.line(margin, y, page_width - margin, y)
        
        # 绘制总计行
        y -= 15
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
        
        c.drawString(x_positions[0], y, "合计")
        
        # 总计金额（右对齐）
        total_str = f"{total_amount:.2f}"
        total_width = c.stringWidth(total_str, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
        c.drawString(x_positions[1] + col_widths[1] - total_width - 5, y, total_str)
        
        # 总计占比（右对齐）
        c.drawString(x_positions[2] + col_widths[2] - c.stringWidth("100.00%", f"{DEFAULT_CHINESE_FONT}-Bold", 10) - 5, y, "100.00%")
    
    def export_settlement_pdf(self, c, month, tenant_name):
        """
        导出结算报表到PDF
        :param c: Canvas对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        """
        # 页面设置
        page_width, page_height = A4
        margin = 50
        
        # 设置标题
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 16)
        title = self.get_text('water_electricity_settlement_report')
        title_width = c.stringWidth(title, f"{DEFAULT_CHINESE_FONT}-Bold", 16)
        c.drawString((page_width - title_width) / 2, page_height - margin - 20, title)
        
        # 设置基本信息
        c.setFont(DEFAULT_CHINESE_FONT, 12)
        y = page_height - margin - 40
        c.drawString(margin, y, f"{self.get_text('report_month')}: {month}")
        
        y -= 15
        c.drawString(margin, y, f"{self.get_text('generate_time')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # 画分隔线
        y -= 15
        c.line(margin, y, page_width - margin, y)
        
        # 获取结算数据
        settlement = Settlement.get_by_month(month)
        
        # 获取收费数据
        payments = Payment.get_by_month(month)
        total_payment = Payment.get_total_by_month(month)
        
        # 收费统计
        y -= 20
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 12)
        c.drawString(margin, y, self.get_text('payment_statistics'))
        
        y -= 15
        c.setFont(DEFAULT_CHINESE_FONT, 10)
        c.drawString(margin + 10, y, f"{self.get_text('total_monthly_payment')}: {total_payment:.2f} {self.get_text('yuan')}")
        
        y -= 15
        c.drawString(margin + 10, y, f"{self.get_text('payment_records')}: {len(payments)} {self.get_text('items')}")
        
        # 按统计方式显示详细收费情况
        # 获取收费数据
        payments = Payment.get_by_month(month)
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 费用ID到租户ID的映射
        charges = Charge.get_by_month(month)
        charge_tenant_map = {c.id: c.tenant_id for c in charges}
        
        # 租户ID到类型的映射
        tenant_type_map = {t.id: t.type for t in tenants}
        
        # 统计数据
        stat_data = {}
        for payment in payments:
            tenant_id = charge_tenant_map.get(payment.charge_id, 0)
            current_tenant_name = tenant_map.get(tenant_id, self.get_text('unknown_tenant'))
            current_tenant_type = tenant_type_map.get(tenant_id, self.get_text('unknown_type'))
            
            # 按租户类型统计
            key = current_tenant_type
            if key not in stat_data:
                stat_data[key] = 0
            stat_data[key] += payment.amount
        
        if stat_data:
            y -= 20
            c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 12)
            c.drawString(margin, y, f"{self.get_text('stat_by_tenant_type')}:")
            
            # 设置表头
            y -= 15
            c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
            headers = [self.get_text('tenant_type'), f"{self.get_text('amount')}({self.get_text('yuan')})"]
            # 调整列宽，使布局更合理
            col_widths = [120, 100, 80]
            x_positions = [margin + 10]
            for i in range(1, len(col_widths)):
                x_positions.append(x_positions[i-1] + col_widths[i-1])
            
            for i, header in enumerate(headers):
                # 计算列标题居中位置
                header_width = c.stringWidth(header, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
                c.drawString(x_positions[i] + (col_widths[i] - header_width) / 2, y, header)
            
            y -= 15
            # 画分隔线
            c.line(margin + 10, y, page_width - margin, y)
            
            # 设置数据字体
            c.setFont(DEFAULT_CHINESE_FONT, 10)
            
            # 填充数据
            for key, amount in stat_data.items():
                y -= 15
                percentage = (amount / total_payment * 100) if total_payment > 0 else 0
                
                # 租户类型（左对齐）
                c.drawString(x_positions[0], y, key)
                
                # 金额（右对齐）
                amount_str = f"{amount:.2f}"
                amount_width = c.stringWidth(amount_str, DEFAULT_CHINESE_FONT, 10)
                c.drawString(x_positions[1] + col_widths[1] - amount_width - 5, y, amount_str)
            
            # 画分隔线
            y -= 15
            c.line(margin + 10, y, page_width - margin, y)
            
            # 绘制总计行
            y -= 15
            c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 10)
            
            c.drawString(x_positions[0], y, self.get_text('total'))
            
            # 总计金额（右对齐）
            total_str = f"{total_payment:.2f}"
            total_width = c.stringWidth(total_str, f"{DEFAULT_CHINESE_FONT}-Bold", 10)
            c.drawString(x_positions[1] + col_widths[1] - total_width - 5, y, total_str)
        
        # 结算信息
        y -= 30
        c.setFont(f"{DEFAULT_CHINESE_FONT}-Bold", 12)
        c.drawString(margin, y, self.get_text('settlement_info'))
        
        y -= 15
        c.setFont(DEFAULT_CHINESE_FONT, 10)
        if settlement:
            c.drawString(margin + 10, y, f"{self.get_text('settlement_date')}: {settlement.settle_date}")
            
            y -= 15
            c.drawString(margin + 10, y, f"{self.get_text('settlement_amount')}: {settlement.total_amount:.2f} {self.get_text('yuan')}")
            
            y -= 15
            c.drawString(margin + 10, y, f"{self.get_text('cashier_name')}: {settlement.cashier}")
            
            y -= 15
            c.drawString(margin + 10, y, f"{self.get_text('remarks')}: {settlement.notes if settlement.notes else self.get_text('none')}")
        else:
            c.drawString(margin + 10, y, self.get_text('no_settlement_this_month'))
            
            y -= 15
            c.drawString(margin + 10, y, f"{self.get_text('suggested_settlement_amount')}: {total_payment:.2f} {self.get_text('yuan')}")
    
    def generate_chart(self, report_type, month, tenant_name, stat_type):
        """
        生成图表
        :param report_type: 报表类型
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式
        """
        # 移除旧的图表和错误信息
        for widget in self.chart_frame.winfo_children():
            widget.destroy()
        
        try:
            # 创建一个新的matplotlib图表
            fig, ax = plt.subplots(figsize=(8, 4))
            
            if report_type == "monthly":
                # 生成月度报表图表（柱状图）
                self.generate_monthly_chart(ax, month, tenant_name, stat_type)
            elif report_type == "tenant_detail":
                # 生成租户明细报表图表（柱状图）
                self.generate_tenant_detail_chart(ax, month, tenant_name, stat_type)
            elif report_type == "payment_stat":
                # 生成收费统计图表（饼图或柱状图）
                self.generate_payment_stat_chart(ax, month, tenant_name, stat_type)
            elif report_type == "settlement":
                # 生成结算报表图表（柱状图）
                self.generate_settlement_chart(ax, month, stat_type)
            
            # 将图表添加到Tkinter窗口
            canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # 设置窗口大小变化时，图表自动调整
            self.parent.bind("<Configure>", lambda event: self.resize_chart(canvas))
        except Exception as e:
            # 如果生成图表失败，显示友好的错误信息
            error_label = ttk.Label(self.chart_frame, text=f"图表生成失败: {str(e)}", foreground="red", wraplength=400)
            error_label.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # 记录详细错误信息到控制台
            print(f"图表生成失败: {str(e)}")
    
    def resize_chart(self, canvas):
        """
        窗口大小变化时，调整图表大小
        :param canvas: 图表画布对象
        """
        try:
            canvas.draw_idle()
        except Exception:
            pass  # 忽略调整大小时的异常
    
    def generate_monthly_chart(self, ax, month, tenant_name, stat_type="按租户"):
        """
        生成月度报表图表（柱状图）
        :param ax: matplotlib轴对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式（按租户/按类型）
        """
        # 获取费用数据
        charges = Charge.get_by_month(month)
        
        # 如果指定了租户，过滤数据
        if tenant_name:
            tenant = next((t for t in Tenant.get_all() if t.name == tenant_name), None)
            if tenant:
                charges = [c for c in charges if c.tenant_id == tenant.id]
        
        # 租户ID到名称和类型的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 根据统计方式准备数据
        if stat_type == self.get_text('by_type'):
            # 按租户类型分组
            type_data = {}
            for charge in charges:
                # 获取租户对象
                tenant = next((t for t in tenants if t.id == charge.tenant_id), None)
                if tenant:
                    # 处理中文租户类型，转换为对应的翻译键
                    if tenant.type == '办公室':
                        translated_tenant_type = self.get_text('office')
                    elif tenant.type == '门面':
                        translated_tenant_type = self.get_text('storefront')
                    else:
                        translated_tenant_type = self.get_text('unknown_type')
                else:
                    translated_tenant_type = self.get_text('unknown_type')
                
                if translated_tenant_type not in type_data:
                    type_data[translated_tenant_type] = {
                        'water_charge': 0,
                        'electricity_charge': 0
                    }
                type_data[translated_tenant_type]['water_charge'] += charge.water_charge
                type_data[translated_tenant_type]['electricity_charge'] += charge.electricity_charge
            
            # 准备图表数据
            labels = list(type_data.keys())
            water_charges = [data['water_charge'] for data in type_data.values()]
            electricity_charges = [data['electricity_charge'] for data in type_data.values()]
            
            x_label = self.get_text('tenant_type')
        else:  # 按租户统计
            # 准备数据
            labels = []
            water_charges = []
            electricity_charges = []
            
            for charge in charges:
                tenant_name = tenant_map.get(charge.tenant_id, self.get_text('unknown_tenant'))
                labels.append(tenant_name[:10])  # 限制显示长度
                water_charges.append(charge.water_charge)
                electricity_charges.append(charge.electricity_charge)
            
            x_label = self.get_text('tenant_name')
        
        # 设置图表
        ax.clear()
        
        # 绘制柱状图
        x = np.arange(len(labels))
        width = 0.35
        
        ax.bar(x - width/2, water_charges, width, label=self.get_text('water_fee'), color='#3498db')
        ax.bar(x + width/2, electricity_charges, width, label=self.get_text('electricity_fee'), color='#e74c3c')
        
        ax.set_xlabel(x_label)
        ax.set_ylabel(f"{self.get_text('amount')} ({self.get_text('yuan')})")
        ax.set_title(f'{month} {self.get_text("monthly_water_electricity_report")} ({stat_type})')
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=45, ha='right')
        ax.legend()
        
        # 调整布局，防止标签重叠
        plt.tight_layout()
    
    def generate_settlement_chart(self, ax, month, stat_type="按租户"):
        """
        生成结算报表图表（柱状图）
        :param ax: matplotlib轴对象
        :param month: 月份
        :param stat_type: 统计方式（按租户/按类型）
        """
        # 获取结算数据
        settlement = Settlement.get_by_month(month)
        
        # 获取收费数据
        payments = Payment.get_by_month(month)
        total_payment = Payment.get_total_by_month(month)
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 获取所有费用记录，构建完整的费用ID到租户ID的映射
        all_charges = Charge.get_all()
        charge_tenant_map = {c.id: c.tenant_id for c in all_charges}
        
        # 根据统计方式分组收费数据
        stat_data = {}
        for payment in payments:
            # 获取支付记录关联的费用记录
            charge = payment.charge
            if not charge:
                payment.load_charge_info()
                charge = payment.charge
            
            # 直接从费用记录中获取租户ID，确保准确性
            tenant_id = charge.tenant_id if charge else 0
            
            # 获取租户名称和类型
            current_tenant_name = tenant_map.get(tenant_id, self.get_text('unknown_tenant'))
            
            # 获取租户对象
            tenant = next((t for t in tenants if t.id == tenant_id), None)
            if tenant:
                # 处理中文租户类型，转换为对应的翻译键
                if tenant.type == '办公室':
                    current_tenant_type = self.get_text('office')
                elif tenant.type == '门面':
                    current_tenant_type = self.get_text('storefront')
                else:
                    current_tenant_type = self.get_text('unknown_type')
            else:
                current_tenant_type = self.get_text('unknown_type')
            
            # 按统计方式分组
            if stat_type == self.get_text('by_tenant'):
                key = current_tenant_name[:10]  # 限制显示长度
            elif stat_type == self.get_text('by_type'):
                key = current_tenant_type
            else:
                key = current_tenant_name[:10]
            
            # 累加金额
            stat_data[key] = stat_data.get(key, 0) + payment.amount
        
        # 设置图表
        ax.clear()
        
        if stat_data:
            # 如果有统计数据，绘制统计图表
            labels = list(stat_data.keys())
            values = list(stat_data.values())
            
            # 根据数据量选择图表类型
            if len(stat_data) <= 8:
                # 数据较少，使用饼图
                ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90, 
                       textprops={'fontsize': 8})
                ax.axis('equal')
                ax.set_title(f'{month} {self.get_text("water_electricity_settlement_report")} ({stat_type})')
            else:
                # 数据较多，使用柱状图
                ax.bar(labels, values, color='#3498db')
                ax.set_xlabel(self.get_text('statistical_item'))
                ax.set_ylabel(f"{self.get_text('amount')} ({self.get_text('yuan')})")
                ax.set_title(f'{month} {self.get_text("water_electricity_settlement_report")} ({stat_type})')
                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
        else:
            # 没有统计数据，显示简单的已收费用和结算金额比较
            labels = [self.get_text('monthly_total_payment'), self.get_text('suggested_settlement_amount')]
            amounts = [total_payment, total_payment]
            
            if settlement:
                labels = [self.get_text('monthly_total_payment'), self.get_text('total_amount')]
                amounts = [total_payment, settlement.total_amount]
            
            # 绘制柱状图
            ax.bar(labels, amounts, color=['#2ecc71', '#f39c12'])
            
            ax.set_ylabel(f"{self.get_text('amount')} ({self.get_text('yuan')})")
            ax.set_title(f'{month} {self.get_text("water_electricity_settlement_report")}')
        
        # 调整布局
        plt.tight_layout()
    
    def generate_payment_stat_chart(self, ax, month, tenant_name, stat_type):
        """
        生成收费统计图表（饼图或柱状图）
        :param ax: matplotlib轴对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式
        """
        # 获取收费数据
        payments = Payment.get_by_month(month)
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 费用ID到租户ID的映射
        charges = Charge.get_by_month(month)
        charge_tenant_map = {c.id: c.tenant_id for c in charges}
        
        # 统计数据
        stat_data = {}
        total_amount = 0
        
        for payment in payments:
            tenant_id = charge_tenant_map.get(payment.charge_id, 0)
            current_tenant_name = tenant_map.get(tenant_id, self.get_text('unknown_tenant'))
            
            # 如果指定了租户，过滤数据
            if tenant_name and current_tenant_name != tenant_name:
                continue
            
            # 按统计方式分组
            if stat_type == self.get_text('by_tenant'):
                key = current_tenant_name[:10]  # 限制显示长度
            elif stat_type == self.get_text('by_type'):
                # 获取租户对象
                tenant = next((t for t in tenants if t.id == tenant_id), None)
                if tenant:
                    # 处理中文租户类型，转换为对应的翻译键
                    if tenant.type == '办公室':
                        key = self.get_text('office')
                    elif tenant.type == '门面':
                        key = self.get_text('storefront')
                    else:
                        key = self.get_text('unknown_type')
                else:
                    key = self.get_text('unknown_type')
            else:
                key = current_tenant_name[:10]
            
            stat_data[key] = stat_data.get(key, 0) + payment.amount
            total_amount += payment.amount
        
        # 设置图表
        ax.clear()
        
        # 如果数据较少，使用饼图
        if len(stat_data) <= 8:
            # 绘制饼图
            labels = list(stat_data.keys())
            sizes = list(stat_data.values())
            
            ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, 
                   textprops={'fontsize': 8})
            ax.axis('equal')
            ax.set_title(f'{month} {self.get_text("payment_stat_report_title")} ({stat_type})')
        else:
            # 绘制柱状图
            labels = list(stat_data.keys())
            values = list(stat_data.values())
            
            ax.bar(labels, values, color='#3498db')
            ax.set_xlabel(self.get_text('statistical_item'))
            ax.set_ylabel(f"{self.get_text('amount')} ({self.get_text('yuan')})")
            ax.set_title(f'{month} {self.get_text("payment_stat_report_title")} ({stat_type})')
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
        
        # 调整布局
        plt.tight_layout()
    
    def generate_tenant_detail_chart(self, ax, month, tenant_name, stat_type="按租户"):
        """
        生成租户明细报表图表（柱状图）
        :param ax: matplotlib轴对象
        :param month: 月份
        :param tenant_name: 租户名称（可选）
        :param stat_type: 统计方式（按租户/按类型）
        """
        # 获取费用数据
        charges = Charge.get_by_month(month)
        
        # 如果指定了租户，过滤数据
        if tenant_name:
            tenant = next((t for t in Tenant.get_all() if t.name == tenant_name), None)
            if tenant:
                charges = [c for c in charges if c.tenant_id == tenant.id]
        
        # 租户ID到名称的映射
        tenants = Tenant.get_all()
        tenant_map = {t.id: t.name for t in tenants}
        
        # 根据统计方式准备数据
        if stat_type == self.get_text('by_type'):
            # 按租户类型分组
            type_data = {}
            for charge in charges:
                # 获取租户对象
                tenant = next((t for t in tenants if t.id == charge.tenant_id), None)
                if tenant:
                    # 处理中文租户类型，转换为对应的翻译键
                    if tenant.type == '办公室':
                        translated_tenant_type = self.get_text('office')
                    elif tenant.type == '门面':
                        translated_tenant_type = self.get_text('storefront')
                    else:
                        translated_tenant_type = self.get_text('unknown_type')
                else:
                    translated_tenant_type = self.get_text('unknown_type')
                
                if translated_tenant_type not in type_data:
                    type_data[translated_tenant_type] = {
                        'water_usage': 0,
                        'electricity_usage': 0,
                        'water_charge': 0,
                        'electricity_charge': 0
                    }
                type_data[translated_tenant_type]['water_usage'] += charge.water_usage
                type_data[translated_tenant_type]['electricity_usage'] += charge.electricity_usage
                type_data[translated_tenant_type]['water_charge'] += charge.water_charge
                type_data[translated_tenant_type]['electricity_charge'] += charge.electricity_charge
            
            # 准备图表数据
            labels = list(type_data.keys())
            water_usages = [data['water_usage'] for data in type_data.values()]
            electricity_usages = [data['electricity_usage'] for data in type_data.values()]
            water_charges = [data['water_charge'] for data in type_data.values()]
            electricity_charges = [data['electricity_charge'] for data in type_data.values()]
            
            x_label = self.get_text('tenant_type')
        else:  # 按租户统计
            # 准备数据
            labels = []
            water_usages = []
            electricity_usages = []
            water_charges = []
            electricity_charges = []
            
            for charge in charges:
                labels.append(tenant_map.get(charge.tenant_id, self.get_text('unknown_tenant'))[:10])  # 限制显示长度
                water_usages.append(charge.water_usage)
                electricity_usages.append(charge.electricity_usage)
                water_charges.append(charge.water_charge)
                electricity_charges.append(charge.electricity_charge)
            
            x_label = self.get_text('tenant')
        
        # 创建柱状图
        x = range(len(labels))
        width = 0.3
        
        # 绘制用水量和用电量
        ax.bar([i - width*1.5 for i in x], water_usages, width, label=self.get_text('water_usage'), color='blue')
        ax.bar([i - width/2 for i in x], electricity_usages, width, label=self.get_text('electricity_usage'), color='lightblue')
        
        # 创建第二个Y轴用于费用
        ax2 = ax.twinx()
        ax2.bar([i + width/2 for i in x], water_charges, width, label=self.get_text('water_fee'), color='red')
        ax2.bar([i + width*1.5 for i in x], electricity_charges, width, label=self.get_text('electricity_fee'), color='orange')
        
        # 设置图表标题和标签
        ax.set_title(f'{month} {self.get_text("tenant_water_electricity_detail_report")} ({stat_type})')
        ax.set_xlabel(x_label)
        ax.set_ylabel(f"{self.get_text('usage')} ({self.get_text('unit')})")
        ax2.set_ylabel(f"{self.get_text('amount')} ({self.get_text('yuan')})")
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=45, ha='right')
        
        # 合并图例
        lines1, labels1 = ax.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax.legend(lines1 + lines2, labels1 + labels2, loc='upper right')
        
        ax.grid(True, alpha=0.3)
    

    


